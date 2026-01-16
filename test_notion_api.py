import os
from pathlib import Path

base_path = Path('D:/?°ì°¨ê³?)

# ?°ë„ë³??µê³„
stats = {}
total_annual = 0
total_half = 0

for year_folder in base_path.iterdir():
    if year_folder.is_dir():
        year = year_folder.name
        files = list(year_folder.glob('*.xlsx'))
        
        annual_count = len([f for f in files if '?°ì°¨ê³? in f.name])
        half_count = len([f for f in files if 'ë°˜ì°¨ê³? in f.name])
        
        # D?«ì ?Œì‹± (?°ì°¨ ?¼ìˆ˜)
        annual_days = 0
        for f in files:
            if '?°ì°¨ê³? in f.name:
                if '_D1' in f.name or 'ê²°í˜¼' in f.name or 'ê²°í˜¼ê¸°ë…?? in f.name:
                    annual_days += 1
                elif '_D2' in f.name:
                    annual_days += 2
                elif '_D3' in f.name:
                    annual_days += 3
                elif '? í˜¼?¬í–‰' in f.name:
                    # ? í˜¼?¬í–‰?€ ë³´í†µ 5??
                    annual_days += 5
                else:
                    annual_days += 1  # ê¸°ë³¸ê°?
        
        half_days = half_count * 0.5
        
        stats[year] = {
            'annual_count': annual_count,
            'half_count': half_count,
            'annual_days': annual_days,
            'half_days': half_days
        }
        
        total_annual += annual_days
        total_half += half_days

print('=== ?°ë„ë³??°ì°¨ ?¬ìš© ?„í™© ===\n')
for year in sorted(stats.keys()):
    s = stats[year]
    total = s['annual_days'] + s['half_days']
    print(f'{year}: ?°ì°¨ {s["annual_count"]}ê±?({s["annual_days"]}?? + ë°˜ì°¨ {s["half_count"]}ê±?({s["half_days"]}?? = ì´?{total}??)

print(f'\n=== ì´??©ê³„ (2021~2025) ===')
print(f'ì´??°ì°¨ ?¬ìš©: {total_annual}??)
print(f'ì´?ë°˜ì°¨ ?¬ìš©: {total_half}??)
print(f'ì´í•©: {total_annual + total_half}??)

# -*- coding: utf-8 -*-
"""
"ì¡°ë½?? Rakeen Jo" ?˜ì´ì§€ ?ì„¸ ?´ìš© ë¶„ì„
"""

from notion_client import Client

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
TARGET_PAGE_ID = "b920a0b4-9159-47ab-8efc-71b7b5a155bf"
notion = Client(auth=NOTION_TOKEN)

print("=" * 80)
print("?“„ 'ì¡°ë½?? Rakeen Jo' ?˜ì´ì§€ ?ì„¸ ë¶„ì„")
print("=" * 80)
print(f"?˜ì´ì§€ URL: https://notion.so/{TARGET_PAGE_ID.replace('-', '')}")
print("=" * 80)

# ëª¨ë“  ë¸”ë¡ ê°€?¸ì˜¤ê¸?(?˜ì´ì§€?¤ì´???¬í•¨)
all_blocks = []
has_more = True
cursor = None

while has_more:
    if cursor:
        response = notion.blocks.children.list(block_id=TARGET_PAGE_ID, start_cursor=cursor, page_size=100)
    else:
        response = notion.blocks.children.list(block_id=TARGET_PAGE_ID, page_size=100)
    
    all_blocks.extend(response['results'])
    has_more = response.get('has_more', False)
    cursor = response.get('next_cursor')

print(f"\nì´?{len(all_blocks)}ê°œì˜ ë¸”ë¡\n")
print("=" * 80)

def extract_text(rich_text_array):
    """Rich text ë°°ì—´?ì„œ ?ìŠ¤??ì¶”ì¶œ"""
    if not rich_text_array:
        return ""
    return ''.join([t['plain_text'] for t in rich_text_array])

def print_block_content(block, indent=0):
    """ë¸”ë¡ ?´ìš©???½ê¸° ?½ê²Œ ì¶œë ¥"""
    block_type = block['type']
    prefix = "  " * indent
    
    if block_type == 'paragraph':
        text = extract_text(block['paragraph']['rich_text'])
        if text.strip():
            print(f"{prefix}{text}")
            print()
    
    elif block_type.startswith('heading_'):
        level = block_type.split('_')[1]
        text = extract_text(block[block_type]['rich_text'])
        marker = "#" * int(level)
        print(f"\n{prefix}{marker} {text}")
        print()
    
    elif block_type == 'bulleted_list_item':
        text = extract_text(block['bulleted_list_item']['rich_text'])
        print(f"{prefix}??{text}")
    
    elif block_type == 'numbered_list_item':
        text = extract_text(block['numbered_list_item']['rich_text'])
        print(f"{prefix}1. {text}")
    
    elif block_type == 'callout':
        icon = block['callout'].get('icon', {})
        emoji = icon.get('emoji', '?’¡') if icon.get('type') == 'emoji' else '?’¡'
        text = extract_text(block['callout']['rich_text'])
        print(f"\n{prefix}?Œâ? {emoji} CALLOUT ?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€")
        print(f"{prefix}??{text}")
        print(f"{prefix}?”â??€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€")
        print()
    
    elif block_type == 'toggle':
        text = extract_text(block['toggle']['rich_text'])
        print(f"\n{prefix}?”½ [TOGGLE] {text}")
    
    elif block_type == 'code':
        language = block['code']['language']
        code = extract_text(block['code']['rich_text'])
        print(f"\n{prefix}```{language}")
        for line in code.split('\n')[:20]:  # ì²˜ìŒ 20ì¤„ë§Œ
            print(f"{prefix}{line}")
        if len(code.split('\n')) > 20:
            print(f"{prefix}... (?ëµ)")
        print(f"{prefix}```")
        print()
    
    elif block_type == 'divider':
        print(f"\n{prefix}{'?€' * 60}\n")
    
    elif block_type == 'table':
        width = block['table']['table_width']
        print(f"\n{prefix}?“Š [TABLE: {width}??\n")
    
    elif block_type == 'table_row':
        cells = block['table_row']['cells']
        row_text = " | ".join([extract_text(cell) for cell in cells])
        print(f"{prefix}| {row_text} |")
    
    elif block_type == 'quote':
        text = extract_text(block['quote']['rich_text'])
        print(f"{prefix}> {text}")
        print()
    
    elif block_type == 'to_do':
        checked = block['to_do']['checked']
        text = extract_text(block['to_do']['rich_text'])
        checkbox = "?? if checked else "??
        print(f"{prefix}{checkbox} {text}")
    
    else:
        print(f"{prefix}[{block_type.upper()}]")

# ë¸”ë¡ ?´ìš© ì¶œë ¥
for idx, block in enumerate(all_blocks, 1):
    print(f"[ë¸”ë¡ {idx}]")
    print_block_content(block)
    
    # ?ì‹ ë¸”ë¡???ˆëŠ” ê²½ìš° (? ê? ??
    if block.get('has_children'):
        try:
            children = notion.blocks.children.list(block_id=block['id'])
            for child in children['results']:
                print_block_content(child, indent=1)
        except:
            pass

print("\n" + "=" * 80)
print("ë¶„ì„ ?„ë£Œ!")
print("=" * 80)

import pandas as pd
from datetime import datetime, timedelta

# ?‘ì? ?Œì¼ ?½ê¸°
file_path = 'D:/Portfolio_Professional/?°ì°¨?¬ìš©?´ì—­_?•ë¦¬??ë³µì‚¬ë³?xlsx'
df = pd.read_excel(file_path, sheet_name=0)

# NaN ???œê±°
df = df.dropna(subset=['? ì§œ'])

# ? ì§œë¥?datetime?¼ë¡œ ë³€??
df['? ì§œ'] = pd.to_datetime(df['? ì§œ'])

# ?…ì‚¬??
hire_date = datetime(2020, 8, 15)
# ?´ì‚¬??(2026??1??31??
resign_date = datetime(2026, 1, 31)

print("=" * 80)
print("?°ì°¨ ë°œìƒ ë°??¬ìš© ?´ì—­ ë¶„ì„")
print("=" * 80)
print(f"\n?…ì‚¬?? {hire_date.strftime('%Y??%m??%d??)}")
print(f"?´ì‚¬ ?ˆì •?? {resign_date.strftime('%Y??%m??%d??)}")
print(f"ì´?ê·¼ë¬´ ê¸°ê°„: {(resign_date - hire_date).days}??({(resign_date - hire_date).days / 365.25:.1f}??")

# ê·¼ë¡œê¸°ì?ë²•ì— ?°ë¥¸ ?°ì°¨ ë°œìƒ ê·œì¹™:
# 1?„ì°¨: ?…ì‚¬ ??1ê°œì›”ë§ˆë‹¤ 1ê°œì”© ë°œìƒ (ìµœë? 11ê°?, 1??ê·¼ì† ??15ê°?ë°œìƒ
# 2???´ìƒ: ë§¤ë…„ 15ê°?+ 2?„ë§ˆ??1ê°?ì¶”ê? (ìµœë? 25ê°?

annual_leave = {}

# 1?„ì°¨ (2020.08.15 ~ 2021.08.14)
# - 1ê°œì›” ê·¼ì† ?œë§ˆ??1ê°?ë°œìƒ (2020.09.15ë¶€??
# - 1??ê·¼ì† ??15ê°?ë°œìƒ (2021.08.15)
year1_start = hire_date
year1_end = hire_date + timedelta(days=365)
year1_monthly = 11  # 1ê°œì›”~11ê°œì›” (2020.09.15 ~ 2021.07.15)
year1_used = df[(df['? ì§œ'] >= year1_start) & (df['? ì§œ'] < year1_end)]['?¬ìš©?¼ìˆ˜'].sum()

annual_leave['1?„ì°¨ (2020.08.15~2021.08.14)'] = {
    'ë°œìƒ': year1_monthly,
    '?¬ìš©': year1_used,
    'ë¹„ê³ ': '?”ì°¨ 11ê°?(1ê°œì›” ê·¼ì†?œë§ˆ??1ê°?'
}

# 2?„ì°¨ (2021.08.15 ~ 2022.08.14)
# 1??ê·¼ì† ?„ë£Œë¡?15ê°?ë°œìƒ
year2_start = year1_end
year2_end = year2_start + timedelta(days=365)
year2_generated = 15
year2_used = df[(df['? ì§œ'] >= year2_start) & (df['? ì§œ'] < year2_end)]['?¬ìš©?¼ìˆ˜'].sum()

annual_leave['2?„ì°¨ (2021.08.15~2022.08.14)'] = {
    'ë°œìƒ': year2_generated,
    '?¬ìš©': year2_used,
    'ë¹„ê³ ': '1??ê·¼ì† ?„ë£Œ (15ê°?'
}

# 3?„ì°¨ (2022.08.15 ~ 2023.08.14)
year3_start = year2_end
year3_end = year3_start + timedelta(days=365)
year3_generated = 15 + 1  # 2??ê·¼ì† ??+1ê°?
year3_used = df[(df['? ì§œ'] >= year3_start) & (df['? ì§œ'] < year3_end)]['?¬ìš©?¼ìˆ˜'].sum()

annual_leave['3?„ì°¨ (2022.08.15~2023.08.14)'] = {
    'ë°œìƒ': year3_generated,
    '?¬ìš©': year3_used,
    'ë¹„ê³ ': '2??ê·¼ì† (15+1ê°?'
}

# 4?„ì°¨ (2023.08.15 ~ 2024.08.14)
year4_start = year3_end
year4_end = year4_start + timedelta(days=366)  # 2024?„ì? ?¤ë…„
year4_generated = 15 + 2  # 4??ê·¼ì† ??+2ê°?
year4_used = df[(df['? ì§œ'] >= year4_start) & (df['? ì§œ'] < year4_end)]['?¬ìš©?¼ìˆ˜'].sum()

annual_leave['4?„ì°¨ (2023.08.15~2024.08.14)'] = {
    'ë°œìƒ': year4_generated,
    '?¬ìš©': year4_used,
    'ë¹„ê³ ': '4??ê·¼ì† (15+2ê°?'
}

# 5?„ì°¨ (2024.08.15 ~ 2025.08.14)
year5_start = year4_end
year5_end = year5_start + timedelta(days=365)
year5_generated = 15 + 2  # 4??ê·¼ì† ??+2ê°?(5?„ì°¨???™ì¼)
year5_used = df[(df['? ì§œ'] >= year5_start) & (df['? ì§œ'] < year5_end)]['?¬ìš©?¼ìˆ˜'].sum()

annual_leave['5?„ì°¨ (2024.08.15~2025.08.14)'] = {
    'ë°œìƒ': year5_generated,
    '?¬ìš©': year5_used,
    'ë¹„ê³ ': '4??ê·¼ì† (15+2ê°?'
}

# 6?„ì°¨ (2025.08.15 ~ 2026.01.31) - ?´ì‚¬?¼ê¹Œì§€
year6_start = year5_end
year6_end = resign_date
# 2026??1??31?¼ê¹Œì§€ ê·¼ë¬´ ?? 2025??8??15?¼ë?????5.5ê°œì›” ê·¼ë¬´
# 15ê°?+ 3ê°?(6??ê·¼ì†) = 18ê°?ë°œìƒ
# ?˜ì?ë§?ì¤‘ë„ ?´ì‚¬ ??ê·¼ë¬´?¼ìˆ˜??ë¹„ë??˜ì—¬ ë°œìƒ
days_worked_year6 = (year6_end - year6_start).days
year6_generated = 15 + 3  # 6??ê·¼ì† ??+3ê°?
# ì¤‘ë„ ?´ì‚¬ ??ë¹„ë? ê³„ì‚°
year6_prorated = year6_generated * (days_worked_year6 / 365.25)
year6_used = df[(df['? ì§œ'] >= year6_start) & (df['? ì§œ'] <= year6_end)]['?¬ìš©?¼ìˆ˜'].sum()

annual_leave['6?„ì°¨ (2025.08.15~2026.01.31)'] = {
    'ë°œìƒ': year6_generated,
    'ë°œìƒ(ë¹„ë?)': round(year6_prorated, 1),
    '?¬ìš©': year6_used,
    'ë¹„ê³ ': f'6??ê·¼ì† (15+3ê°?, {days_worked_year6}??ê·¼ë¬´ (ë¹„ë? ?ìš©)'
}

print("\n" + "=" * 80)
print("?„ë„ë³??°ì°¨ ë°œìƒ ë°??¬ìš© ?´ì—­")
print("=" * 80)

total_generated = 0
total_used = 0
remaining_by_year = []

for year, info in annual_leave.items():
    print(f"\n{year}")
    print(f"  ë°œìƒ: {info['ë°œìƒ']}ê°?, end='')
    if 'ë°œìƒ(ë¹„ë?)' in info:
        print(f" (ë¹„ë? ?ìš©: {info['ë°œìƒ(ë¹„ë?)']}ê°?")
        generated = info['ë°œìƒ(ë¹„ë?)']
    else:
        print()
        generated = info['ë°œìƒ']
    
    print(f"  ?¬ìš©: {info['?¬ìš©']}ê°?)
    remaining = generated - info['?¬ìš©']
    print(f"  ?”ì—¬: {remaining}ê°?)
    print(f"  ë¹„ê³ : {info['ë¹„ê³ ']}")
    
    total_generated += generated
    total_used += info['?¬ìš©']
    remaining_by_year.append(remaining)

print("\n" + "=" * 80)
print("ì´??©ê³„")
print("=" * 80)
print(f"ì´?ë°œìƒ ?°ì°¨: {total_generated:.1f}ê°?)
print(f"ì´??¬ìš© ?°ì°¨: {total_used:.1f}ê°?)
print(f"ì´??”ì—¬ ?°ì°¨: {total_generated - total_used:.1f}ê°?)

# ?„ì  ?”ì—¬ ê³„ì‚° (?´ì›” ?¬í•¨)
print("\n" + "=" * 80)
print("?°ì°¨ ?˜ë‹¹ ê³„ì‚° (2026??1??31???´ì‚¬ ê¸°ì?)")
print("=" * 80)

# ê·¼ë¡œê¸°ì?ë²? ?´ì‚¬ ??ë¯¸ì‚¬???°ì°¨???€???˜ë‹¹ ì§€ê¸?
# ?? ?°ì°¨???´ì›” ë¶ˆê? ?ì¹™?´ì?ë§? ?¬ìš©?ì˜ ê·€ì±…ì‚¬? ê? ?†ëŠ” ê²½ìš° ?´ì›” ê°€??
# ?¤ì œë¡œëŠ” ?„ë„ë³„ë¡œ ê³„ì‚°?´ì•¼ ??

print("\n[ë°©ë²• 1] ?¨ìˆœ ?©ê³„ ë°©ì‹")
print(f"ë¯¸ì‚¬???°ì°¨ ?˜ë‹¹: {total_generated - total_used:.1f}ê°?)

print("\n[ë°©ë²• 2] ?´ì›” ?†ì´ ?„ë„ë³?ê³„ì‚° (ë²•ì  ?ì¹™)")
carryover = 0
final_remaining = 0

for i, (year, info) in enumerate(annual_leave.items()):
    generated = info.get('ë°œìƒ(ë¹„ë?)', info['ë°œìƒ'])
    used = info['?¬ìš©']
    available = carryover + generated
    remaining = available - used
    
    if i < len(annual_leave) - 1:  # ë§ˆì?ë§??´ê? ?„ë‹Œ ê²½ìš°
        # ?°ì°¨???´ì›”?˜ì? ?ŠìŒ (ë²•ì  ?ì¹™)
        if remaining > 0:
            print(f"{year}: ë°œìƒ {generated}ê°?+ ?´ì›” {carryover}ê°?= ?¬ìš©ê°€??{available}ê°????¬ìš© {used}ê°????Œë©¸ {remaining}ê°?)
            carryover = 0
        else:
            print(f"{year}: ë°œìƒ {generated}ê°?+ ?´ì›” {carryover}ê°?= ?¬ìš©ê°€??{available}ê°????¬ìš© {used}ê°???ë¶€ì¡?{abs(remaining)}ê°?)
            carryover = remaining  # ?Œìˆ˜ë©??¤ìŒ?´ë¡œ ?´ì›”
    else:  # ë§ˆì?ë§???(?´ì‚¬?„ë„)
        final_remaining = remaining
        print(f"{year}: ë°œìƒ {generated:.1f}ê°?+ ?´ì›” {carryover}ê°?= ?¬ìš©ê°€??{available:.1f}ê°????¬ìš© {used}ê°????”ì—¬ {remaining:.1f}ê°?)

print(f"\n?´ì‚¬ ???˜ë ¹ ê°€?¥í•œ ?°ì°¨ ?˜ë‹¹: {max(0, final_remaining):.1f}ê°?)

print("\n" + "=" * 80)
print("?¤ì œ ?Œì‚¬ ?•ì±… ê¸°ì? (?¬ìš©?ê? ?•ë¦¬???°ì´??")
print("=" * 80)
print("24???´ì›”: 3ê°? ë°œìƒ: 16ê°????¬ìš©ê°€?? 19ê°?)
print("25???´ì›”: 3.5ê°? ë°œìƒ: 17ê°????¬ìš©ê°€?? 20.5ê°?)
print("26???´ì›”: 6ê°? ë°œìƒ: 17ê°????¬ìš©ê°€?? 23ê°?)
print(f"\n26??1??31?¼ê¹Œì§€ ?¬ìš© ?ˆìƒ: {df[df['? ì§œ'] <= resign_date]['?¬ìš©?¼ìˆ˜'].sum():.1f}ê°?)
print(f"?”ì—¬ ?°ì°¨ (?˜ë‹¹ ?˜ë ¹ ê°€??: ??23 - ?¬ìš©ë¶?= ì¶”ì • ?„ìš”")

# ?¤ì œ 2026??1?”ê¹Œì§€ ?¬ìš©??
used_2026_jan = df[(df['? ì§œ'] >= datetime(2026, 1, 1)) & (df['? ì§œ'] <= resign_date)]['?¬ìš©?¼ìˆ˜'].sum()
print(f"\n2026??1???¬ìš©: {used_2026_jan}ê°?)
print(f"?ˆìƒ ?”ì—¬: 23 - {used_2026_jan} = {23 - used_2026_jan}ê°?)

import pandas as pd
from datetime import datetime

# ?‘ì? ?Œì¼ ?½ê¸°
file_path = 'D:/Portfolio_Professional/?°ì°¨?¬ìš©?´ì—­_?•ë¦¬??ë³µì‚¬ë³?xlsx'
df = pd.read_excel(file_path, sheet_name=0)

# ?¤ì œ ?°ì°¨ ?¬ìš© ?°ì´?°ë§Œ ?„í„°ë§?(92ë²????´ì „)
df_actual = df.iloc[:92].copy()
df_actual = df_actual.dropna(subset=['? ì§œ'])

# ? ì§œë¥?datetime?¼ë¡œ ë³€??
df_actual['? ì§œ'] = pd.to_datetime(df_actual['? ì§œ'], errors='coerce')

# ?…ì‚¬??
hire_date = datetime(2020, 8, 15)
# ?´ì‚¬??(2026??1??31??
resign_date = datetime(2026, 1, 31)

print("=" * 100)
print("?°ì°¨ ë°œìƒ ë°??¬ìš© ?´ì—­ ë¶„ì„ (?…ì‚¬?? 2020.08.15 ~ ?´ì‚¬ ?ˆì •?? 2026.01.31)")
print("=" * 100)

# ?¬ìš©?ê? ?•ë¦¬???°ì´???½ê¸° (93~94ë²???
user_data = {
    '1?„ì°¨ (2020.08.15~2021.08.14)': 16.5,
    '2?„ì°¨ (2021.08.15~2022.08.14)': 15.5,
    '3?„ì°¨ (2022.08.15~2023.08.14)': 14.0,
    '4?„ì°¨ (2023.08.15~2024.08.14)': 8.5,
    '5?„ì°¨ (2024.08.15~2025.08.14)': 17.0,
    '6?„ì°¨ (2025.08.15~2026.01.31)': 7.5
}

# ?¬ìš©?ê? ?•ë¦¬???Œì‚¬ ?•ì±… ?°ì´??
company_policy = {
    '24??: {'?´ì›”': 3, 'ë°œìƒ': 16, '?¬ìš©ê°€??: 19},
    '25??: {'?´ì›”': 3.5, 'ë°œìƒ': 17, '?¬ìš©ê°€??: 20.5},
    '26??: {'?´ì›”': 6, 'ë°œìƒ': 17, '?¬ìš©ê°€??: 23}
}

print("\n[1] ?¤ì œ ?¬ìš© ?°ì´??(?Œì¼ ê¸°ë°˜)")
print("-" * 100)

# ?„ë„ë³??¬ìš© ?µê³„
periods = [
    ('1?„ì°¨ (2020.08.15~2021.08.14)', datetime(2020, 8, 15), datetime(2021, 8, 14)),
    ('2?„ì°¨ (2021.08.15~2022.08.14)', datetime(2021, 8, 15), datetime(2022, 8, 14)),
    ('3?„ì°¨ (2022.08.15~2023.08.14)', datetime(2022, 8, 15), datetime(2023, 8, 14)),
    ('4?„ì°¨ (2023.08.15~2024.08.14)', datetime(2023, 8, 15), datetime(2024, 8, 14)),
    ('5?„ì°¨ (2024.08.15~2025.08.14)', datetime(2024, 8, 15), datetime(2025, 8, 14)),
    ('6?„ì°¨ (2025.08.15~2026.01.31)', datetime(2025, 8, 15), datetime(2026, 1, 31))
]

for period_name, start_date, end_date in periods:
    used = df_actual[(df_actual['? ì§œ'] >= start_date) & (df_actual['? ì§œ'] <= end_date)]['?¬ìš©?¼ìˆ˜'].sum()
    print(f"{period_name}: ?¬ìš© {used}??)

print("\n[2] ?Œì‚¬ ?•ì±…???°ë¥¸ ?°ì°¨ ë°œìƒ ë°??´ì›”")
print("-" * 100)

for year, data in company_policy.items():
    print(f"{year}: ?´ì›” {data['?´ì›”']}ê°?+ ë°œìƒ {data['ë°œìƒ']}ê°?= ?¬ìš©ê°€??{data['?¬ìš©ê°€??]}ê°?)

print("\n[3] 2026??1??31???´ì‚¬ ???°ì°¨ ?˜ë‹¹ ê³„ì‚°")
print("-" * 100)

# 2024?„ë???ê³„ì‚° (24???´ì›” 3ê°œë????œì‘)
print("\n24?„ë„:")
print(f"  ?´ì›”: 3ê°?)
print(f"  ë°œìƒ: 16ê°?)
print(f"  ?¬ìš©ê°€?? 19ê°?)
used_24 = df_actual[(df_actual['? ì§œ'] >= datetime(2024, 1, 1)) & (df_actual['? ì§œ'] <= datetime(2024, 12, 31))]['?¬ìš©?¼ìˆ˜'].sum()
print(f"  ?¤ì œ ?¬ìš©: {used_24}ê°?)
remaining_24 = 19 - used_24
print(f"  ?”ì—¬ ??25???´ì›”: {remaining_24}ê°?)

print("\n25?„ë„:")
print(f"  ?´ì›”: {remaining_24}ê°?(?¤ì œ ?•ë¦¬ë³¸ì—??3.5ê°?")
print(f"  ë°œìƒ: 17ê°?)
carryover_25 = 3.5  # ?¬ìš©?ê? ?•ë¦¬???´ì›”ë¶??¬ìš©
total_25 = carryover_25 + 17
print(f"  ?¬ìš©ê°€?? {total_25}ê°?)
used_25 = df_actual[(df_actual['? ì§œ'] >= datetime(2025, 1, 1)) & (df_actual['? ì§œ'] <= datetime(2025, 12, 31))]['?¬ìš©?¼ìˆ˜'].sum()
print(f"  ?¤ì œ ?¬ìš©: {used_25}ê°?)
remaining_25 = total_25 - used_25
print(f"  ?”ì—¬ ??26???´ì›”: {remaining_25}ê°?)

print("\n26?„ë„ (1??31???´ì‚¬):")
print(f"  ?´ì›”: {remaining_25}ê°?(?¤ì œ ?•ë¦¬ë³¸ì—??6ê°?")
carryover_26 = 6  # ?¬ìš©?ê? ?•ë¦¬???´ì›”ë¶??¬ìš©
print(f"  ë°œìƒ: 17ê°?(?°ì´ˆ ê¸°ì?)")
total_26 = carryover_26 + 17
print(f"  ?¬ìš©ê°€?? {total_26}ê°?)
used_26 = df_actual[(df_actual['? ì§œ'] >= datetime(2026, 1, 1)) & (df_actual['? ì§œ'] <= datetime(2026, 1, 31))]['?¬ìš©?¼ìˆ˜'].sum()
print(f"  1??31?¼ê¹Œì§€ ?¬ìš©: {used_26}ê°?)
remaining_26 = total_26 - used_26
print(f"  ?”ì—¬: {remaining_26}ê°?)

print("\n" + "=" * 100)
print("?“Š ìµœì¢… ê²°ê³¼")
print("=" * 100)
print(f"\n2026??1??31???´ì‚¬ ??ë°›ì„ ???ˆëŠ” ?°ì°¨ ?˜ë‹¹: {remaining_26}ê°?)
print(f"\n???Œì‚¬ ?•ì±…???°ë¥¸ ?´ì›” ë°?ë°œìƒ ê¸°ì? ?ìš©")
print(f"??2026??1?”ê¹Œì§€ ?¤ì œ ?¬ìš©: {used_26}ê°?)

# ?ì„¸ ?´ì—­
print("\n[?ì„¸ ?´ì—­]")
print(f"  2024???¬ìš©: {used_24}ê°?)
print(f"  2025???¬ìš©: {used_25}ê°?)
print(f"  2026??1???¬ìš©: {used_26}ê°?)
print(f"  ì´??¬ìš© (2024~2026.01): {used_24 + used_25 + used_26}ê°?)

print("\n" + "=" * 100)

import pandas as pd
from datetime import datetime, timedelta

# ?‘ì? ?Œì¼ ?½ê¸°
file_path = 'D:/Portfolio_Professional/?°ì°¨?¬ìš©?´ì—­_?•ë¦¬??ë³µì‚¬ë³?xlsx'
df = pd.read_excel(file_path, sheet_name=0)

# ?¤ì œ ?°ì°¨ ?¬ìš© ?°ì´?°ë§Œ ?„í„°ë§?
df_actual = df.iloc[:92].copy()
df_actual = df_actual.dropna(subset=['? ì§œ'])
df_actual['? ì§œ'] = pd.to_datetime(df_actual['? ì§œ'], errors='coerce')

# ?…ì‚¬??ë°??´ì‚¬??
hire_date = datetime(2020, 8, 15)
resign_date = datetime(2026, 1, 31)

print("=" * 100)
print("?´ì‚¬ ???°ì°¨ ?˜ë‹¹ ê³„ì‚° (?…ì‚¬??ê¸°ì?)")
print("=" * 100)
print(f"\n?…ì‚¬?? {hire_date.strftime('%Y??%m??%d??)}")
print(f"?´ì‚¬?? {resign_date.strftime('%Y??%m??%d??)}")
print(f"ì´?ê·¼ë¬´ ê¸°ê°„: {(resign_date - hire_date).days}??({(resign_date - hire_date).days / 365.25:.2f}??")

print("\n" + "=" * 100)
print("ê·¼ë¡œê¸°ì?ë²•ì— ?°ë¥¸ ?…ì‚¬??ê¸°ì? ?°ì°¨ ë°œìƒ")
print("=" * 100)

# ê·¼ë¡œê¸°ì?ë²?
# - 1??ë¯¸ë§Œ: ë§¤ì›” 1ê°?ë°œìƒ (ìµœë? 11ê°?
# - 1???´ìƒ: 15ê°?+ 2?„ë§ˆ??1ê°?ì¶”ê? (ìµœë? 25ê°?
#   * 3??ë¯¸ë§Œ: 15ê°?
#   * 3~5??ë¯¸ë§Œ: 16ê°?
#   * 5~7??ë¯¸ë§Œ: 17ê°?
#   * 7~9??ë¯¸ë§Œ: 18ê°?

periods = []

# 1?„ì°¨ (2020.08.15 ~ 2021.08.14) - ?”ì°¨ 11ê°?
period1_start = hire_date
period1_end = hire_date + timedelta(days=365)
period1_generated = 11  # 1ê°œì›” ~ 11ê°œì›” (1??ë¯¸ë§Œ)
period1_used = df_actual[(df_actual['? ì§œ'] >= period1_start) & (df_actual['? ì§œ'] < period1_end)]['?¬ìš©?¼ìˆ˜'].sum()
periods.append({
    'ê¸°ê°„': '1?„ì°¨ (2020.08.15~2021.08.14)',
    'ë°œìƒ': period1_generated,
    '?¬ìš©': period1_used,
    '?”ì—¬': period1_generated - period1_used,
    'ë¹„ê³ ': '?…ì‚¬ ??1??ë¯¸ë§Œ (?”ì°¨ 11ê°?'
})

# 2?„ì°¨ (2021.08.15 ~ 2022.08.14) - ?°ì°¨ 15ê°?
period2_start = period1_end
period2_end = period2_start + timedelta(days=365)
period2_generated = 15  # 1??ê·¼ì†
period2_used = df_actual[(df_actual['? ì§œ'] >= period2_start) & (df_actual['? ì§œ'] < period2_end)]['?¬ìš©?¼ìˆ˜'].sum()
periods.append({
    'ê¸°ê°„': '2?„ì°¨ (2021.08.15~2022.08.14)',
    'ë°œìƒ': period2_generated,
    '?¬ìš©': period2_used,
    '?”ì—¬': period2_generated - period2_used,
    'ë¹„ê³ ': '1??ê·¼ì† (15ê°?'
})

# 3?„ì°¨ (2022.08.15 ~ 2023.08.14) - ?°ì°¨ 16ê°?
period3_start = period2_end
period3_end = period3_start + timedelta(days=365)
period3_generated = 16  # 3??ê·¼ì† (15 + 1)
period3_used = df_actual[(df_actual['? ì§œ'] >= period3_start) & (df_actual['? ì§œ'] < period3_end)]['?¬ìš©?¼ìˆ˜'].sum()
periods.append({
    'ê¸°ê°„': '3?„ì°¨ (2022.08.15~2023.08.14)',
    'ë°œìƒ': period3_generated,
    '?¬ìš©': period3_used,
    '?”ì—¬': period3_generated - period3_used,
    'ë¹„ê³ ': '3??ê·¼ì† (15+1ê°?'
})

# 4?„ì°¨ (2023.08.15 ~ 2024.08.14) - ?°ì°¨ 16ê°?
period4_start = period3_end
period4_end = period4_start + timedelta(days=366)  # 2024???¤ë…„
period4_generated = 16  # 4??ê·¼ì† (15 + 1, 5??ë¯¸ë§Œ?´ë?ë¡??„ì§ 16ê°?
period4_used = df_actual[(df_actual['? ì§œ'] >= period4_start) & (df_actual['? ì§œ'] < period4_end)]['?¬ìš©?¼ìˆ˜'].sum()
periods.append({
    'ê¸°ê°„': '4?„ì°¨ (2023.08.15~2024.08.14)',
    'ë°œìƒ': period4_generated,
    '?¬ìš©': period4_used,
    '?”ì—¬': period4_generated - period4_used,
    'ë¹„ê³ ': '4??ê·¼ì† (15+1ê°?'
})

# 5?„ì°¨ (2024.08.15 ~ 2025.08.14) - ?°ì°¨ 17ê°?
period5_start = period4_end
period5_end = period5_start + timedelta(days=365)
period5_generated = 17  # 5??ê·¼ì† (15 + 2)
period5_used = df_actual[(df_actual['? ì§œ'] >= period5_start) & (df_actual['? ì§œ'] < period5_end)]['?¬ìš©?¼ìˆ˜'].sum()
periods.append({
    'ê¸°ê°„': '5?„ì°¨ (2024.08.15~2025.08.14)',
    'ë°œìƒ': period5_generated,
    '?¬ìš©': period5_used,
    '?”ì—¬': period5_generated - period5_used,
    'ë¹„ê³ ': '5??ê·¼ì† (15+2ê°?'
})

# 6?„ì°¨ (2025.08.15 ~ 2026.01.31) - ì¤‘ë„ ?´ì‚¬
period6_start = period5_end
period6_end = resign_date
days_worked = (period6_end - period6_start).days + 1  # 169??
period6_base = 17  # 6??ê·¼ì† ê¸°ì? (15 + 2, 7??ë¯¸ë§Œ?´ë?ë¡?17ê°?
# ì¤‘ë„ ?´ì‚¬ ??ë¹„ë? ê³„ì‚°
period6_generated = period6_base * (days_worked / 365)
period6_used = df_actual[(df_actual['? ì§œ'] >= period6_start) & (df_actual['? ì§œ'] <= period6_end)]['?¬ìš©?¼ìˆ˜'].sum()
periods.append({
    'ê¸°ê°„': '6?„ì°¨ (2025.08.15~2026.01.31)',
    'ë°œìƒ': round(period6_generated, 1),
    'ë°œìƒ(ê¸°ì?)': period6_base,
    '?¬ìš©': period6_used,
    '?”ì—¬': round(period6_generated - period6_used, 1),
    'ë¹„ê³ ': f'6??ê·¼ì† (15+2ê°?, {days_worked}??ê·¼ë¬´ ë¹„ë? ?ìš©'
})

print("\n[?„ë„ë³??°ì°¨ ë°œìƒ ë°??¬ìš© ?´ì—­]")
print("-" * 100)

total_generated = 0
total_used = 0
total_remaining = 0

for i, p in enumerate(periods, 1):
    print(f"\n{i}. {p['ê¸°ê°„']}")
    if 'ë°œìƒ(ê¸°ì?)' in p:
        print(f"   ë°œìƒ: {p['ë°œìƒ']}ê°?(ê¸°ì?: {p['ë°œìƒ(ê¸°ì?)']}ê°?")
    else:
        print(f"   ë°œìƒ: {p['ë°œìƒ']}ê°?)
    print(f"   ?¬ìš©: {p['?¬ìš©']}ê°?)
    print(f"   ?”ì—¬: {p['?”ì—¬']}ê°?)
    print(f"   ë¹„ê³ : {p['ë¹„ê³ ']}")
    
    total_generated += p['ë°œìƒ']
    total_used += p['?¬ìš©']
    total_remaining += p['?”ì—¬']

print("\n" + "=" * 100)
print("ì´??©ê³„")
print("=" * 100)
print(f"ì´?ë°œìƒ ?°ì°¨: {total_generated:.1f}ê°?)
print(f"ì´??¬ìš© ?°ì°¨: {total_used:.1f}ê°?)
print(f"ì´??”ì—¬ ?°ì°¨: {total_remaining:.1f}ê°?)

print("\n" + "=" * 100)
print("?“Š ?´ì‚¬ ???°ì°¨ ?˜ë‹¹ (2026??1??31??ê¸°ì?)")
print("=" * 100)
print(f"\n?…ì‚¬??ê¸°ì? ê³„ì‚°:")
print(f"  - 1?„ì°¨ (?”ì°¨): 11ê°?ë°œìƒ ??{period1_used}ê°??¬ìš© ??{periods[0]['?”ì—¬']}ê°??Œë©¸")
print(f"  - 2?„ì°¨: 15ê°?ë°œìƒ ??{period2_used}ê°??¬ìš© ??{periods[1]['?”ì—¬']}ê°??Œë©¸")
print(f"  - 3?„ì°¨: 16ê°?ë°œìƒ ??{period3_used}ê°??¬ìš© ??{periods[2]['?”ì—¬']}ê°??Œë©¸")
print(f"  - 4?„ì°¨: 16ê°?ë°œìƒ ??{period4_used}ê°??¬ìš© ??{periods[3]['?”ì—¬']}ê°??Œë©¸")
print(f"  - 5?„ì°¨: 17ê°?ë°œìƒ ??{period5_used}ê°??¬ìš© ??{periods[4]['?”ì—¬']}ê°??”ì—¬")
print(f"  - 6?„ì°¨: {period6_generated:.1f}ê°?ë°œìƒ(ë¹„ë?) ??{period6_used}ê°??¬ìš© ??{periods[5]['?”ì—¬']}ê°??”ì—¬")

print(f"\n??ë²•ì ?¼ë¡œ??1~4?„ì°¨ ë¯¸ì‚¬???°ì°¨???Œë©¸??(1??ê²½ê³¼ ??")
print(f"???´ì‚¬ ?œì  ê¸°ì? ? íš¨ ?°ì°¨:")
print(f"  - 5?„ì°¨ ?”ì—¬: {periods[4]['?”ì—¬']}ê°?)
print(f"  - 6?„ì°¨ ?”ì—¬: {periods[5]['?”ì—¬']}ê°?)

valid_remaining = periods[4]['?”ì—¬'] + periods[5]['?”ì—¬']

print(f"\n?¯ ?´ì‚¬ ???˜ë ¹ ê°€?¥í•œ ?°ì°¨ ?˜ë‹¹: {valid_remaining:.1f}ê°?)

# ?ì„¸ ?¬ìš© ?´ì—­
print("\n[ì°¸ê³ : ?„ë„ë³??ì„¸ ?¬ìš© ?´ì—­]")
print("-" * 100)
for year in [2020, 2021, 2022, 2023, 2024, 2025, 2026]:
    year_data = df_actual[df_actual['? ì§œ'].dt.year == year]
    if len(year_data) > 0:
        year_total = year_data['?¬ìš©?¼ìˆ˜'].sum()
        print(f"{year}?? {year_total}ê°??¬ìš©")

print("\n" + "=" * 100)

import pandas as pd
from datetime import datetime, timedelta

# ?‘ì? ?Œì¼ ?½ê¸°
file_path = 'D:/Portfolio_Professional/?°ì°¨?¬ìš©?´ì—­_?•ë¦¬??ë³µì‚¬ë³?xlsx'
df = pd.read_excel(file_path, sheet_name=0)

# ?¤ì œ ?°ì°¨ ?¬ìš© ?°ì´?°ë§Œ ?„í„°ë§?
df_actual = df.iloc[:92].copy()
df_actual = df_actual.dropna(subset=['? ì§œ'])
df_actual['? ì§œ'] = pd.to_datetime(df_actual['? ì§œ'], errors='coerce')

# ?…ì‚¬??ë°??´ì‚¬??
hire_date = datetime(2020, 8, 15)
resign_date = datetime(2026, 1, 31)

print("=" * 100)
print("?´ì‚¬ ???°ì°¨ ?˜ë‹¹ ê³„ì‚° (?…ì‚¬??ê¸°ì? + ?´ì›” ?ìš©)")
print("=" * 100)
print(f"\n?…ì‚¬?? {hire_date.strftime('%Y??%m??%d??)}")
print(f"?´ì‚¬?? {resign_date.strftime('%Y??%m??%d??)}")
print(f"ì´?ê·¼ë¬´ ê¸°ê°„: {(resign_date - hire_date).days}??({(resign_date - hire_date).days / 365.25:.2f}??")

print("\n" + "=" * 100)
print("ê·¼ë¡œê¸°ì?ë²•ì— ?°ë¥¸ ?…ì‚¬??ê¸°ì? ?°ì°¨ ë°œìƒ (?´ì›” ?¬í•¨)")
print("=" * 100)

# ê°?ê¸°ê°„ë³??°ì´??
periods_data = [
    ('1?„ì°¨ (2020.08.15~2021.08.14)', datetime(2020, 8, 15), datetime(2021, 8, 14), 11, '?”ì°¨ 11ê°?(1??ë¯¸ë§Œ)'),
    ('2?„ì°¨ (2021.08.15~2022.08.14)', datetime(2021, 8, 15), datetime(2022, 8, 14), 15, '1??ê·¼ì† (15ê°?'),
    ('3?„ì°¨ (2022.08.15~2023.08.14)', datetime(2022, 8, 15), datetime(2023, 8, 14), 16, '3??ê·¼ì† (15+1ê°?'),
    ('4?„ì°¨ (2023.08.15~2024.08.14)', datetime(2023, 8, 15), datetime(2024, 8, 15), 16, '4??ê·¼ì† (15+1ê°?'),
    ('5?„ì°¨ (2024.08.15~2025.08.14)', datetime(2024, 8, 15), datetime(2025, 8, 14), 17, '5??ê·¼ì† (15+2ê°?'),
]

# 6?„ì°¨??ë¹„ë? ê³„ì‚°
period6_start = datetime(2025, 8, 15)
period6_end = resign_date
days_worked = (period6_end - period6_start).days + 1
period6_base = 17
period6_generated = round(period6_base * (days_worked / 365), 1)
periods_data.append(
    ('6?„ì°¨ (2025.08.15~2026.01.31)', period6_start, period6_end, period6_generated, 
     f'6??ê·¼ì† (15+2ê°?, {days_worked}??ë¹„ë?')
)

print("\n[?„ë„ë³??°ì°¨ ë°œìƒ ë°??¬ìš© ?´ì—­ (?´ì›” ?¬í•¨)]")
print("-" * 100)

carryover = 0  # ?´ì›” ?°ì°¨
total_generated = 0
total_used = 0

for i, (period_name, start_date, end_date, generated, note) in enumerate(periods_data, 1):
    # ?´ë‹¹ ê¸°ê°„ ?¬ìš©??
    used = df_actual[(df_actual['? ì§œ'] >= start_date) & (df_actual['? ì§œ'] <= end_date)]['?¬ìš©?¼ìˆ˜'].sum()
    
    # ?¬ìš© ê°€??= ?´ì›” + ë°œìƒ
    available = carryover + generated
    
    # ?”ì—¬ = ?¬ìš©ê°€??- ?¬ìš©
    remaining = available - used
    
    print(f"\n{i}. {period_name}")
    print(f"   ?´ì›”: {carryover}ê°?)
    print(f"   ë°œìƒ: {generated}ê°?)
    print(f"   ?¬ìš©ê°€?? {available}ê°?)
    print(f"   ?¬ìš©: {used}ê°?)
    print(f"   ?”ì—¬: {remaining}ê°?)
    print(f"   ë¹„ê³ : {note}")
    
    # ?¤ìŒ ?„ë„ë¡??´ì›” (?”ì—¬ê°€ ?Œìˆ˜ë©?0?¼ë¡œ, ë²•ì ?¼ë¡œ??ë§ˆì´?ˆìŠ¤ ì²˜ë¦¬ ?ˆë¨)
    if i < len(periods_data):  # ë§ˆì?ë§??„ë„ê°€ ?„ë‹ˆë©??´ì›”
        carryover = max(0, remaining)  # ?Œìˆ˜??0?¼ë¡œ (ë²•ì ?¼ë¡œ ? ì°¨ê°?ë¶ˆê?)
    
    total_generated += generated
    total_used += used

final_remaining = remaining

print("\n" + "=" * 100)
print("ì´??©ê³„")
print("=" * 100)
print(f"ì´?ë°œìƒ ?°ì°¨: {total_generated}ê°?)
print(f"ì´??¬ìš© ?°ì°¨: {total_used}ê°?)

print("\n" + "=" * 100)
print("?“Š ?´ì‚¬ ???°ì°¨ ?˜ë‹¹ (2026??1??31??ê¸°ì?)")
print("=" * 100)

print(f"\n???„ì  ?´ì›”??ê³ ë ¤??ìµœì¢… ?”ì—¬:")
print(f"\n?¯ ?´ì‚¬ ???˜ë ¹ ê°€?¥í•œ ?°ì°¨ ?˜ë‹¹: {max(0, final_remaining)}ê°?)

if final_remaining < 0:
    print(f"\n? ï¸  ì£¼ì˜: ?”ì—¬ê°€ ?Œìˆ˜({final_remaining}ê°?ë¡??˜í??¬ìŠµ?ˆë‹¤.")
    print(f"   ?´ëŠ” ë°œìƒ???°ì°¨ë³´ë‹¤ ??ë§ì´ ?¬ìš©?ˆë‹¤???˜ë??…ë‹ˆ??")
    print(f"   - ?Œì‚¬ ?•ì±…?¼ë¡œ ? ì°¨ê°?ë¯¸ë˜ ?°ì°¨ ë¯¸ë¦¬ ?¬ìš©)???ˆìš©??ê²½ìš°")
    print(f"   - ?ëŠ” ?Œì‚¬ê°€ ì¶”ê? ?°ì°¨ë¥?ë¶€?¬í•œ ê²½ìš°")
    print(f"   ?¤ì œ ?˜ë‹¹?€ 0ê°œì…?ˆë‹¤.")

print("\n[ì°¸ê³ : ?´ì›” ?ë¦„??")
print("-" * 100)

# ?¤ì‹œ ê³„ì‚°?˜ì—¬ ?ë¦„ ë³´ì—¬ì£¼ê¸°
carryover_flow = 0
for i, (period_name, start_date, end_date, generated, note) in enumerate(periods_data, 1):
    used = df_actual[(df_actual['? ì§œ'] >= start_date) & (df_actual['? ì§œ'] <= end_date)]['?¬ìš©?¼ìˆ˜'].sum()
    available = carryover_flow + generated
    remaining = available - used
    
    print(f"{i}?„ì°¨: ?´ì›”{carryover_flow} + ë°œìƒ{generated} = {available} ???¬ìš©{used} = ?”ì—¬{remaining}", end='')
    if i < len(periods_data):
        carryover_flow = max(0, remaining)
        print(f" ???¤ìŒ ?´ì›”: {carryover_flow}")
    else:
        print(f" ??ìµœì¢… ?”ì—¬")

print("\n" + "=" * 100)

# ?Œì‚¬ ?•ì±…ê³?ë¹„êµ
print("\n[ì°¸ê³ : ?Œì‚¬ê°€ ê³µì????°ì´?°ì? ë¹„êµ]")
print("-" * 100)
print("?Œì‚¬ ê³µì? (?Œê³„?„ë„ ê¸°ì?):")
print("  24?? ?´ì›” 3 + ë°œìƒ 16 = 19ê°?)
print("  25?? ?´ì›” 3.5 + ë°œìƒ 17 = 20.5ê°?)
print("  26?? ?´ì›” 6 + ë°œìƒ 17 = 23ê°?)
print("\n?…ì‚¬??ê¸°ì? ê³„ì‚°ê³??Œê³„?„ë„ ê¸°ì? ê³„ì‚°??ì°¨ì´ë¡??¸í•´")
print("?¤ì œ ?´ì‚¬ ???˜ë‹¹?€ ?Œì‚¬?€ ?‘ì˜ê°€ ?„ìš”?????ˆìŠµ?ˆë‹¤.")

print("\n" + "=" * 100)

# -*- coding: utf-8 -*-
from notion_client import Client
import json

notion = Client(auth='YOUR_NOTION_TOKEN_HERE')
parent_id = '2e8f0746-9821-809c-9331-e34ae2d5e03e'

print("=== ?¸ì…˜ ?˜ì´ì§€ ëª©ë¡ ===\n")
results = notion.blocks.children.list(block_id=parent_id)
for block in results['results']:
    if block['type'] == 'child_page':
        title = block['child_page']['title']
        page_id = block['id']
        print(f"{title}: {page_id}")

# -*- coding: utf-8 -*-
from notion_client import Client

notion = Client(auth='YOUR_NOTION_TOKEN_HERE')
v5_page_id = '2e8f0746-9821-818f-82a0-d0421f8d53ea'

def extract_text(rich_text_array):
    return ''.join([text['plain_text'] for text in rich_text_array])

print("=== ?¸ì…˜ v5?€ HTML ë¹„êµ ë¶„ì„ ===\n")

# ?¸ì…˜ v5 ?˜ì´ì§€??ëª¨ë“  ? ê? ë¸”ë¡ ê°€?¸ì˜¤ê¸?
blocks = notion.blocks.children.list(block_id=v5_page_id, page_size=100)

projects = {}
current_project = None

for block in blocks['results']:
    if block['type'] == 'toggle':
        # ? ê? ?œëª© ì¶”ì¶œ
        title = extract_text(block['toggle']['rich_text'])
        current_project = title
        projects[current_project] = {'title': title, 'content': []}
        
        # ? ê? ?ì‹ ë¸”ë¡ ê°€?¸ì˜¤ê¸?
        if block['has_children']:
            children = notion.blocks.children.list(block_id=block['id'], page_size=100)
            
            for child in children['results']:
                if child['type'] == 'bulleted_list_item':
                    text = extract_text(child['bulleted_list_item']['rich_text'])
                    projects[current_project]['content'].append(text)
                elif child['type'] == 'heading_3':
                    text = extract_text(child['heading_3']['rich_text'])
                    projects[current_project]['content'].append(f"### {text}")

# ?„ë¡œ?íŠ¸ë³?ì¶œë ¥
print("?“Š ?¸ì…˜ v5 ?„ë¡œ?íŠ¸ ëª©ë¡:\n")
for idx, (key, proj) in enumerate(projects.items(), 1):
    print(f"{idx}. {proj['title']}")
    print(f"   ?µì‹¬ ê¸°ìˆ :")
    for item in proj['content']:
        if '?µì‹¬ ê¸°ìˆ ' in item or not item.startswith('###'):
            print(f"   - {item}")
    print()

print("\n=== ì£¼ìš” ì°¨ì´??ë¶„ì„ ===\n")

# Psi-1000 ë¶„ì„
print("?”¬ Psi-1000:")
print("?¸ì…˜ v5: ì§„ê³µ ê²Œì´ì§€ ëª¨ë‹ˆ?°ë§ + 250ms ì£¼ê¸° PID")
print("HTML: Honeywell ?•ë ¥?¼ì„œ + BLDC ëª¨í„°\n")

# L-Titrator ë¶„ì„  
print("?§ª L-Titrator:")
print("?¸ì…˜ v5: Hamilton ?œë¦°ì§€ ?Œí”„ (RS485)")
print("HTML: DRV8825 ?¤í…Œ??ëª¨í„° (GPIO)\n")

"""
Á¶¶ôÇö Æ÷Æ®Æú¸®¿À ³ë¼Ç »ı¼º ½ºÅ©¸³Æ®
"""
from notion_client import Client

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"

notion = Client(auth=NOTION_TOKEN)

print("=" * 60)
print("?? Á¶¶ôÇö Æ÷Æ®Æú¸®¿À ³ë¼Ç ÀÚµ¿ »ı¼º")
print("=" * 60)

# 1. ¸ŞÀÎ ÆäÀÌÁö »ı¼º
print("\n?? ¸ŞÀÎ ÆäÀÌÁö »ı¼º Áß...")

page = notion.pages.create(
    parent={"page_id": PARENT_ID},
    icon={"type": "emoji", "emoji": "?????"},
    cover={
        "type": "external",
        "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1200"}
    },
    properties={
        "title": {"title": [{"text": {"content": "Á¶¶ôÇö | Senior Embedded Systems Engineer (9³â)"}}]}
    },
    children=[
        # Çì´õ Callout
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "??"},
                "color": "blue_background",
                "rich_text": [{"type": "text", "text": {"content": "9³â °æ·Â | Hardware & Firmware Full-Stack °³¹ß | STM32 Expert | FreeRTOS | PID Control"}}]
            }
        },
        # ¿¬¶ôÃ³
        {
            "object": "block",
            "type": "paragraph",
            "paragraph": {
                "rich_text": [
                    {"type": "text", "text": {"content": "?? 92lock@kakao.com  |  ?? 010-7311-0402  |  "}, "annotations": {"color": "gray"}},
                    {"type": "text", "text": {"content": "?? GitHub", "link": {"url": "https://github.com/gari210404"}}, "annotations": {"color": "blue"}}
                ]
            }
        },
        {"object": "block", "type": "divider", "divider": {}},
    ]
)

page_id = page["id"]
print(f"  ? ¸ŞÀÎ ÆäÀÌÁö »ı¼º: {page_id}")

# 2. Executive Summary
print("\n?? Executive Summary Ãß°¡...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Executive Summary"}}]}},
        {
            "object": "block",
            "type": "quote",
            "quote": {
                "rich_text": [{"type": "text", "text": {"content": "9³â °æ·ÂÀÇ ½Ã´Ï¾î ÀÓº£µğµå ½Ã½ºÅÛ ¿£Áö´Ï¾î·Î, ÇÏµå¿ş¾î ¼³°èºÎÅÍ Æß¿ş¾î °³¹ß±îÁö Full-Stack Embedded °³¹ß Àü¹®°¡ÀÔ´Ï´Ù. ¹İµµÃ¼ °øÁ¤ Àåºñ, ±¤ÇĞ ºĞ¼® ±â±â, »ê¾÷¿ë Á¦¾î ½Ã½ºÅÛÀÇ ¼³°è ¹× °³¹ß¿¡¼­ °ËÁõµÈ ½ÇÀûÀ» º¸À¯ÇÏ°í ÀÖ½À´Ï´Ù."}}]
            }
        },
        {"object": "block", "type": "divider", "divider": {}},
    ]
)

# 3. Core Competencies
print("?? Core Competencies Ãß°¡...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Core Competencies"}}]}},
        {
            "object": "block",
            "type": "table",
            "table": {
                "table_width": 3,
                "has_column_header": True,
                "has_row_header": False,
                "children": [
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "ºĞ¾ß"}}], [{"type": "text", "text": {"content": "±â¼ú"}}], [{"type": "text", "text": {"content": "°æ·Â"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "?? Firmware"}}], [{"type": "text", "text": {"content": "STM32, FreeRTOS, HAL/LL, Bootloader"}}], [{"type": "text", "text": {"content": "9³â"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "? Hardware"}}], [{"type": "text", "text": {"content": "Analog/Digital Circuit, PCB Layout"}}], [{"type": "text", "text": {"content": "9³â"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "?? Communication"}}], [{"type": "text", "text": {"content": "Modbus RTU/ASCII/TCP, UART, SPI, I2C, CAN"}}], [{"type": "text", "text": {"content": "8³â"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "??? Control"}}], [{"type": "text", "text": {"content": "PID, Auto-tuning, State Machine, Kalman"}}], [{"type": "text", "text": {"content": "6³â"}}]]}},
                ]
            }
        },
        {"object": "block", "type": "divider", "divider": {}},
    ]
)

# 4. Projects
print("?? Major Projects Ãß°¡...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Major Projects"}}]}},
    ]
)

# Project 1: L-LPC
toggle1 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "1?? L-LPC (H-Sensor) System | 2022-2024 | ? ´ëÇ¥ ÇÁ·ÎÁ§Æ®"}, "annotations": {"bold": True}}],
                "color": "blue_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "¹İµµÃ¼ Àú¾Ğ Ã¨¹ö Á¦¾î ½Ã½ºÅÛ - Hardware & Firmware Full-Stack °³¹ß"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "?? ÇÏµå¿ş¾î ¼³°è:"}, "annotations": {"bold": True, "color": "blue"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Main Board: STM32H743 (480MHz, 2MB Flash, 1MB RAM)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Analog Frontend: °íÁ¤¹Ğ TIA (S/N > 60dB)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "FPGA Interface: Xilinx Artix-7 ¿¬µ¿"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Power: Multi-rail SMPS (È¿À² > 90%, Ripple < 30mVpp)"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "?? Æß¿ş¾î °³¹ß:"}, "annotations": {"bold": True, "color": "purple"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "FreeRTOS ±â¹İ ¸ÖÆ¼ÅÂ½ºÅ© ±¸Á¶ (7°³ Task)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "DMA ±â¹İ °í¼Ó ADC µ¥ÀÌÅÍ ¼öÁı (2MSPS)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "TCP/IP (lwIP), Modbus RTU Åë½Å ±¸Çö"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "?? µğ¹ö±ë °æÇè:"}, "annotations": {"bold": True, "color": "red"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ADC ³ëÀÌÁî ÀÌ½´: ¡¾500mV ¡æ ¡¾5mV (100¹è °³¼±)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "FreeRTOS µ¥µå¶ô ÇØ°á: Mutex È¹µæ ¼ø¼­ ÅëÀÏ"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "callout", "callout": {"icon": {"type": "emoji", "emoji": "??"}, "color": "green_background", "rich_text": [{"type": "text", "text": {"content": "¼º°ú: ¾ç»ê 100+ units, ÇÊµå ºÒ·®·ü < 0.2%"}}]}},
    ]
)

# Project 2: Psi-1000
toggle2 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "2?? Psi-1000/3000 Pressure Controller | 2019-2023 | ? PID Àü¹®"}, "annotations": {"bold": True}}],
                "color": "purple_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle2["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "Á¤¹Ğ Áø°ø ¾Ğ·Â Á¦¾î ½Ã½ºÅÛ - PID ¾Ë°í¸®Áò °³¹ß ¹× ÃÖÀûÈ­"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "??? PID Á¦¾î ¾Ë°í¸®Áò:"}, "annotations": {"bold": True, "color": "purple"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Anti-windup, Derivative filtering ±¸Çö"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Auto-tuning (Relay Feedback, Z-N Method)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Cascade control (Pressure + Temperature)"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "?? Åë½Å ÇÁ·ÎÅäÄİ:"}, "annotations": {"bold": True, "color": "blue"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Modbus RTU/ASCII Master/Slave ±¸Çö"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Ethernet (Modbus TCP)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Beckhoff PLC ¿¬µ¿"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "callout", "callout": {"icon": {"type": "emoji", "emoji": "??"}, "color": "green_background", "rich_text": [{"type": "text", "text": {"content": "¼º°ú: ¡¾0.1% ¾Ğ·Â Á¤È®µµ, < 2ÃÊ Á¤Âø½Ã°£, µ¿¾Æ´ë ¿¬±¸ Çù·Â"}}]}},
    ]
)

# Project 3: Nu-2000
toggle3 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "3?? Nu-2000 Optical Analysis System | 2021-2022"}, "annotations": {"bold": True}}],
                "color": "yellow_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle3["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "±¤ÇĞ Èí¼ö ºĞ±¤ ½Ã½ºÅÛ - Multi-channel Æß¿ş¾î °³¹ß"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "FreeRTOS 5°³ Task ±¸Á¶"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "4Ã¤³Î LED Driver PWM Á¦¾î"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "4Ã¤³Î Photodiode °í¼Ó ADC (250kSPS)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "IAP Bootloader °³¹ß (UART/USB)"}}]}},
    ]
)

# Project 4: MS
toggle4 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "4?? MS (Mass Spectrometer) | 2024-2025 | ?? ÃÖ½Å"}, "annotations": {"bold": True}}],
                "color": "green_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle4["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "¹İµµÃ¼ °øÁ¤ °¡½º ºĞ¼® Mass Spectrometer"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "µ¥ÀÌÅÍ ¼öÁı Æß¿ş¾î °³¹ß"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "RF Á¦¾î ÀÎÅÍÆäÀÌ½º"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Åë½Å ÇÁ·ÎÅäÄİ ±¸Çö"}}]}},
    ]
)

# Project 5: LE_Laser
toggle5 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "5?? LE_Laser (Mantis SSC) | 2025 | ?? ÃÖ½Å"}, "annotations": {"bold": True}}],
                "color": "orange_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle5["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "·¹ÀÌÀú ±â¹İ Á¤¹Ğ ¼¾¼­ ½Ã½ºÅÛ"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "STM32G071RB (ÃÖ½Å Cortex-M0+)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "22-bit ADC °íºĞÇØ´É ÃøÁ¤"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Compact 2-board design"}}]}},
    ]
)

# 5. Code Samples
print("?? Code Samples Ãß°¡...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Code Samples"}}]}},
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "PID Controller Implementation"}}]}},
        {
            "object": "block",
            "type": "code",
            "code": {
                "language": "c",
                "rich_text": [{"type": "text", "text": {"content": """// PID Update with Anti-windup
float PID_Update(PID_t* pid, float setpoint, float measurement) {
    float error = setpoint - measurement;
    
    // Proportional
    float P = pid->Kp * error;
    
    // Integral with Anti-windup
    pid->integral += pid->Ki * 0.001f * error;
    pid->integral = CLAMP(pid->integral, -pid->i_max, pid->i_max);
    
    // Derivative with Filter
    float deriv = (error - pid->prev_error) / 0.001f;
    pid->filtered_d = 0.1f * deriv + 0.9f * pid->filtered_d;
    float D = pid->Kd * pid->filtered_d;
    
    pid->prev_error = error;
    return CLAMP(P + pid->integral + D, pid->out_min, pid->out_max);
}"""}}]
            }
        },
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "FreeRTOS Task with Queue"}}]}},
        {
            "object": "block",
            "type": "code",
            "code": {
                "language": "c",
                "rich_text": [{"type": "text", "text": {"content": """// ADC Acquisition Task
void ADC_Task(void *arg) {
    uint16_t adc_data[4];
    
    while(1) {
        HAL_ADC_Start_DMA(&hadc1, (uint32_t*)adc_data, 4);
        
        if (xSemaphoreTake(ADC_Sem, pdMS_TO_TICKS(100)) == pdTRUE) {
            xQueueSend(ADC_Queue, &adc_data, 0);
        }
        osDelay(10);  // 100Hz
    }
}"""}}]
            }
        },
    ]
)

# 6. Technical Skills
print("??? Technical Skills Ãß°¡...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "??? Technical Skills"}}]}},
        {
            "object": "block",
            "type": "table",
            "table": {
                "table_width": 2,
                "has_column_header": True,
                "has_row_header": False,
                "children": [
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Category"}}], [{"type": "text", "text": {"content": "Technologies"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "MCU"}}], [{"type": "text", "text": {"content": "STM32 F1/F4/F7/H7/G0, ARM Cortex-M"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "RTOS"}}], [{"type": "text", "text": {"content": "FreeRTOS (Task, Queue, Semaphore, Mutex)"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Languages"}}], [{"type": "text", "text": {"content": "C (Expert), C++, Python"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Protocols"}}], [{"type": "text", "text": {"content": "Modbus RTU/ASCII/TCP, UART, SPI, I2C, CAN"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Tools"}}], [{"type": "text", "text": {"content": "STM32CubeIDE, Altium Designer, Git, Logic Analyzer"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Control"}}], [{"type": "text", "text": {"content": "PID, Auto-tuning, Kalman Filter, State Machine"}}]]}},
                ]
            }
        },
    ]
)

# 7. Key Achievements
print("?? Key Achievements Ãß°¡...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Key Achievements"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "12+ ¾ç»ê Æß¿ş¾î ¹èÆ÷, Ä¡¸íÀû ¹ö±× Zero"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "99.8%+ ½Ã½ºÅÛ °¡µ¿·ü ´Ş¼º"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "PID ¶óÀÌºê·¯¸®, IAP Bootloader - 5°³ Á¦Ç° Àç»ç¿ë"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "20+ ±â¼ú ¹®¼­, 15+ »ç¿ëÀÚ ¸Å´º¾ó ÀÛ¼º"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "µ¿¾Æ´ëÇĞ±³ PID ¾Ë°í¸®Áò ¿¬±¸ Çù·Â"}}]}},
    ]
)

# 8. Footer
print("?? Contact Ãß°¡...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "divider", "divider": {}},
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "??"},
                "color": "gray_background",
                "rich_text": [{"type": "text", "text": {"content": "Contact: 92lock@kakao.com | 010-7311-0402 | github.com/gari210404"}}]
            }
        },
    ]
)

print("\n" + "=" * 60)
print("? Æ÷Æ®Æú¸®¿À »ı¼º ¿Ï·á!")
print("=" * 60)
print(f"\n?? ÆäÀÌÁö URL: https://notion.so/{page_id.replace('-', '')}")
print("\n³ë¼Ç¿¡¼­ '»õ ÆäÀÌÁö' ÇÏÀ§¿¡¼­ È®ÀÎÇÏ¼¼¿ä!")

"""
ì¡°ë½???¬íŠ¸?´ë¦¬???¸ì…˜ ?ì„± ?¤í¬ë¦½íŠ¸ (ê³ ê¸‰ ë²„ì „)
- ë¸”ë¡ ?¤ì´?´ê·¸???´ë?ì§€ ?¬í•¨
- ?ì„¸??ê¸°ìˆ  ?¤ëª…
- ì½”ë“œ ?˜í”Œ ?¬í•¨
"""
from notion_client import Client
import os
import time

NOTION_TOKEN = os.getenv('NOTION_TOKEN', 'YOUR_NOTION_TOKEN_HERE')
PARENT_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"

notion = Client(auth=NOTION_TOKEN)

print("=" * 60)
print("?? ì¡°ë½???¬íŠ¸?´ë¦¬???¸ì…˜ ?ë™ ?ì„± (Enhanced Version)")
print("=" * 60)

# 1. ë©”ì¸ ?˜ì´ì§€ ?ì„±
print("\n?“ ë©”ì¸ ?˜ì´ì§€ ?ì„± ì¤?..")

page = notion.pages.create(
    parent={"page_id": PARENT_ID},
    icon={"type": "emoji", "emoji": "?‘¨?ğŸ’?},
    cover={
        "type": "external",
        "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1920"}
    },
    properties={
        "title": {"title": [{"text": {"content": "ì¡°ë½??| Senior Embedded Systems Engineer Portfolio"}}]}
    },
    children=[
        # ?¤ë” Callout
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "?¯"},
                "color": "blue_background",
                "rich_text": [{"type": "text", "text": {"content": "9??ê²½ë ¥ Hardware & Firmware Full-Stack ê°œë°œ??| STM32 Expert | FreeRTOS | PID Control | PCB Design"}, "annotations": {"bold": True}}]
            }
        },
        # ?°ë½ì²?
        {
            "object": "block",
            "type": "paragraph",
            "paragraph": {
                "rich_text": [
                    {"type": "text", "text": {"content": "?“§ "}, "annotations": {"color": "gray"}},
                    {"type": "text", "text": {"content": "92lock@kakao.com", "link": {"url": "mailto:92lock@kakao.com"}}, "annotations": {"color": "blue"}},
                    {"type": "text", "text": {"content": "  |  ?“± 010-7311-0402  |  "}, "annotations": {"color": "gray"}},
                    {"type": "text", "text": {"content": "?”— GitHub", "link": {"url": "https://github.com/gari210404"}}, "annotations": {"color": "blue"}},
                    {"type": "text", "text": {"content": "  |  ?“ ê²½ê¸°???©ì¸??}, "annotations": {"color": "gray"}}
                ]
            }
        },
        {"object": "block", "type": "divider", "divider": {}},
    ]
)

page_id = page["id"]
print(f"  ??ë©”ì¸ ?˜ì´ì§€ ?ì„±: {page_id}")

# 2. Executive Summary
print("\n?“‹ Executive Summary ì¶”ê?...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?“‹ Executive Summary"}, "annotations": {"bold": True}}], "color": "blue"}},
        {
            "object": "block",
            "type": "quote",
            "quote": {
                "rich_text": [{"type": "text", "text": {"content": "9??ê²½ë ¥???œë‹ˆ???„ë² ?”ë“œ ?œìŠ¤???”ì??ˆì–´ë¡? ?˜ë“œ?¨ì–´ ?¤ê³„ë¶€???Œì›¨??ê°œë°œê¹Œì? Full-Stack Embedded ê°œë°œ ?„ë¬¸ê°€?…ë‹ˆ?? ?¹íˆ STM32 MCU ê¸°ë°˜???°ì—…???œì–´ ?œìŠ¤?? ë°˜ë„ì²?ê³µì • ?¥ë¹„, ê´‘í•™ ë¶„ì„ ê¸°ê¸° ê°œë°œ?ì„œ ê²€ì¦ëœ ?¤ì ??ë³´ìœ ?˜ê³  ?ˆìŠµ?ˆë‹¤."}, "annotations": {"italic": True}}]
            }
        },
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "?’¡"},
                "color": "yellow_background",
                "rich_text": [{"type": "text", "text": {"content": "?µì‹¬ ê°•ì : PID ?œì–´ ?Œê³ ë¦¬ì¦˜, FreeRTOS ?¤ì‹œê°??œìŠ¤?? ê³ ì† ADC ?°ì´??ì²˜ë¦¬, Modbus ?µì‹  ?„ë¡œ? ì½œ, ê³ ì •ë°€ ?„ë‚ ë¡œê·¸ ?Œë¡œ ?¤ê³„"}, "annotations": {"bold": True}}]
            }
        },
        {"object": "block", "type": "divider", "divider": {}},
    ]
)

# 3. Technical Stack
print("?› ï¸?Technical Stack ì¶”ê?...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?› ï¸?Technical Stack"}, "annotations": {"bold": True}}], "color": "purple"}},
    ]
)

# 3-1. Firmware
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_2", "heading_2": {"rich_text": [{"type": "text", "text": {"content": "?’» Firmware Development"}}]}},
        {
            "object": "block",
            "type": "table",
            "table": {
                "table_width": 3,
                "has_column_header": True,
                "has_row_header": False,
                "children": [
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Category"}}], [{"type": "text", "text": {"content": "Technology"}}], [{"type": "text", "text": {"content": "Proficiency"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "MCU"}}], [{"type": "text", "text": {"content": "STM32F4/F7/H7, Arduino, ESP32"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "RTOS"}}], [{"type": "text", "text": {"content": "FreeRTOS (Task, Queue, Semaphore, Mutex)"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Peripherals"}}], [{"type": "text", "text": {"content": "ADC, DAC, PWM, Timer, DMA, UART, SPI, I2C, CAN"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Communication"}}], [{"type": "text", "text": {"content": "Modbus RTU/ASCII/TCP, TCP/IP (lwIP), Ethernet"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Control"}}], [{"type": "text", "text": {"content": "PID (Anti-windup, D-filter), Auto-tuning, State Machine"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Bootloader"}}], [{"type": "text", "text": {"content": "IAP, Y-modem, UART/USB Firmware Update"}}], [{"type": "text", "text": {"content": "â­â­â­â­"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "UI/HMI"}}], [{"type": "text", "text": {"content": "Nextion LCD, TouchGFX, Custom UI Design"}}], [{"type": "text", "text": {"content": "â­â­â­â­"}}]]}},
                ]
            }
        },
    ]
)

# 3-2. Hardware
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_2", "heading_2": {"rich_text": [{"type": "text", "text": {"content": "?”Œ Hardware Design"}}]}},
        {
            "object": "block",
            "type": "table",
            "table": {
                "table_width": 3,
                "has_column_header": True,
                "has_row_header": False,
                "children": [
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Category"}}], [{"type": "text", "text": {"content": "Technology"}}], [{"type": "text", "text": {"content": "Proficiency"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Analog"}}], [{"type": "text", "text": {"content": "OP-Amp (TIA, Inst-Amp), ADC/DAC, Filter Design"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Power"}}], [{"type": "text", "text": {"content": "SMPS, LDO, Multi-rail Design, EMI/EMC"}}], [{"type": "text", "text": {"content": "â­â­â­â­"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Digital"}}], [{"type": "text", "text": {"content": "MCU Circuit, Logic Level, Pull-up/down, ESD"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "PCB Design"}}], [{"type": "text", "text": {"content": "PADS, Altium Designer, 4-layer, Stack-up"}}], [{"type": "text", "text": {"content": "â­â­â­â­"}}]]}},
                    {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Sensor"}}], [{"type": "text", "text": {"content": "Pressure, Temperature, Optical, pH Electrode"}}], [{"type": "text", "text": {"content": "â­â­â­â­â­?}}]]}},
                ]
            }
        },
        {"object": "block", "type": "divider", "divider": {}},
    ]
)

# 4. Major Projects
print("?¯ Major Projects ì¶”ê?...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?¯ Major Projects"}, "annotations": {"bold": True}}], "color": "orange"}},
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "?“Š"},
                "color": "gray_background",
                "rich_text": [{"type": "text", "text": {"content": "ê°??„ë¡œ?íŠ¸???˜ë“œ?¨ì–´ ?¤ê³„, ?Œì›¨??ê°œë°œ, ?”ë²„ê¹?ê²½í—˜???¬í•¨?©ë‹ˆ?? ë¸”ë¡ ?¤ì´?´ê·¸?¨ê³¼ ì½”ë“œ ?˜í”Œ???¨ê»˜ ?œê³µ?©ë‹ˆ??"}, "annotations": {"italic": True}}]
            }
        },
    ]
)

# ==================== PROJECT 1: L-LPC (H-Sensor) ====================
print("  ?“Œ L-LPC ?„ë¡œ?íŠ¸ ì¶”ê?...")
toggle1 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "1ï¸âƒ£ L-LPC (H-Sensor) System | 2022-2024 | â­â­â­??€???„ë¡œ?íŠ¸"}, "annotations": {"bold": True}}],
                "color": "blue_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?“ ?„ë¡œ?íŠ¸ ê°œìš”"}, "annotations": {"bold": True}}], "color": "blue"}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "ë°˜ë„ì²?ê³µì •???€??ì±”ë²„ ?œì–´ ?œìŠ¤??}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??• : "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Hardware & Firmware Full-Stack ê°œë°œ (?¨ë…)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ê¸°ê°„: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "2022.01 ~ 2024.12 (3??"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ëª©ì : "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "ë°˜ë„ì²?ê³µì • ì±”ë²„???•ë? ?•ë ¥ ?œì–´ ë°??°ì´???˜ì§‘"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        
        # ë¸”ë¡ ?¤ì´?´ê·¸??
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?“ System Block Diagram"}, "annotations": {"bold": True}}], "color": "purple"}},
        {"object": "block", "type": "callout", "callout": {"icon": {"type": "emoji", "emoji": "?–¼ï¸?}, "color": "gray_background", "rich_text": [{"type": "text", "text": {"content": "H-Sensor FPGA ?œìŠ¤???„í‚¤?ì²˜ (9ê°?ë¸”ë¡ ?¤ì´?´ê·¸??"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "???¤ì œ ?„ë¡œ?íŠ¸??ë¸”ë¡ ?¤ì´?´ê·¸???´ë?ì§€??Project_Files/02_L-LPC/Images/ ?´ë”???ˆìŠµ?ˆë‹¤."}, "annotations": {"color": "gray", "italic": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "FPGA Interface: Xilinx Artix-7 SPI ?µì‹ "}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Analog Frontend: TIA + Inst-Amp + ADC"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Main Controller: STM32H743 + FreeRTOS"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        
        # ?˜ë“œ?¨ì–´ ?¤ê³„
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?”§ Hardware Design"}, "annotations": {"bold": True}}], "color": "green"}},
        {"object": "block", "type": "toggle", "toggle": {"rich_text": [{"type": "text", "text": {"content": "Main Board Specifications"}, "annotations": {"bold": True}}]}},
    ]
)

# Main Board ?ì„¸ ?´ìš©
main_board_toggle = notion.blocks.children.list(block_id=toggle1["results"][0]["id"])
main_board_id = [b for b in main_board_toggle["results"] if b["type"] == "toggle"][-1]["id"]

notion.blocks.children.append(
    block_id=main_board_id,
    children=[
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "MCU: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "STM32H743VIT6 (480MHz Cortex-M7, 2MB Flash, 1MB RAM)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Analog: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "16-bit ADC (ADS1115), 16-bit DAC (DAC8552)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Communication: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Ethernet (LAN8742A), RS485, SPI, I2C"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Power: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Multi-rail SMPS (TPS54360, LM2596, AMS1117) - ?¨ìœ¨ > 90%"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "PCB: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "4-layer, PADS Layout, 120mm x 80mm"}}]}},
    ]
)

time.sleep(0.5)  # Rate limit ë°©ì?

# Analog Frontend ?ì„¸
notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "toggle", "toggle": {"rich_text": [{"type": "text", "text": {"content": "Analog Frontend Circuit"}, "annotations": {"bold": True}}]}},
    ]
)

analog_toggle = notion.blocks.children.list(block_id=toggle1["results"][0]["id"])
analog_id = [b for b in analog_toggle["results"] if b["type"] == "toggle"][-1]["id"]

notion.blocks.children.append(
    block_id=analog_id,
    children=[
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Transimpedance Amplifier (TIA): OPA657 (1GHz GBW)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Instrumentation Amplifier: INA826 (CMRR > 100dB)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Anti-aliasing Filter: 2nd-order Butterworth (fc = 100kHz)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Signal-to-Noise Ratio: > 60dB"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Dynamic Range: 16-bit (0.0015% resolution)"}}]}},
    ]
)

time.sleep(0.5)

# ?Œì›¨??ê°œë°œ
notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?’» Firmware Development"}, "annotations": {"bold": True}}], "color": "purple"}},
        {"object": "block", "type": "toggle", "toggle": {"rich_text": [{"type": "text", "text": {"content": "FreeRTOS Task Architecture (7ê°?Task)"}, "annotations": {"bold": True}}]}},
    ]
)

firmware_toggle = notion.blocks.children.list(block_id=toggle1["results"][0]["id"])
firmware_id = [b for b in firmware_toggle["results"] if b["type"] == "toggle"][-1]["id"]

notion.blocks.children.append(
    block_id=firmware_id,
    children=[
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "SensorTask: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "ADC ?°ì´???˜ì§‘ (2MSPS, DMA), Kalman ?„í„°ë§?}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "ControlTask: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "PID ?œì–´ ë£¨í”„ (1kHz), DAC ì¶œë ¥"}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "CommunicationTask: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Modbus RTU/TCP, Beckhoff PLC ?µì‹ "}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "FPGATask: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "SPI ?¸í„°?˜ì´?? ?°ì´???™ê¸°??}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "LogTask: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Flash ë©”ëª¨ë¦?ë¡œê¹…, ?´ë²¤??ê¸°ë¡"}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "UITask: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "LCD ?”ìŠ¤?Œë ˆ?? ?¬ìš©???…ë ¥ ì²˜ë¦¬"}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "WatchdogTask: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "?œìŠ¤??ëª¨ë‹ˆ?°ë§, ?¤ë¥˜ ë³µêµ¬"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "callout", "callout": {"icon": {"type": "emoji", "emoji": "?™ï¸"}, "color": "blue_background", "rich_text": [{"type": "text", "text": {"content": "Task ê°??µì‹ : Queue 3ê°? Semaphore 5ê°? Mutex 2ê°??¬ìš©"}}]}},
    ]
)

time.sleep(0.5)

# ì½”ë“œ ?˜í”Œ
notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "toggle", "toggle": {"rich_text": [{"type": "text", "text": {"content": "?“ Firmware Code Sample"}, "annotations": {"bold": True, "code": True}}]}},
    ]
)

code_toggle = notion.blocks.children.list(block_id=toggle1["results"][0]["id"])
code_id = [b for b in code_toggle["results"] if b["type"] == "toggle"][-1]["id"]

notion.blocks.children.append(
    block_id=code_id,
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "FreeRTOS Sensor Task ?ˆì œ:"}, "annotations": {"bold": True}}]}},
        {
            "object": "block",
            "type": "code",
            "code": {
                "language": "c",
                "rich_text": [{"type": "text", "text": {"content": """void SensorTask(void *pvParameters) {
    TickType_t xLastWakeTime = xTaskGetTickCount();
    const TickType_t xFrequency = pdMS_TO_TICKS(1); // 1ms period
    
    while(1) {
        // DMAë¡?ADC ?°ì´???˜ì§‘
        HAL_ADC_Start_DMA(&hadc1, (uint32_t*)adc_buffer, ADC_BUFFER_SIZE);
        
        // Kalman ?„í„° ?ìš©
        kalman_update(&kf, adc_buffer[0]);
        float filtered_value = kalman_get_state(&kf);
        
        // Queueë¡??œì–´ Task???„ì†¡
        xQueueSend(xSensorQueue, &filtered_value, 0);
        
        // 1ms ì£¼ê¸°ë¡??¤í–‰
        vTaskDelayUntil(&xLastWakeTime, xFrequency);
    }
}"""}}]
            }
        },
    ]
)

time.sleep(0.5)

# ?”ë²„ê¹?ê²½í—˜
notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?› Debugging Experience"}, "annotations": {"bold": True}}], "color": "red"}},
        {"object": "block", "type": "toggle", "toggle": {"rich_text": [{"type": "text", "text": {"content": "Issue #1: ADC ?¸ì´ì¦?ë¬¸ì œ (Â±500mV ??Â±5mV)"}, "annotations": {"bold": True}}]}},
    ]
)

debug1_toggle = notion.blocks.children.list(block_id=toggle1["results"][0]["id"])
debug1_id = [b for b in debug1_toggle["results"] if b["type"] == "toggle"][-1]["id"]

notion.blocks.children.append(
    block_id=debug1_id,
    children=[
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ë¬¸ì œ: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "ADC ?…ë ¥?ì„œ Â±500mV?????¸ì´ì¦?ë°œìƒ"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?ì¸ ë¶„ì„: "}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??SMPS ?¤ìœ„ì¹??¸ì´ì¦ˆê? ?„ë‚ ë¡œê·¸ GNDë¡?? ì…"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??ADC VREF ë¶ˆì•ˆ??(ë¦¬í”Œ > 50mVpp)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??PCB ?ˆì´?„ì›ƒ?ì„œ ?”ì????„ë‚ ë¡œê·¸ GND ?¼ì¬"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?´ê²° ë°©ë²•: "}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??AGND/DGND ë¶„ë¦¬ ??Single-point ?°ê²°"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??ADC VREF??2??LC ?„í„° ì¶”ê? (ë¦¬í”Œ < 5mVpp)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??Ferrite Bead ì¶”ê?, Star-ground ?¨í„´ ?ìš©"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ê²°ê³¼: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "?¸ì´ì¦?Â±500mV ??Â±5mV (100ë°?ê°œì„ ), S/N ë¹„ìœ¨ > 60dB ?¬ì„±"}, "annotations": {"color": "green"}}]}},
    ]
)

time.sleep(0.5)

# ?”ë²„ê¹?#2
notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "toggle", "toggle": {"rich_text": [{"type": "text", "text": {"content": "Issue #2: FreeRTOS ?°ë“œ??ë¬¸ì œ"}, "annotations": {"bold": True}}]}},
    ]
)

debug2_toggle = notion.blocks.children.list(block_id=toggle1["results"][0]["id"])
debug2_id = [b for b in debug2_toggle["results"] if b["type"] == "toggle"][-1]["id"]

notion.blocks.children.append(
    block_id=debug2_id,
    children=[
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ë¬¸ì œ: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "?œìŠ¤?œì´ ?œë¤?˜ê²Œ ë©ˆì¶”???„ìƒ (??2?œê°„ë§ˆë‹¤)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?ì¸: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "SensorTask?€ ControlTask ê°?Mutex ?ë“ ?œì„œ ë¶ˆì¼ì¹˜ë¡œ Deadlock ë°œìƒ"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?´ê²°: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Mutex ?ë“ ?œì„œë¥??„ì—­?ìœ¼ë¡??µì¼, Timeout ì¶”ê? (100ms)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ê²°ê³¼: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "72?œê°„ ?°ì† ?™ì‘ ?ŒìŠ¤???µê³¼, ?„ë“œ ë°°í¬ ??ë¬´ì •ì§€ ?´ì˜"}, "annotations": {"color": "green"}}]}},
    ]
)

time.sleep(0.5)

# ?±ê³¼
notion.blocks.children.append(
    block_id=toggle1["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?“Š Results & Achievements"}, "annotations": {"bold": True}}], "color": "green"}},
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "?†"},
                "color": "green_background",
                "rich_text": [
                    {"type": "text", "text": {"content": "???‘ì‚°: 120+ units ì¶œí•˜\n"}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "???„ë“œ ë¶ˆëŸ‰ë¥? < 0.2% (ëª©í‘œ < 1%)\n"}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "??ê³ ê°???‰ê?: 4.8/5.0\n"}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "???©ê¸°: ?ˆì • ?€ë¹?-2ì£??¨ì¶•"}, "annotations": {"bold": True}}
                ]
            }
        },
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
    ]
)

print("  ??L-LPC ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")

# ==================== PROJECT 2: Psi-1000/3000 ====================
print("  ?“Œ Psi-1000/3000 ?„ë¡œ?íŠ¸ ì¶”ê?...")
toggle2 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "2ï¸âƒ£ Psi-1000/3000 Pressure Controller | 2019-2023 | â­â­ PID ?œì–´ ?„ë¬¸"}, "annotations": {"bold": True}}],
                "color": "purple_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle2["results"][0]["id"],
    children=[
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?“ ?„ë¡œ?íŠ¸ ê°œìš”"}, "annotations": {"bold": True}}], "color": "purple"}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "?•ë? ì§„ê³µ ?•ë ¥ ?œì–´ ?œìŠ¤??(ë°˜ë„ì²?ê³µì •??"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "??• : "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "PID ?Œê³ ë¦¬ì¦˜ ê°œë°œ, Firmware ?¤ê³„"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?•ë ¥ ë²”ìœ„: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "0.01 ~ 1000 Torr (4 decades)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?•í™•?? "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Â±0.1% FS, ?¬í˜„??Â±0.05%"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        
        # PID ?œì–´ ?Œê³ ë¦¬ì¦˜
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?›ï¸?PID Control Algorithm"}, "annotations": {"bold": True}}], "color": "purple"}},
        {"object": "block", "type": "toggle", "toggle": {"rich_text": [{"type": "text", "text": {"content": "Advanced PID Implementation"}, "annotations": {"bold": True}}]}},
    ]
)

pid_toggle = notion.blocks.children.list(block_id=toggle2["results"][0]["id"])
pid_id = [b for b in pid_toggle["results"] if b["type"] == "toggle"][-1]["id"]

notion.blocks.children.append(
    block_id=pid_id,
    children=[
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "Anti-windup: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "?ë¶„???œí•œ, Conditional Integration"}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "Derivative Filter: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "1st-order Low-pass filter (fc = 10Hz), ?¸ì´ì¦??œê±°"}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "Gain Scheduling: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "?•ë ¥ ë²”ìœ„ë³?PID ê²Œì¸ ?ë™ ?„í™˜ (3?¨ê³„)"}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "Auto-tuning: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Relay Feedback Method + Ziegler-Nichols ?œë‹"}}]}},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": "Cascade Control: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Pressure (Outer) + Temperature (Inner) 2ì¤?ë£¨í”„"}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
    ]
)

time.sleep(0.5)

# ?µì‹  ?„ë¡œ? ì½œ
notion.blocks.children.append(
    block_id=toggle2["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {"object": "block", "type": "heading_3", "heading_3": {"rich_text": [{"type": "text", "text": {"content": "?“¡ Communication Protocol"}, "annotations": {"bold": True}}], "color": "blue"}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Modbus RTU: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "RS485, Master/Slave, 115200 baud"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Modbus ASCII: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "RS232, ?¸í™˜???„í•´ ë³‘í–‰ ì§€??}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Modbus TCP: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Ethernet, lwIP ?¤íƒ, ?ê²© ëª¨ë‹ˆ?°ë§"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "PLC ?°ë™: "}, "annotations": {"bold": True}}, {"type": "text", "text": {"content": "Beckhoff TwinCAT, Siemens S7 ?µì‹  ê²€ì¦?}}]}},
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": ""}}]}},
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "?†"},
                "color": "green_background",
                "rich_text": [
                    {"type": "text", "text": {"content": "?“Š ?±ê³¼:\n"}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "???•ë ¥ ?•í™•?? Â±0.1% FS\n"}},
                    {"type": "text", "text": {"content": "???‘ë‹µ ?œê°„: < 2ì´?(10% ??90%)\n"}},
                    {"type": "text", "text": {"content": "???¤ë²„?ˆíŠ¸: < 5%\n"}},
                    {"type": "text", "text": {"content": "???™ì•„?€?™êµ ?°êµ¬???‘ë ¥ ?„ë¡œ?íŠ¸"}}
                ]
            }
        },
    ]
)

print("  ??Psi-1000/3000 ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")

# ==================== ê°„ëµ ?„ë¡œ?íŠ¸??====================
print("  ?“Œ ?˜ë¨¸ì§€ ?„ë¡œ?íŠ¸ ì¶”ê?...")

# Nu-2000
toggle3 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "3ï¸âƒ£ Nu-2000 Optical Analysis System | 2021-2022"}, "annotations": {"bold": True}}],
                "color": "yellow_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle3["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "ê´‘í•™ ?¡ìˆ˜ ë¶„ê´‘ ?ë„ ì¸¡ì • ?œìŠ¤??}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "4ì±„ë„ LED Driver PWM ?œì–´ (UV/IR)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "4ì±„ë„ Photodiode ADC (250kSPS)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "IAP Bootloader (UART/USB)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?¹í—ˆ: ?°ì´??? íš¨???•ì¸ ê¸°ëŠ¥ ë°œëª… (ì¶œì›)"}}]}},
    ]
)

# MS
toggle4 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "4ï¸âƒ£ MS (Mass Spectrometer) | 2024-2025 | ?†•"}, "annotations": {"bold": True}}],
                "color": "green_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle4["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "ë°˜ë„ì²?ê³µì • ê°€??ë¶„ì„ Mass Spectrometer"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "ê³ ì† ?°ì´???˜ì§‘ ?Œì›¨??(1MSPS)"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "RF Generator ?œì–´ ?¸í„°?˜ì´??}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "SPI/UART ?µì‹  ?„ë¡œ? ì½œ"}}]}},
    ]
)

# LE_Laser
toggle5 = notion.blocks.children.append(
    block_id=page_id,
    children=[
        {
            "object": "block",
            "type": "toggle",
            "toggle": {
                "rich_text": [{"type": "text", "text": {"content": "5ï¸âƒ£ LE_Laser (Mantis SSC) | 2024-2025 | ?†•"}, "annotations": {"bold": True}}],
                "color": "orange_background"
            }
        }
    ]
)

notion.blocks.children.append(
    block_id=toggle5["results"][0]["id"],
    children=[
        {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": "?ˆì´?€ ?Œì›Œ & ë¹??¬ê¸° ë¶„ì„ ?¥ë¹„"}, "annotations": {"bold": True}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Arduino ê¸°ë°˜ ?„ë¡œ? í???ê°œë°œ"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "Photodiode array ?°ì´???˜ì§‘"}}]}},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "?¤ì‹œê°?ë¹??„ë¡œ?Œì¼ ë¶„ì„"}}]}},
    ]
)

print("  ???˜ë¨¸ì§€ ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")

# 5. Contact & Links
print("\n?“ Contact ì¶”ê?...")
notion.blocks.children.append(
    block_id=page_id,
    children=[
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_1", "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?“ Contact & Links"}, "annotations": {"bold": True}}], "color": "blue"}},
        {
            "object": "block",
            "type": "callout",
            "callout": {
                "icon": {"type": "emoji", "emoji": "?‰ï¸"},
                "color": "blue_background",
                "rich_text": [
                    {"type": "text", "text": {"content": "?“§ Email: "}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "92lock@kakao.com\n", "link": {"url": "mailto:92lock@kakao.com"}}},
                    {"type": "text", "text": {"content": "?“± Mobile: "}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "010-7311-0402\n"}},
                    {"type": "text", "text": {"content": "?”— GitHub: "}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "github.com/gari210404\n", "link": {"url": "https://github.com/gari210404"}}},
                    {"type": "text", "text": {"content": "?“ Portfolio: "}, "annotations": {"bold": True}},
                    {"type": "text", "text": {"content": "github.com/gari210404/Portfolio_Professional", "link": {"url": "https://github.com/gari210404/Portfolio_Professional"}}}
                ]
            }
        },
    ]
)

print("\n" + "=" * 60)
print("???¬íŠ¸?´ë¦¬???ì„± ?„ë£Œ!")
print("=" * 60)
print(f"\n?“ ?˜ì´ì§€ URL: https://notion.so/{page_id.replace('-', '')}")
print("\n?’¡ ë¸”ë¡ ?¤ì´?´ê·¸???´ë?ì§€??ì§ì ‘ ?…ë¡œ?œí•´ì£¼ì„¸??")
print("   - Project_Files/02_L-LPC/Images/")
print("   - Project_Files/03_Psi-1000/Images/")
print("   - Project_Files/01_MS_Mass_Spectrometer/Images/")

# -*- coding: utf-8 -*-
"""
ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v10 - Professional Design (?ì–´ + ì·¨ë? ?¹ì…˜)
"""

from notion_client import Client
import time

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {"object": "block", "type": heading_type, heading_type: {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_paragraph(text="", bold=False):
    if not text:
        return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}
    return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": text}, "annotations": {"bold": bold}}]}}

def create_bulleted_list(text):
    return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_callout(text, emoji="?’¡", color="default"):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": text}}], "icon": {"type": "emoji", "emoji": emoji}, "color": color}}

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def create_toggle(title, emoji="?“"):
    return {
        "object": "block",
        "type": "toggle",
        "toggle": {
            "rich_text": [{"type": "text", "text": {"content": f"{emoji} {title}"}}],
            "children": []
        }
    }

def create_column_list_2col():
    """2ê°?ì»¬ëŸ¼??ê°€ì§?column_list ?ì„± (ê°?ì»¬ëŸ¼??ë¹?paragraph ?„ìˆ˜)"""
    return {
        "object": "block",
        "type": "column_list",
        "column_list": {
            "children": [
                {
                    "object": "block", 
                    "type": "column", 
                    "column": {
                        "children": [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}]
                    }
                },
                {
                    "object": "block", 
                    "type": "column", 
                    "column": {
                        "children": [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}]
                    }
                }
            ]
        }
    }

def create_column_list_3col():
    """3ê°?ì»¬ëŸ¼ (ì·¨ë???"""
    return {
        "object": "block",
        "type": "column_list",
        "column_list": {
            "children": [
                {"object": "block", "type": "column", "column": {"children": [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}]}},
                {"object": "block", "type": "column", "column": {"children": [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}]}},
                {"object": "block", "type": "column", "column": {"children": [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}]}}
            ]
        }
    }

def main():
    print("=" * 80)
    print("?“„ ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v10 - Professional Design")
    print("=" * 80)
    
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={"title": {"title": [{"type": "text", "text": {"content": "ì¡°ë½??| Rakeen Jo - Resume"}}]}},
        icon={"type": "emoji", "emoji": "?’¼"},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1451187580459-43490279c0fa?w=1200"}}
    )
    page_id = new_page["id"]
    print(f"???˜ì´ì§€ ?ì„±: {page_id}\n")
    
    # ===== ?¤ë” (?„ë¡œ???ì—­) =====
    header_blocks = [
        create_callout("?‘¨?ğŸ’?, "?¯", "blue_background"),
        create_heading("ì¡°ë½??| Rakeen Jo", 1),
        create_paragraph(""),
        create_callout("??Embedded System Engineer | 9 Years (2016~Present) | HW/FW Full-Stack", "?’¼", "blue_background"),
        create_paragraph(""),
        create_callout("?“§ 92lock@kakao.com  |  ?“± 010-7311-0402  |  ?™ github.com/Rakeen-Jo", "?“", "gray_background"),
        create_callout("?”— Portfolio: [ë§í¬ ì¶”ê? ?ˆì •]", "?”—", "gray_background"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=header_blocks)
    print("???¤ë”")
    time.sleep(0.3)
    
    # ===== ?µì‹¬ ??Ÿ‰ (?œê? ? ì?) =====
    summary_blocks = [
        create_heading("?’¼ ?µì‹¬ ??Ÿ‰ | Core Competencies", 2),
        create_callout("9?„ê°„ STM32 ê¸°ë°˜ ?„ë² ?”ë“œ ?œìŠ¤??Full-Stack ê°œë°œ ?„ë¬¸ (?Œë¡œ ?¤ê³„ ???Œì›¨?????‘ì‚°)", "?¯", "yellow_background"),
        create_paragraph(""),
        create_bulleted_list("??8+ ë¶„ì„ ?¥ë¹„ ë©”ì¸ë³´ë“œ ?¤ê³„ ë°??‘ì‚° ?„ë£Œ"),
        create_bulleted_list("??HW/FW ?µí•© ê°œë°œë¡??”ë²„ê¹??¨ìœ¨ 40% ?¥ìƒ"),
        create_bulleted_list("??FreeRTOS, PID Control, Ethernet ?µì‹  ?„ë¬¸"),
        create_bulleted_list("???ë™???ŒìŠ¤??JIG ê°œë°œë¡?ê²€???œê°„ 60% ?¨ì¶•"),
        create_bulleted_list("???¹í—ˆ ë°œëª… 2ê±?ì¶œì› (ê°€???œì–´, ê´‘í•™ ë¶„ì„)"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=summary_blocks)
    print("???µì‹¬ ??Ÿ‰")
    time.sleep(0.3)
    
    # ===== 2??ì»¬ëŸ¼ ?ˆì´?„ì›ƒ ?ì„± =====
    print("??2??ì»¬ëŸ¼ ?ì„± ì¤?..")
    column_list_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_column_list_2col()]
    )
    column_list_id = column_list_response["results"][0]["id"]
    
    # ì»¬ëŸ¼ ID ê°€?¸ì˜¤ê¸?
    time.sleep(0.3)
    columns = notion.blocks.children.list(block_id=column_list_id)
    left_col_id = columns["results"][0]["id"]
    right_col_id = columns["results"][1]["id"]
    print(f"  ???¼ìª½ ì»¬ëŸ¼ ID: {left_col_id}")
    print(f"  ???¤ë¥¸ìª?ì»¬ëŸ¼ ID: {right_col_id}")
    
    # ===== ?¼ìª½ ì»¬ëŸ¼: ê²½ë ¥ (?ì–´) =====
    print("\n?“ ?¼ìª½ ì»¬ëŸ¼: ê²½ë ¥ ì¶”ê? ì¤?..")
    left_content = [
        create_heading("?¢ Work Experience", 2),
        create_paragraph(""),
        
        # ATIK
        create_callout("ATIK Co., Ltd. - Senior R&D Engineer", "?¢", "blue_background"),
        create_paragraph("?“… Aug 2020 ~ Jan 2025 (4y 6m)", bold=True),
        create_paragraph("?’¼ R&D Center / Main Board HW/FW Design", bold=True),
    ]
    notion.blocks.children.append(block_id=left_col_id, children=left_content)
    time.sleep(0.3)
    
    # ATIK ?„ë¡œ?íŠ¸ ? ê?
    atik_toggle = notion.blocks.children.append(
        block_id=left_col_id,
        children=[create_toggle("8 Major Projects", "?“‚")]
    )
    atik_toggle_id = atik_toggle["results"][0]["id"]
    
    atik_proj = [
        create_bulleted_list("Psi-1000/3000: Precision Gas Pressure Control System"),
        create_bulleted_list("LE_Laser: Laser Particle Analyzer"),
        create_bulleted_list("MS: Mass Spectrometer DRV Board"),
        create_bulleted_list("JIG System: Automated Board Test Equipment"),
        create_bulleted_list("L-Titrator: Titration Concentration Analyzer"),
        create_bulleted_list("Nu-2000: UV+IR Optical Concentration Analyzer"),
        create_bulleted_list("GasQuantrol: Gas Control Equipment"),
        create_bulleted_list("Sigma/Epsilon/L-LPC Series"),
    ]
    notion.blocks.children.append(block_id=atik_toggle_id, children=atik_proj)
    print("  ??ATIK")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=left_col_id, children=[create_paragraph("")])
    
    # Chemtronics
    chem_content = [
        create_callout("Chemtronics Co., Ltd. - Junior R&D Engineer", "?¢", "blue_background"),
        create_paragraph("?“… Aug 2019 ~ Aug 2020 (1y)", bold=True),
        create_paragraph("?’¼ Consumer Electronics Team / Touch Module", bold=True),
    ]
    notion.blocks.children.append(block_id=left_col_id, children=chem_content)
    time.sleep(0.3)
    
    chem_toggle = notion.blocks.children.append(
        block_id=left_col_id,
        children=[create_toggle("4 Major Projects", "?“‚")]
    )
    chem_toggle_id = chem_toggle["results"][0]["id"]
    
    chem_proj = [
        create_bulleted_list("LG Induction/Oven Touch PBA Mass Production"),
        create_bulleted_list("Samsung Air Conditioner Touch PBA Mass Production"),
        create_bulleted_list("Winix Air Purifier Touch PBA"),
        create_bulleted_list("Haatz Hood Proximity Sensor Touch PBA"),
    ]
    notion.blocks.children.append(block_id=chem_toggle_id, children=chem_proj)
    print("  ??Chemtronics")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=left_col_id, children=[create_paragraph("")])
    
    # Dreamtech
    dream_content = [
        create_callout("Dreamtech Co., Ltd. - Junior R&D Engineer", "?¢", "blue_background"),
        create_paragraph("?“… Nov 2016 ~ May 2019 (2y 7m)", bold=True),
        create_paragraph("?’¼ Advanced R&D Lab / Fingerprint Module", bold=True),
    ]
    notion.blocks.children.append(block_id=left_col_id, children=dream_content)
    time.sleep(0.3)
    
    dream_toggle = notion.blocks.children.append(
        block_id=left_col_id,
        children=[create_toggle("4 Major Projects", "?“‚")]
    )
    dream_toggle_id = dream_toggle["results"][0]["id"]
    
    dream_proj = [
        create_bulleted_list("Hyundai Motor Fingerprint Recognition PBA (Mass Production)"),
        create_bulleted_list("Mobile Ultrasonic Fingerprint Recognition PBA"),
        create_bulleted_list("FPC Fingerprint Recognition PBA ODM"),
        create_bulleted_list("3D MID Automotive Module (Government Project)"),
    ]
    notion.blocks.children.append(block_id=dream_toggle_id, children=dream_proj)
    print("  ??Dreamtech")
    time.sleep(0.3)
    
    # ===== ?¤ë¥¸ìª?ì»¬ëŸ¼: ê¸°ìˆ  ?¤íƒ (?ì–´) =====
    print("\n?“ ?¤ë¥¸ìª?ì»¬ëŸ¼: ê¸°ìˆ  ?¤íƒ ì¶”ê? ì¤?..")
    right_content = [
        create_heading("?› ï¸?Technical Skills", 2),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=right_col_id, children=right_content)
    time.sleep(0.3)
    
    # Circuit Design ? ê?
    circuit_toggle = notion.blocks.children.append(
        block_id=right_col_id,
        children=[create_toggle("Circuit Design", "??)]
    )
    circuit_id = circuit_toggle["results"][0]["id"]
    
    circuit_skills = [
        create_bulleted_list("STM32F4/F7 Main Board Design"),
        create_bulleted_list("Switching/Linear Regulator Power Supply"),
        create_bulleted_list("16-bit ADC + OP-Amp/TIA Sensor Circuit"),
        create_bulleted_list("16-bit DAC External Interface"),
        create_bulleted_list("Photocoupler Isolation Circuit"),
        create_bulleted_list("PCB Artwork (PADS, Altium Designer)"),
    ]
    notion.blocks.children.append(block_id=circuit_id, children=circuit_skills)
    print("  ??Circuit Design")
    time.sleep(0.3)
    
    # Firmware Design ? ê?
    firmware_toggle = notion.blocks.children.append(
        block_id=right_col_id,
        children=[create_toggle("Firmware Design", "?’»")]
    )
    firmware_id = firmware_toggle["results"][0]["id"]
    
    firmware_skills = [
        create_bulleted_list("STM32 Firmware Development (C)"),
        create_bulleted_list("FreeRTOS Task Management"),
        create_bulleted_list("PID Control Algorithm"),
        create_bulleted_list("Y-modem IAP Bootloader"),
        create_bulleted_list("Ethernet TCP/UDP Communication"),
        create_bulleted_list("UART/SPI/I2C Protocol"),
        create_bulleted_list("Touch LCD UI/UX Design"),
    ]
    notion.blocks.children.append(block_id=firmware_id, children=firmware_skills)
    print("  ??Firmware Design")
    time.sleep(0.3)
    
    # Tools ? ê?
    tools_toggle = notion.blocks.children.append(
        block_id=right_col_id,
        children=[create_toggle("Tools & Software", "?–¥ï¸?)]
    )
    tools_id = tools_toggle["results"][0]["id"]
    
    tools_list = [
        create_bulleted_list("Circuit: OrCAD, PADS, Altium Designer"),
        create_bulleted_list("Firmware: STM32 IDE, Keil Î¼Vision, VS Code"),
        create_bulleted_list("Software: C#, Arduino, ESP32"),
        create_bulleted_list("Version Control: Git, SVN"),
    ]
    notion.blocks.children.append(block_id=tools_id, children=tools_list)
    print("  ??Tools")
    time.sleep(0.3)
    
    # ?™ë ¥ & ?ê²© (?ì–´)
    edu_content = [
        create_paragraph(""),
        create_heading("?“ Education & Certificate", 2),
        create_callout("Myongji University - Electronics Engineering (2011~2017)", "?“", "gray_background"),
        create_paragraph(""),
        create_callout("Certifications: Computer Specialist Level 3, ITQ, TOEIC Speaking 110", "?“œ", "gray_background"),
    ]
    notion.blocks.children.append(block_id=right_col_id, children=edu_content)
    print("  ???™ë ¥ & ?ê²©")
    time.sleep(0.3)
    
    # ===== ê°œì¸ ì·¨ë?/ê´€?¬ì‚¬ (?„ì²´ ?ˆë¹„, 3??ì»¬ëŸ¼) =====
    hobby_header = [
        create_paragraph(""),
        create_divider(),
        create_heading("?¨ Personal Interests & Hobbies", 2),
        create_callout("?¼ê³¼ ?¶ì˜ ê· í˜•??ì¶”êµ¬?˜ë©°, ?¤ì–‘??ì·¨ë?ë¥??µí•´ ì°½ì˜?±ê³¼ ?¬ì¶©?„ì˜ ?œê°„??ê°€ì§‘ë‹ˆ??", "??, "purple_background"),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=page_id, children=hobby_header)
    print("\n??ì·¨ë? ?¹ì…˜ ?¤ë”")
    time.sleep(0.3)
    
    # 3??ì»¬ëŸ¼ ?ì„±
    hobby_col_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_column_list_3col()]
    )
    hobby_col_id = hobby_col_response["results"][0]["id"]
    time.sleep(0.3)
    
    hobby_cols = notion.blocks.children.list(block_id=hobby_col_id)
    hobby_col1_id = hobby_cols["results"][0]["id"]
    hobby_col2_id = hobby_cols["results"][1]["id"]
    hobby_col3_id = hobby_cols["results"][2]["id"]
    
    # ì»¬ëŸ¼ 1: ?Œì•… & LP
    hobby1 = [
        create_callout("?µ Music & Vinyl", "?¶", "pink_background"),
        create_paragraph("LP ?ˆì½”???˜ì§‘ê³??Œì•… ê°ìƒ??ì¦ê¹?ˆë‹¤. ?„ë‚ ë¡œê·¸ ?¬ìš´?œì˜ ?°ëœ»?¨ì„ ?¬ë‘?©ë‹ˆ??", bold=False),
    ]
    notion.blocks.children.append(block_id=hobby_col1_id, children=hobby1)
    
    # ì»¬ëŸ¼ 2: ?„ì›ƒ?„ì–´
    hobby2 = [
        create_callout("?£ Outdoor Activities", "??, "green_background"),
        create_paragraph("?šì‹œ?€ ìº í•‘??ì¢‹ì•„?©ë‹ˆ?? ?ì—° ?ì—?œì˜ ?ë§ê³??¬í–‰??ì¦ê¹?ˆë‹¤.", bold=False),
    ]
    notion.blocks.children.append(block_id=hobby_col2_id, children=hobby2)
    
    # ì»¬ëŸ¼ 3: ?Œì‹ & ë¬¸í™”
    hobby3 = [
        create_callout("?£ Food & Culture", "?—¾", "orange_background"),
        create_paragraph("?¼ë³¸ ?Œì‹??ì¢‹ì•„?˜ë©°, ?¤ì–‘??ë¬¸í™”ë¥?ê²½í—˜?˜ëŠ” ?¬í–‰??ì¦ê¹?ˆë‹¤.", bold=False),
    ]
    notion.blocks.children.append(block_id=hobby_col3_id, children=hobby3)
    print("??ì·¨ë? 3??ì»¬ëŸ¼ ?„ë£Œ")
    time.sleep(0.3)
    
    # ===== ?˜ë‹¨ ë©”ì‹œì§€ (?„ì²´ ?ˆë¹„) =====
    footer_blocks = [
        create_paragraph(""),
        create_divider(),
        create_callout("?“ Detailed project information and code samples are available in the portfolio!", "?”—", "green_background"),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=page_id, children=footer_blocks)
    
    print("\n" + "=" * 80)
    print("?‰ ê²½ë ¥ê¸°ìˆ ??v10 ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print("=" * 80)
    
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??- ?¸ì…˜ ?˜ì´ì§€ ?ì„±
ê¸°ì¡´ "ì¡°ë½?? Rakeen Jo" ?˜ì´ì§€?€ ? ì‚¬???ˆì´?„ì›ƒ + ê²½ë ¥ ì¤‘ì‹¬
"""

from notion_client import Client
import time

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {"object": "block", "type": heading_type, heading_type: {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_paragraph(text="", bold=False, color="default"):
    if not text:
        return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}
    return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": text}, "annotations": {"bold": bold}, "color": color}]}}

def create_bulleted_list(text):
    return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_numbered_list(text):
    return {"object": "block", "type": "numbered_list_item", "numbered_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_callout(text, emoji="?’¡"):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": text}}], "icon": {"type": "emoji", "emoji": emoji}}}

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def create_column_list_with_columns(num_columns=2):
    """ì»¬ëŸ¼ ë¦¬ìŠ¤?¸ë? ?ì‹ ì»¬ëŸ¼?¤ê³¼ ?¨ê»˜ ?ì„±"""
    columns = [{"object": "block", "type": "column", "column": {}} for _ in range(num_columns)]
    return {"object": "block", "type": "column_list", "column_list": {"children": columns}}

def main():
    print("=" * 80)
    print("?“„ ì¡°ë½??ê²½ë ¥ê¸°ìˆ ???¸ì…˜ ?˜ì´ì§€ ?ì„±")
    print("=" * 80)
    
    # ???˜ì´ì§€ ?ì„±
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={"title": {"title": [{"type": "text", "text": {"content": "ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v7"}}]}},
        icon={"type": "emoji", "emoji": "?‘¨?ğŸ’?},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1517694712202-14dd9538aa97?w=1200"}}
    )
    page_id = new_page["id"]
    print(f"???˜ì´ì§€ ?ì„±: {page_id}\n")
    
    # ===== PROLOGUE ?¤ë” =====
    print("?“ PROLOGUE ?¹ì…˜ ì¶”ê?...")
    prologue_blocks = [
        create_heading("PROLOGUE ?€?€", 2),
        create_paragraph("")
    ]
    notion.blocks.children.append(block_id=page_id, children=prologue_blocks)
    time.sleep(0.3)
    
    # 2??ì»¬ëŸ¼ (?„ë¡œ???¬ì§„ + ?¸ì‚¬ë§?
    column_list = notion.blocks.children.append(
        block_id=page_id,
        children=[create_column_list_with_columns(2)]
    )
    column_list_id = column_list["results"][0]["id"]
    
    # ?¼ìª½ ì»¬ëŸ¼ ID ê°€?¸ì˜¤ê¸?
    columns = notion.blocks.children.list(block_id=column_list_id)
    left_col_id = columns["results"][0]["id"]
    notion.blocks.children.append(
        block_id=left_col_id,
        children=[
            create_callout("?¬ê¸°???„ë¡œ???¬ì§„ ?½ì…", "?“·")
        ]
    )
    
    # ?¤ë¥¸ìª?ì»¬ëŸ¼ ID ê°€?¸ì˜¤ê¸?
    right_col_id = columns["results"][1]["id"]
    notion.blocks.children.append(
        block_id=right_col_id,
        children=[
            create_heading("?ˆë…•?˜ì„¸???‘", 2),
            create_heading("9?„ê°„ ?„ë² ?”ë“œ ?œìŠ¤??ê°œë°œ ?„ë¬¸?±ì„ ?“ì•„??ì¡°ë½?„ì…?ˆë‹¤!", 2),
            create_paragraph(""),
            create_heading("STM32 MCU ê¸°ë°˜ ?˜ë“œ?¨ì–´ ?¤ê³„ë¶€??FreeRTOS ?Œì›¨??ê°œë°œê¹Œì?", 3),
            create_heading("Full-Stack ?„ë² ?”ë“œ ?œìŠ¤??ê°œë°œ ??Ÿ‰??ë³´ìœ ?˜ê³  ?ˆìŠµ?ˆë‹¤.", 3),
            create_heading("?€ ì§ë¬´???€???´í•´?€ ê³µê° ?¥ë ¥?¼ë¡œ ?¨ê³¼?ì¸ ?‘ì—…???´ë£¨ë©?", 3),
            create_heading("ì£¼ë„?ì¸ ?œë„ë¡?ìµœê³ ???±ê³¼ë¥??´ëŒ?´ë‚´ê³??ˆìŠµ?ˆë‹¤.", 3),
            create_paragraph("")
        ]
    )
    
    print("  ??PROLOGUE ?„ë£Œ\n")
    time.sleep(0.5)
    
    # ===== ë©”ì¸ 3??ì»¬ëŸ¼ =====
    print("?“ ë©”ì¸ ì»¨í…ì¸?(3??ì»¬ëŸ¼) ì¶”ê?...")
    
    main_column_list = notion.blocks.children.append(
        block_id=page_id,
        children=[create_column_list_with_columns(3)]
    )
    main_col_list_id = main_column_list["results"][0]["id"]
    
    # ì»¬ëŸ¼ ID??ê°€?¸ì˜¤ê¸?
    main_columns = notion.blocks.children.list(block_id=main_col_list_id)
    col1_id = main_columns["results"][0]["id"]
    col2_id = main_columns["results"][1]["id"]
    col3_id = main_columns["results"][2]["id"]
    
    col1_content = [
        create_heading("CONTACT ?€?€", 2),
        create_callout("?“§ Email: 92lock@kakao.com", "?“§"),
        create_callout("?“± Mobile: 010-7311-0402", "?“±"),
        create_callout("?™ GitHub: github.com/Rakeen-Jo", "?™"),
        create_callout("?”— ?¬íŠ¸?´ë¦¬?? notion.so/portfolio", "?”—"),
        create_paragraph(""),
        create_paragraph(""),
        
        create_heading("CAREER SUMMARY ?€?€", 2),
        create_callout("ì´?ê²½ë ¥ 9??(2016.11 ~ ?„ì¬)", "?’¼"),
        create_paragraph(""),
        create_bulleted_list("?„ë² ?”ë“œ HW/FW Full-Stack ê°œë°œ"),
        create_bulleted_list("8ê°??´ìƒ ë¶„ì„ ?¥ë¹„ ë©”ì¸ë³´ë“œ ?¤ê³„"),
        create_bulleted_list("STM32/ARM ê¸°ë°˜ ?œí’ˆ ?‘ì‚° ê²½í—˜"),
        create_bulleted_list("HW/FW ?µí•© ?”ë²„ê¹??„ë¬¸"),
        create_paragraph(""),
        create_paragraph(""),
        
        create_heading("EDUCATION ?€?€", 2),
        create_callout("ëª…ì??€?™êµ ?„ìê³µí•™ê³?, "?“"),
        create_bulleted_list("2011.03 ~ 2017.08 ì¡¸ì—…"),
        create_bulleted_list("ê·¼ë¡œ?¥í•™??(?œì„¤ê´€ë¦¬í?)"),
        create_paragraph(""),
        create_heading("ë³‘ì—­", 3),
        create_bulleted_list("?¡êµ° ë³‘ì¥ ë§Œê¸° ?œë? (2012~2013)"),
        create_paragraph(""),
        
        create_heading("CERTIFICATION ?€?€", 2),
        create_bulleted_list("ì»´í“¨?°í™œ?©ëŠ¥??3ê¸?),
        create_bulleted_list("ITQ ?•ë³´ê¸°ìˆ ?ê²© (?œê?/Excel/PPT)"),
        create_bulleted_list("TOEIC Speaking 110??(2016.08)"),
        create_bulleted_list("êµ´ì‚­ê¸??´ì „ê¸°ëŠ¥??(2012.11)"),
    ]
    notion.blocks.children.append(block_id=col1_id, children=col1_content)
    print("  ???¼ìª½ ì»¬ëŸ¼ ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ì¤‘ê°„ ì»¬ëŸ¼: ê²½ë ¥ ?ì„¸ =====
    
    col2_content = [
        create_heading("WORK EXPERIENCE ?€?€", 2),
        create_paragraph(""),
        
        # ATIK (2020~2025)
        create_callout("ATIK (?„í‹±) - ? ì„ ?°êµ¬??, "?¢"),
        create_paragraph("?“… 2020.08 ~ 2025.01 (4??6ê°œì›”)", bold=True),
        create_paragraph("?’¼ ?°êµ¬??/ ë©”ì¸ë³´ë“œ ?Œë¡œ & ?Œì›¨???¤ê³„", bold=True),
        create_paragraph(""),
        create_heading("?´ë‹¹ ?…ë¬´", 3),
        create_bulleted_list("ë¶„ì„ ?¥ë¹„ Main Board ?Œë¡œ ë°??Œì›¨???¤ê³„"),
        create_bulleted_list("PID ?œì–´ ê¸°ë°˜ Pressure Controller ê°œë°œ"),
        create_bulleted_list("FreeRTOS Task ?¤ê³„ ë°?ìµœì ??),
        create_bulleted_list("ë³´ë“œ ê²€?¬ìš© JIG System ë°?Program ?œì‘"),
        create_bulleted_list("Sub-Project Manager (ê°œë°œ ?¼ì • ê´€ë¦?"),
        create_bulleted_list("PCB Artwork ?¤ê³„ (PADS, Altium)"),
        create_bulleted_list("? ê·œ ?¼ì„œ ?°êµ¬ ë°?? ê¸°???¹í—ˆ ê°œë°œ"),
        create_paragraph(""),
        
        create_heading("ì£¼ìš” ?„ë¡œ?íŠ¸", 3),
        create_numbered_list("Psi-1000/3000 Pressure Controller (2024.06~2025.01)"),
        create_numbered_list("LE_Laser (Mantis SSC) ê°œë°œ (2024.01~2024.12)"),
        create_numbered_list("MS (Mass Spectrometer) ê°œë°œ (2023.07~2024.06)"),
        create_numbered_list("JIG System ê°œë°œ (2022.10~2023.06)"),
        create_numbered_list("L-Titrator ?ë„ ë¶„ì„ ?¥ë¹„ (2022.01~2022.12)"),
        create_numbered_list("Nu-2000 (Lux) ê´‘í•™ ?ë„ ë¶„ì„ê¸?(2021.05~2022.10)"),
        create_numbered_list("GasQuantrol ê°€???œì–´ ?¥ë¹„ (2021.01~2021.08)"),
        create_numbered_list("Sigma/Epsilon/L-LPC ?œë¦¬ì¦?(2020.08~2024.12)"),
        create_paragraph(""),
        create_callout("?ì„¸ ?„ë¡œ?íŠ¸ ?´ìš©: ?¬íŠ¸?´ë¦¬??ë§í¬ ì°¸ì¡°", "?”—"),
        create_paragraph(""),
        create_paragraph(""),
        
        # Chemtronics (2019~2020)
        create_callout("Chemtronics (ì¼íŠ¸ë¡œë‹‰?? - ì£¼ì„ ?°êµ¬??, "?¢"),
        create_paragraph("?“… 2019.08 ~ 2020.08 (1??", bold=True),
        create_paragraph("?’¼ ê°€??ê°œë°œ?€ / ?°ì¹˜ ëª¨ë“ˆ ?Œë¡œ & ?Œì›¨???¤ê³„", bold=True),
        create_paragraph(""),
        create_heading("?´ë‹¹ ?…ë¬´", 3),
        create_bulleted_list("?°ì¹˜ ?¼ì„œ ëª¨ë“ˆ ?Œë¡œ, PCB ë°??Œì›¨???¤ê³„"),
        create_bulleted_list("?”ë²„ê¹?ë°?? ë¢°??ê²€??),
        create_bulleted_list("SMT ê³µì • ê´€ë¦?),
        create_bulleted_list("?¹ì¸???‘ì„± ë°??‘ì‚° ê´€ë¦?),
        create_bulleted_list("?„ë“œ ë¶ˆëŸ‰ ?œí’ˆ ë¶„ì„"),
        create_paragraph(""),
        
        create_heading("ì£¼ìš” ?„ë¡œ?íŠ¸", 3),
        create_numbered_list("LG ?¸ë•???¤ë¸/?Œí˜•ê°€??Touch PBA ?‘ì‚°"),
        create_numbered_list("?¼ì„± ?ì–´ì»?Touch PBA ?‘ì‚°"),
        create_numbered_list("Winix ê³µê¸°ì²? •ê¸?Touch PBA ê°œë°œ"),
        create_numbered_list("Haatz ?„ë“œ ê·¼ì ‘?¼ì„œ Touch PBA ê°œë°œ"),
        create_paragraph(""),
        create_paragraph(""),
        
        # Dreamtech (2016~2019)
        create_callout("Dreamtech (?œë¦¼?? - ì£¼ì„ ?°êµ¬??, "?¢"),
        create_paragraph("?“… 2016.11 ~ 2019.05 (2??7ê°œì›”)", bold=True),
        create_paragraph("?’¼ ? í–‰ ?°êµ¬??/ ì§€ë¬?ëª¨ë“ˆ ?Œë¡œ & PCB ?¤ê³„", bold=True),
        create_paragraph(""),
        create_heading("?´ë‹¹ ?…ë¬´", 3),
        create_bulleted_list("ì§€ë¬??¸ì‹ ëª¨ë“ˆ ?Œë¡œ ë°?PCB ?¤ê³„"),
        create_bulleted_list("?”ë²„ê¹?ë°?? ë¢°??ê²€??),
        create_bulleted_list("SMT ê³µì • ê´€ë¦?),
        create_bulleted_list("Chip mount ê³µì • ê°œë°œ ë°??¥ë¹„ ?´ì˜"),
        create_bulleted_list("? í–‰ ê¸°ìˆ  ë¶„ì„ ë°?ê°œë°œ"),
        create_bulleted_list("?•ë?ê³¼ì œ ì§€??(3D-MID ?Œì¬ ê°œë°œ)"),
        create_paragraph(""),
        
        create_heading("ì£¼ìš” ?„ë¡œ?íŠ¸", 3),
        create_numbered_list("?„ë??ë™ì°?ì°¨ëŸ‰??ì§€ë¬¸ì¸??PBA ?‘ì‚°"),
        create_numbered_list("ëª¨ë°”??ì´ˆìŒ??ì§€ë¬¸ì¸??PBA ê°œë°œ"),
        create_numbered_list("FPCç¤?ì§€ë¬¸ì¸??PBA ODM ê°œë°œ"),
        create_numbered_list("3D MID ê¸°ìˆ ê¸°ë°˜ ?„ì¥??ëª¨ë“ˆ ê°œë°œ (?•ë?ê³¼ì œ)"),
    ]
    notion.blocks.children.append(block_id=col2_id, children=col2_content)
    print("  ??ì¤‘ê°„ ì»¬ëŸ¼ ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?¤ë¥¸ìª?ì»¬ëŸ¼: ê¸°ìˆ  ?¤íƒ =====
    
    col3_content = [
        create_heading("SKILLS ?€?€", 2),
        create_paragraph(""),
        
        create_callout("Circuit Design Skills", "??),
        create_numbered_list("STM32F4/F7 MCU ê¸°ë°˜ Main Board ?¤ê³„"),
        create_numbered_list("Switching/Linear Regulator ?„ì›ë¶€ ?¤ê³„"),
        create_numbered_list("16ë¹„íŠ¸ ADC + OP-Amp/TIA ?¼ì„œ ?Œë¡œ ?¤ê³„"),
        create_numbered_list("ê³ ì •ë°€ PID ?œì–´ ?Œë¡œ ?¤ê³„"),
        create_numbered_list("16ë¹„íŠ¸ DAC ?¸ë? ?¸í„°?˜ì´???¤ê³„"),
        create_numbered_list("Photocoupler ?ˆì—° ?Œë¡œ ?¤ê³„"),
        create_numbered_list("Digital I/O ?Œë¡œ ?¤ê³„"),
        create_numbered_list("PCB Artwork (PADS, Altium Designer)"),
        create_paragraph(""),
        
        create_callout("Firmware Design Skills", "?’»"),
        create_numbered_list("STM32F4/F7 MCU ?Œì›¨??ê°œë°œ"),
        create_numbered_list("FreeRTOS Task ê´€ë¦?& ?¤ì‹œê°?ì²˜ë¦¬"),
        create_numbered_list("PID ?œì–´ ?Œê³ ë¦¬ì¦˜ êµ¬í˜„"),
        create_numbered_list("Y-modem IAP Bootloader ê°œë°œ"),
        create_numbered_list("Ethernet TCP/UDP + FreeRTOS"),
        create_numbered_list("UART (RS232/485), SPI, I2C ?µì‹ "),
        create_numbered_list("Touch LCD HMI UI/UX ?¤ê³„"),
        create_numbered_list("Flash ë©”ëª¨ë¦?ìµœì ??),
        create_paragraph(""),
        
        create_callout("Software Development", "?–¥ï¸?),
        create_numbered_list("C# PC ?ŒìŠ¤???„ë¡œê·¸ë¨ ê°œë°œ"),
        create_numbered_list("Arduino/ESP32 ?„ë¡œ? í??´í•‘"),
        create_numbered_list("?¤ì‹œê°?ê²€??ë°??°ì´??ë¡œê¹…"),
        create_paragraph(""),
        
        create_callout("Tools & Technologies", "?› ï¸?),
        create_bulleted_list("Circuit: OrCAD, PADS, KiCad, Altium"),
        create_bulleted_list("Firmware: STM32 CubeIDE, Keil, VS Code"),
        create_bulleted_list("Editor: Source Insight, Visual Studio"),
        create_bulleted_list("Version: Git, Tortoise SVN"),
        create_bulleted_list("MCU: STM32 (F1/F4/G0/H7), Renesas RA6M4"),
        create_bulleted_list("RTOS: FreeRTOS"),
        create_paragraph(""),
        create_paragraph(""),
        
        create_heading("KEY ACHIEVEMENTS ?€?€", 2),
        create_callout("?µì‹¬ ?±ê³¼", "?†"),
        create_bulleted_list("8ê°??´ìƒ ë¶„ì„ ?¥ë¹„ ë©”ì¸ë³´ë“œ ?‘ì‚° ?„ë£Œ"),
        create_bulleted_list("HW/FW ?µí•© ê°œë°œë¡??”ë²„ê¹??œê°„ 40% ?¨ì¶•"),
        create_bulleted_list("?ë™??JIG ê°œë°œë¡?ê²€???œê°„ 60% ?ˆê°"),
        create_bulleted_list("?¹í—ˆ ë°œëª…: ?°ì´??? íš¨???•ì¸ ê¸°ëŠ¥ (2ê±?"),
        create_bulleted_list("FreeRTOS ìµœì ?”ë¡œ ?‘ë‹µ?ë„ 35% ?¥ìƒ"),
        create_bulleted_list("ì½”ë“œ ë¦¬íŒ©? ë§?¼ë¡œ ë©”ëª¨ë¦?25% ?ˆê°"),
        create_paragraph(""),
        
        create_heading("ETC ?€?€", 2),
        create_bulleted_list("?œì´??ICT ê³µëª¨??500?€ ì¤?100?€ ? ì • (2016)"),
        create_bulleted_list("?„ìê³µí•™ê³??™ìƒ???´ì˜êµ?¥ (2015~2016)"),
        create_bulleted_list("?„ìê³µí•™ê³?ì¡±êµ¬ ?™íšŒ ë¶€?Œì¥ (2014~2015)"),
    ]
    notion.blocks.children.append(block_id=col3_id, children=col3_content)
    print("  ???¤ë¥¸ìª?ì»¬ëŸ¼ ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?˜ë‹¨ ë©”ì‹œì§€ =====
    print("?“ ?˜ë‹¨ ë©”ì‹œì§€ ì¶”ê?...")
    footer_blocks = [
        create_divider(),
        create_paragraph(""),
        create_callout("?ì„¸???„ë¡œ?íŠ¸ ?´ìš©?€ ?¬íŠ¸?´ë¦¬???˜ì´ì§€ë¥?ì°¸ê³ ?´ì£¼?¸ìš”!", "?™"),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=page_id, children=footer_blocks)
    
    print("\n" + "=" * 80)
    print("?‰ ê²½ë ¥ê¸°ìˆ ???˜ì´ì§€ ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print("=" * 80)
    
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v7 - ê°„ë‹¨ ë²„ì „ (ì»¬ëŸ¼ ?†ì´ ?œì°¨??êµ¬ì„±)
"""

from notion_client import Client
import time

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {"object": "block", "type": heading_type, heading_type: {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_paragraph(text="", bold=False):
    if not text:
        return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}
    return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": text}, "annotations": {"bold": bold}}]}}

def create_bulleted_list(text):
    return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_callout(text, emoji="?’¡"):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": text}}], "icon": {"type": "emoji", "emoji": emoji}}}

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def main():
    print("=" * 80)
    print("?“„ ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v7 ?ì„±")
    print("=" * 80)
    
    # ???˜ì´ì§€ ?ì„±
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={"title": {"title": [{"type": "text", "text": {"content": "ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v7"}}]}},
        icon={"type": "emoji", "emoji": "?‘¨?ğŸ’?},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1517694712202-14dd9538aa97?w=1200"}}
    )
    page_id = new_page["id"]
    print(f"???˜ì´ì§€ ?ì„±: {page_id}\n")
    
    # ===== ?¤ë” =====
    header_blocks = [
        create_heading("?‘¤ ì¡°ë½??(Rakeen Jo)", 1),
        create_callout("Embedded System Engineer | 9?„ì°¨ | HW/FW Full-Stack", "??),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=header_blocks)
    print("???¤ë” ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ?°ë½ì²?=====
    contact_blocks = [
        create_heading("?“ CONTACT", 2),
        create_bulleted_list("?“§ Email: 92lock@kakao.com"),
        create_bulleted_list("?“± Mobile: 010-7311-0402"),
        create_bulleted_list("?™ GitHub: github.com/Rakeen-Jo"),
        create_bulleted_list("?”— ?¬íŠ¸?´ë¦¬?? [ë§í¬ ì¶”ê? ?ˆì •]"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=contact_blocks)
    print("???°ë½ì²?ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ê²½ë ¥ ?”ì•½ =====
    summary_blocks = [
        create_heading("?’¼ CAREER SUMMARY", 2),
        create_callout("ì´?ê²½ë ¥ 9??(2016.11 ~ 2025.01)", "?“…"),
        create_paragraph(""),
        create_bulleted_list("???„ë² ?”ë“œ HW/FW Full-Stack ê°œë°œ (STM32 ?„ë¬¸)"),
        create_bulleted_list("??8ê°??´ìƒ ë¶„ì„ ?¥ë¹„ ë©”ì¸ë³´ë“œ ?¤ê³„ ë°??‘ì‚°"),
        create_bulleted_list("??FreeRTOS ?¤ì‹œê°??œìŠ¤??ê°œë°œ"),
        create_bulleted_list("??PID ?œì–´, Ethernet ?µì‹ , IAP Bootloader êµ¬í˜„"),
        create_bulleted_list("??HW/FW ?µí•© ?”ë²„ê¹…ìœ¼ë¡?ê°œë°œ ?œê°„ 40% ?¨ì¶•"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=summary_blocks)
    print("??ê²½ë ¥ ?”ì•½ ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ?Œì‚¬ ê²½ë ¥ 1: ATIK =====
    atik_blocks = [
        create_heading("?¢ WORK EXPERIENCE", 2),
        create_paragraph(""),
        create_heading("ATIK (?„í‹±) - ? ì„ ?°êµ¬??, 3),
        create_callout("?“… 2020.08 ~ 2025.01 (4??6ê°œì›”)", "?“…"),
        create_callout("?’¼ ?°êµ¬??/ ë©”ì¸ë³´ë“œ ?Œë¡œ & ?Œì›¨???¤ê³„", "?’¼"),
        create_paragraph(""),
        create_paragraph("?´ë‹¹ ?…ë¬´:", bold=True),
        create_bulleted_list("ë¶„ì„ ?¥ë¹„ Main Board ?Œë¡œ ë°??Œì›¨???¤ê³„"),
        create_bulleted_list("PID ?œì–´ ê¸°ë°˜ Pressure Controller ê°œë°œ"),
        create_bulleted_list("FreeRTOS Task ?¤ê³„ ë°?ìµœì ??),
        create_bulleted_list("ë³´ë“œ ê²€?¬ìš© JIG System ë°?C# Program ?œì‘"),
        create_bulleted_list("Sub-Project Manager (ê°œë°œ ?¼ì • ê´€ë¦?"),
        create_bulleted_list("PCB Artwork ?¤ê³„ (PADS, Altium Designer)"),
        create_bulleted_list("? ê·œ ?¼ì„œ ?°êµ¬ ë°??¹í—ˆ ë°œëª… (2ê±?"),
        create_paragraph(""),
        create_paragraph("ì£¼ìš” ?„ë¡œ?íŠ¸:", bold=True),
        create_bulleted_list("Psi-1000/3000 Pressure Controller (2024.06~2025.01)"),
        create_bulleted_list("LE_Laser (Mantis SSC) ê°œë°œ (2024.01~2024.12)"),
        create_bulleted_list("MS (Mass Spectrometer) ê°œë°œ (2023.07~2024.06)"),
        create_bulleted_list("JIG System ê°œë°œ (2022.10~2023.06)"),
        create_bulleted_list("L-Titrator ?ë„ ë¶„ì„ ?¥ë¹„ (2022.01~2022.12)"),
        create_bulleted_list("Nu-2000 (Lux) ê´‘í•™ ?ë„ ë¶„ì„ê¸?(2021.05~2022.10)"),
        create_bulleted_list("GasQuantrol ê°€???œì–´ ?¥ë¹„ (2021.01~2021.08)"),
        create_bulleted_list("Sigma/Epsilon/L-LPC ?œë¦¬ì¦?(2020.08~2024.12)"),
        create_paragraph(""),
        create_callout("?“Š ?ì„¸ ?„ë¡œ?íŠ¸ ?´ìš©: ?¬íŠ¸?´ë¦¬??ë§í¬ ì°¸ì¡°", "?”—"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=atik_blocks)
    print("??ATIK ê²½ë ¥ ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ?Œì‚¬ ê²½ë ¥ 2: Chemtronics =====
    chem_blocks = [
        create_heading("Chemtronics (ì¼íŠ¸ë¡œë‹‰?? - ì£¼ì„ ?°êµ¬??, 3),
        create_callout("?“… 2019.08 ~ 2020.08 (1??", "?“…"),
        create_callout("?’¼ ê°€??ê°œë°œ?€ / ?°ì¹˜ ëª¨ë“ˆ ?Œë¡œ & ?Œì›¨???¤ê³„", "?’¼"),
        create_paragraph(""),
        create_paragraph("?´ë‹¹ ?…ë¬´:", bold=True),
        create_bulleted_list("?°ì¹˜ ?¼ì„œ ëª¨ë“ˆ ?Œë¡œ, PCB ë°??Œì›¨???¤ê³„"),
        create_bulleted_list("?”ë²„ê¹?ë°?? ë¢°??ê²€??),
        create_bulleted_list("SMT ê³µì • ê´€ë¦?),
        create_bulleted_list("?¹ì¸???‘ì„± ë°??‘ì‚° ê´€ë¦?),
        create_bulleted_list("?„ë“œ ë¶ˆëŸ‰ ?œí’ˆ ë¶„ì„"),
        create_paragraph(""),
        create_paragraph("ì£¼ìš” ?„ë¡œ?íŠ¸:", bold=True),
        create_bulleted_list("LG ?¸ë•???¤ë¸/?Œí˜•ê°€??Touch PBA ?‘ì‚°"),
        create_bulleted_list("?¼ì„± ?ì–´ì»?Touch PBA ?‘ì‚°"),
        create_bulleted_list("Winix ê³µê¸°ì²? •ê¸?Touch PBA ê°œë°œ"),
        create_bulleted_list("Haatz ?„ë“œ ê·¼ì ‘?¼ì„œ Touch PBA ê°œë°œ"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=chem_blocks)
    print("??Chemtronics ê²½ë ¥ ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ?Œì‚¬ ê²½ë ¥ 3: Dreamtech =====
    dream_blocks = [
        create_heading("Dreamtech (?œë¦¼?? - ì£¼ì„ ?°êµ¬??, 3),
        create_callout("?“… 2016.11 ~ 2019.05 (2??7ê°œì›”)", "?“…"),
        create_callout("?’¼ ? í–‰ ?°êµ¬??/ ì§€ë¬?ëª¨ë“ˆ ?Œë¡œ & PCB ?¤ê³„", "?’¼"),
        create_paragraph(""),
        create_paragraph("?´ë‹¹ ?…ë¬´:", bold=True),
        create_bulleted_list("ì§€ë¬??¸ì‹ ëª¨ë“ˆ ?Œë¡œ ë°?PCB ?¤ê³„ (RFPCB)"),
        create_bulleted_list("?”ë²„ê¹?ë°?? ë¢°??ê²€??),
        create_bulleted_list("SMT ë°?Chip mount ê³µì • ê´€ë¦?),
        create_bulleted_list("? í–‰ ê¸°ìˆ  ë¶„ì„ ë°?ê°œë°œ"),
        create_bulleted_list("?•ë?ê³¼ì œ ì§€??(3D-MID ?Œì¬ ê°œë°œ)"),
        create_paragraph(""),
        create_paragraph("ì£¼ìš” ?„ë¡œ?íŠ¸:", bold=True),
        create_bulleted_list("?„ë??ë™ì°?ì°¨ëŸ‰??ì§€ë¬¸ì¸??PBA ?‘ì‚°"),
        create_bulleted_list("ëª¨ë°”??ì´ˆìŒ??ì§€ë¬¸ì¸??PBA ê°œë°œ"),
        create_bulleted_list("FPCç¤?ì§€ë¬¸ì¸??PBA ODM ê°œë°œ"),
        create_bulleted_list("3D MID ê¸°ìˆ ê¸°ë°˜ ?„ì¥??ëª¨ë“ˆ ê°œë°œ (?•ë?ê³¼ì œ)"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=dream_blocks)
    print("??Dreamtech ê²½ë ¥ ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ê¸°ìˆ  ?¤íƒ =====
    skills_blocks = [
        create_heading("?› ï¸?SKILLS", 2),
        create_paragraph(""),
        create_heading("Circuit Design", 3),
        create_bulleted_list("STM32F4/F7 MCU ê¸°ë°˜ Main Board ?¤ê³„"),
        create_bulleted_list("Switching/Linear Regulator ?„ì›ë¶€ ?¤ê³„"),
        create_bulleted_list("16ë¹„íŠ¸ ADC + OP-Amp/TIA ?¼ì„œ ?Œë¡œ ?¤ê³„"),
        create_bulleted_list("ê³ ì •ë°€ PID ?œì–´ ?Œë¡œ ?¤ê³„"),
        create_bulleted_list("16ë¹„íŠ¸ DAC ?¸ë? ?¸í„°?˜ì´???¤ê³„"),
        create_bulleted_list("Photocoupler ?ˆì—° ?Œë¡œ ?¤ê³„"),
        create_bulleted_list("Digital I/O ?Œë¡œ ?¤ê³„"),
        create_bulleted_list("PCB Artwork (PADS, Altium Designer)"),
        create_paragraph(""),
        
        create_heading("Firmware Design", 3),
        create_bulleted_list("STM32F4/F7 MCU ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("FreeRTOS Task ê´€ë¦?& ?¤ì‹œê°?ì²˜ë¦¬"),
        create_bulleted_list("PID ?œì–´ ?Œê³ ë¦¬ì¦˜ êµ¬í˜„"),
        create_bulleted_list("Y-modem IAP Bootloader ê°œë°œ"),
        create_bulleted_list("Ethernet TCP/UDP + FreeRTOS"),
        create_bulleted_list("UART (RS232/485), SPI, I2C ?µì‹ "),
        create_bulleted_list("Touch LCD HMI UI/UX ?¤ê³„ (Nextion, TouchGFX)"),
        create_bulleted_list("Flash ë©”ëª¨ë¦?ìµœì ??),
        create_paragraph(""),
        
        create_heading("Software & Tools", 3),
        create_bulleted_list("C# PC ?ŒìŠ¤???„ë¡œê·¸ë¨ ê°œë°œ"),
        create_bulleted_list("Arduino/ESP32 ?„ë¡œ? í??´í•‘"),
        create_bulleted_list("Circuit: OrCAD, PADS, KiCad, Altium"),
        create_bulleted_list("Firmware: STM32 CubeIDE, Keil, VS Code"),
        create_bulleted_list("Version: Git, Tortoise SVN"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=skills_blocks)
    print("??ê¸°ìˆ  ?¤íƒ ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ?µì‹¬ ?±ê³¼ =====
    achievement_blocks = [
        create_heading("?† KEY ACHIEVEMENTS", 2),
        create_bulleted_list("8ê°??´ìƒ ë¶„ì„ ?¥ë¹„ ë©”ì¸ë³´ë“œ ?‘ì‚° ?„ë£Œ"),
        create_bulleted_list("HW/FW ?µí•© ê°œë°œë¡??”ë²„ê¹??œê°„ 40% ?¨ì¶•"),
        create_bulleted_list("?ë™??JIG ê°œë°œë¡?ê²€???œê°„ 60% ?ˆê°"),
        create_bulleted_list("?¹í—ˆ ë°œëª…: ?°ì´??? íš¨???•ì¸ ê¸°ëŠ¥ (2ê±?ì¶œì›)"),
        create_bulleted_list("FreeRTOS ìµœì ?”ë¡œ ?‘ë‹µ?ë„ 35% ?¥ìƒ"),
        create_bulleted_list("ì½”ë“œ ë¦¬íŒ©? ë§?¼ë¡œ ë©”ëª¨ë¦??¬ìš©??25% ?ˆê°"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=achievement_blocks)
    print("???µì‹¬ ?±ê³¼ ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ?™ë ¥ & ?ê²©ì¦?=====
    edu_blocks = [
        create_heading("?“ EDUCATION", 2),
        create_callout("ëª…ì??€?™êµ ?„ìê³µí•™ê³?ì¡¸ì—…", "?“"),
        create_bulleted_list("2011.03 ~ 2017.08"),
        create_bulleted_list("ê·¼ë¡œ?¥í•™??(?œì„¤ê´€ë¦¬í? ê·¼ë¬´)"),
        create_paragraph(""),
        create_paragraph("ë³‘ì—­:", bold=True),
        create_bulleted_list("?¡êµ° ???¼ì „?˜ì†¡êµìœ¡??ë³‘ì¥ ë§Œê¸° ?œë? (2012~2013)"),
        create_paragraph(""),
        create_divider(),
        
        create_heading("?“œ CERTIFICATION", 2),
        create_bulleted_list("ì»´í“¨?°í™œ?©ëŠ¥??3ê¸?),
        create_bulleted_list("ITQ ?•ë³´ê¸°ìˆ ?ê²© (?œê?B/?‘ì?B/PPT-A)"),
        create_bulleted_list("TOEIC Speaking 110??(2016.08)"),
        create_bulleted_list("êµ´ì‚­ê¸??´ì „ê¸°ëŠ¥??(2012.11)"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=edu_blocks)
    print("???™ë ¥ & ?ê²©ì¦?ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ê¸°í? ?œë™ =====
    etc_blocks = [
        create_heading("?¯ ETC", 2),
        create_bulleted_list("?œì´??ICT ê³µëª¨??500?€ ì¤?100?€ ? ì • ë°??„ì‹œ??ì°¸ê? (2016)"),
        create_bulleted_list("?„ìê³µí•™ê³??™ìƒ???´ì˜êµ?¥ (2015~2016)"),
        create_bulleted_list("?„ìê³µí•™ê³?ì¡±êµ¬ ?™íšŒ ë¶€?Œì¥ (2014~2015)"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=etc_blocks)
    print("??ê¸°í? ?œë™ ì¶”ê?")
    time.sleep(0.3)
    
    # ===== ?˜ë‹¨ ë©”ì‹œì§€ =====
    footer_blocks = [
        create_paragraph(""),
        create_callout("?ì„¸???„ë¡œ?íŠ¸ ?´ìš© ë°?ì½”ë“œ ?˜í”Œ?€ ?¬íŠ¸?´ë¦¬?¤ë? ì°¸ê³ ?´ì£¼?¸ìš”! ?™", "?“"),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=page_id, children=footer_blocks)
    
    print("\n" + "=" * 80)
    print("?‰ ê²½ë ¥ê¸°ìˆ ???˜ì´ì§€ ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print("=" * 80)
    
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v8 - ê¹”ë”???”ì??+ ? ê?ë¡??ì„¸ ?•ë³´ ?‘ê¸°
"""

from notion_client import Client
import time

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {"object": "block", "type": heading_type, heading_type: {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_paragraph(text="", bold=False):
    if not text:
        return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}
    return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": text}, "annotations": {"bold": bold}}]}}

def create_bulleted_list(text):
    return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_callout(text, emoji="?’¡", color="default"):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": text}}], "icon": {"type": "emoji", "emoji": emoji}, "color": color}}

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def create_toggle(title, emoji="?“"):
    return {
        "object": "block",
        "type": "toggle",
        "toggle": {
            "rich_text": [{"type": "text", "text": {"content": f"{emoji} {title}"}}],
            "children": []
        }
    }

def main():
    print("=" * 80)
    print("?“„ ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v8 - ê¹”ë”???”ì??)
    print("=" * 80)
    
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={"title": {"title": [{"type": "text", "text": {"content": "ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v8"}}]}},
        icon={"type": "emoji", "emoji": "?‘¨?ğŸ’?},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1517694712202-14dd9538aa97?w=1200"}}
    )
    page_id = new_page["id"]
    print(f"???˜ì´ì§€ ?ì„±: {page_id}\n")
    
    # ===== ?¤ë” =====
    header_blocks = [
        create_heading("ì¡°ë½??(Rakeen Jo)", 1),
        create_callout("Embedded System Engineer | 9?„ì°¨ (2016~?„ì¬) | HW/FW Full-Stack", "??, "blue_background"),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=page_id, children=header_blocks)
    print("???¤ë”")
    time.sleep(0.3)
    
    # ===== ?°ë½ì²?& ë§í¬ =====
    contact_blocks = [
        create_callout("?“§ 92lock@kakao.com  |  ?“± 010-7311-0402  |  ?™ github.com/Rakeen-Jo", "?“", "gray_background"),
        create_callout("?”— ?¬íŠ¸?´ë¦¬?? [ë§í¬ ì¶”ê? ?ˆì •]", "?”—", "gray_background"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=contact_blocks)
    print("???°ë½ì²?)
    time.sleep(0.3)
    
    # ===== ?µì‹¬ ??Ÿ‰ ?”ì•½ =====
    summary_blocks = [
        create_heading("?’¼ ?µì‹¬ ??Ÿ‰", 2),
        create_callout("9?„ê°„ STM32 ê¸°ë°˜ ?„ë² ?”ë“œ ?œìŠ¤??Full-Stack ê°œë°œ ?„ë¬¸", "?¯", "yellow_background"),
        create_paragraph(""),
        create_bulleted_list("??8ê°??´ìƒ ë¶„ì„ ?¥ë¹„ ë©”ì¸ë³´ë“œ ?¤ê³„ ë°??‘ì‚° ?„ë£Œ"),
        create_bulleted_list("??HW/FW ?µí•© ê°œë°œë¡??”ë²„ê¹??¨ìœ¨ 40% ?¥ìƒ"),
        create_bulleted_list("??FreeRTOS, PID ?œì–´, Ethernet ?µì‹  ?„ë¬¸"),
        create_bulleted_list("???ë™???ŒìŠ¤??JIG ê°œë°œë¡?ê²€???œê°„ 60% ?¨ì¶•"),
        create_bulleted_list("???¹í—ˆ ë°œëª… 2ê±?ì¶œì›"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=summary_blocks)
    print("???µì‹¬ ??Ÿ‰")
    time.sleep(0.3)
    
    # ===== ?Œì‚¬ ê²½ë ¥ (? ê? ë°©ì‹) =====
    print("???Œì‚¬ ê²½ë ¥ ì¶”ê? ì¤?..")
    
    # ATIK
    atik_toggle = notion.blocks.children.append(
        block_id=page_id,
        children=[create_heading("?¢ ê²½ë ¥", 2)]
    )
    time.sleep(0.2)
    
    atik_header = [
        create_callout("ATIK (?„í‹±) - ? ì„ ?°êµ¬??, "?¢", "blue_background"),
        create_paragraph("?“… 2020.08 ~ 2025.01 (4??6ê°œì›”) | ?’¼ ?°êµ¬??/ ë©”ì¸ë³´ë“œ HW/FW ?¤ê³„", bold=True),
    ]
    notion.blocks.children.append(block_id=page_id, children=atik_header)
    time.sleep(0.2)
    
    # ATIK ì£¼ìš” ?„ë¡œ?íŠ¸ ? ê?
    atik_proj_toggle = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("ì£¼ìš” ?„ë¡œ?íŠ¸ 8ê°?(?´ë¦­?˜ì—¬ ?¼ì¹˜ê¸?", "?“‚")]
    )
    atik_proj_id = atik_proj_toggle["results"][0]["id"]
    
    atik_proj_content = [
        create_bulleted_list("Psi-1000/3000: ?•ë? ê°€???•ë ¥ ?œì–´ ?œìŠ¤??(2024)"),
        create_bulleted_list("LE_Laser (Mantis SSC): ?ˆì´?€ ?Œí‹°??ë¶„ì„ê¸?(2024)"),
        create_bulleted_list("MS (Mass Spectrometer): ì§ˆëŸ‰ë¶„ì„ê¸?DRV Board (2023)"),
        create_bulleted_list("JIG System: ?ë™??ë³´ë“œ ?ŒìŠ¤???¥ë¹„ + C# ?„ë¡œê·¸ë¨ (2022~2023)"),
        create_bulleted_list("L-Titrator: ?ì •???ë„ ë¶„ì„ ?¥ë¹„ (2022)"),
        create_bulleted_list("Nu-2000 (Lux): UV+IR ê´‘í•™ ?ë„ ë¶„ì„ê¸?(2021~2022)"),
        create_bulleted_list("GasQuantrol: ê°€???œì–´ ?¥ë¹„ (2021)"),
        create_bulleted_list("Sigma/Epsilon/L-LPC ?œë¦¬ì¦? COD/?„ë„??ë¶„ì„ (2020~2024)"),
    ]
    notion.blocks.children.append(block_id=atik_proj_id, children=atik_proj_content)
    print("  ??ATIK")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=page_id, children=[create_paragraph("")])
    
    # Chemtronics
    chem_blocks = [
        create_callout("Chemtronics (ì¼íŠ¸ë¡œë‹‰?? - ì£¼ì„ ?°êµ¬??, "?¢", "blue_background"),
        create_paragraph("?“… 2019.08 ~ 2020.08 (1?? | ?’¼ ê°€??ê°œë°œ?€ / ?°ì¹˜ ëª¨ë“ˆ HW/FW ?¤ê³„", bold=True),
    ]
    notion.blocks.children.append(block_id=page_id, children=chem_blocks)
    time.sleep(0.2)
    
    chem_proj_toggle = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("ì£¼ìš” ?„ë¡œ?íŠ¸ 4ê°?(?´ë¦­?˜ì—¬ ?¼ì¹˜ê¸?", "?“‚")]
    )
    chem_proj_id = chem_proj_toggle["results"][0]["id"]
    
    chem_proj_content = [
        create_bulleted_list("LG ?¸ë•???¤ë¸/?Œí˜•ê°€??Touch PBA ?‘ì‚°"),
        create_bulleted_list("?¼ì„± ?ì–´ì»?Touch PBA ?‘ì‚°"),
        create_bulleted_list("Winix ê³µê¸°ì²? •ê¸?Touch PBA ê°œë°œ"),
        create_bulleted_list("Haatz ?„ë“œ ê·¼ì ‘?¼ì„œ Touch PBA ê°œë°œ"),
    ]
    notion.blocks.children.append(block_id=chem_proj_id, children=chem_proj_content)
    print("  ??Chemtronics")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=page_id, children=[create_paragraph("")])
    
    # Dreamtech
    dream_blocks = [
        create_callout("Dreamtech (?œë¦¼?? - ì£¼ì„ ?°êµ¬??, "?¢", "blue_background"),
        create_paragraph("?“… 2016.11 ~ 2019.05 (2??7ê°œì›”) | ?’¼ ? í–‰ ?°êµ¬??/ ì§€ë¬?ëª¨ë“ˆ HW ?¤ê³„", bold=True),
    ]
    notion.blocks.children.append(block_id=page_id, children=dream_blocks)
    time.sleep(0.2)
    
    dream_proj_toggle = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("ì£¼ìš” ?„ë¡œ?íŠ¸ 4ê°?(?´ë¦­?˜ì—¬ ?¼ì¹˜ê¸?", "?“‚")]
    )
    dream_proj_id = dream_proj_toggle["results"][0]["id"]
    
    dream_proj_content = [
        create_bulleted_list("?„ë??ë™ì°?ì°¨ëŸ‰??ì§€ë¬¸ì¸??PBA ?‘ì‚°"),
        create_bulleted_list("ëª¨ë°”??ì´ˆìŒ??ì§€ë¬¸ì¸??PBA ê°œë°œ"),
        create_bulleted_list("FPCç¤?ì§€ë¬¸ì¸??PBA ODM ê°œë°œ"),
        create_bulleted_list("3D MID ê¸°ìˆ ê¸°ë°˜ ?„ì¥??ëª¨ë“ˆ ê°œë°œ (?•ë?ê³¼ì œ)"),
    ]
    notion.blocks.children.append(block_id=dream_proj_id, children=dream_proj_content)
    print("  ??Dreamtech")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=page_id, children=[create_paragraph(""), create_divider()])
    
    # ===== ê¸°ìˆ  ?¤íƒ (? ê?) =====
    print("??ê¸°ìˆ  ?¤íƒ ì¶”ê? ì¤?..")
    
    notion.blocks.children.append(block_id=page_id, children=[create_heading("?› ï¸?ê¸°ìˆ  ?¤íƒ", 2)])
    
    # Circuit Design ? ê?
    circuit_toggle = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("Circuit Design (?Œë¡œ ?¤ê³„)", "??)]
    )
    circuit_id = circuit_toggle["results"][0]["id"]
    
    circuit_content = [
        create_bulleted_list("STM32F4/F7 MCU ê¸°ë°˜ Main Board ?¤ê³„"),
        create_bulleted_list("Switching/Linear Regulator ?„ì›ë¶€ ?¤ê³„"),
        create_bulleted_list("16ë¹„íŠ¸ ADC + OP-Amp/TIA ?¼ì„œ ?Œë¡œ ?¤ê³„"),
        create_bulleted_list("16ë¹„íŠ¸ DAC ?¸ë? ?¸í„°?˜ì´???¤ê³„"),
        create_bulleted_list("Photocoupler ?ˆì—° ?Œë¡œ ?¤ê³„"),
        create_bulleted_list("PCB Artwork (PADS, Altium Designer)"),
    ]
    notion.blocks.children.append(block_id=circuit_id, children=circuit_content)
    time.sleep(0.2)
    
    # Firmware Design ? ê?
    firmware_toggle = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("Firmware Design (?Œì›¨???¤ê³„)", "?’»")]
    )
    firmware_id = firmware_toggle["results"][0]["id"]
    
    firmware_content = [
        create_bulleted_list("STM32F4/F7 MCU ?Œì›¨??ê°œë°œ (C)"),
        create_bulleted_list("FreeRTOS Task ê´€ë¦?& ?¤ì‹œê°?ì²˜ë¦¬"),
        create_bulleted_list("PID ?œì–´ ?Œê³ ë¦¬ì¦˜ êµ¬í˜„"),
        create_bulleted_list("Y-modem IAP Bootloader ê°œë°œ"),
        create_bulleted_list("Ethernet TCP/UDP + FreeRTOS"),
        create_bulleted_list("UART (RS232/485), SPI, I2C ?µì‹ "),
        create_bulleted_list("Touch LCD HMI UI/UX (Nextion, TouchGFX)"),
    ]
    notion.blocks.children.append(block_id=firmware_id, children=firmware_content)
    time.sleep(0.2)
    
    # Tools ? ê?
    tools_toggle = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("Tools & Software", "?–¥ï¸?)]
    )
    tools_id = tools_toggle["results"][0]["id"]
    
    tools_content = [
        create_bulleted_list("Circuit: OrCAD, PADS, KiCad, Altium Designer"),
        create_bulleted_list("Firmware: STM32 CubeIDE, Keil, VS Code, Source Insight"),
        create_bulleted_list("Software: C# (PC ?ŒìŠ¤???„ë¡œê·¸ë¨), Arduino/ESP32"),
        create_bulleted_list("Version: Git, Tortoise SVN"),
    ]
    notion.blocks.children.append(block_id=tools_id, children=tools_content)
    print("  ??ê¸°ìˆ  ?¤íƒ ?„ë£Œ")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=page_id, children=[create_paragraph(""), create_divider()])
    
    # ===== ?™ë ¥ & ?ê²©ì¦?(ê°„ê²°?˜ê²Œ) =====
    edu_blocks = [
        create_heading("?“ ?™ë ¥ & ?ê²©", 2),
        create_callout("ëª…ì??€?™êµ ?„ìê³µí•™ê³?ì¡¸ì—… (2011~2017)", "?“", "gray_background"),
        create_callout("ë³‘ì—­: ?¡êµ° ë³‘ì¥ ë§Œê¸° ?œë? (2012~2013)", "?ª–", "gray_background"),
        create_paragraph(""),
        create_callout("?ê²©ì¦? ì»´í“¨?°í™œ?©ëŠ¥??3ê¸? ITQ, TOEIC Speaking 110?? êµ´ì‚­ê¸??´ì „ê¸°ëŠ¥??, "?“œ", "gray_background"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=edu_blocks)
    print("???™ë ¥ & ?ê²©")
    time.sleep(0.3)
    
    # ===== ?˜ë‹¨ ë©”ì‹œì§€ =====
    footer_blocks = [
        create_paragraph(""),
        create_callout("?“ ?ì„¸ ?„ë¡œ?íŠ¸ ?´ìš© ë°?ì½”ë“œ ?˜í”Œ?€ ?¬íŠ¸?´ë¦¬?¤ë? ì°¸ê³ ?´ì£¼?¸ìš”!", "?”—", "green_background"),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=page_id, children=footer_blocks)
    
    print("\n" + "=" * 80)
    print("?‰ ê²½ë ¥ê¸°ìˆ ??v8 ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print("=" * 80)
    
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v9 - 2??ì»¬ëŸ¼ ?ˆì´?„ì›ƒ (ì¢? ê²½ë ¥, ?? ê¸°ìˆ ?¤íƒ)
"""

from notion_client import Client
import time

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {"object": "block", "type": heading_type, heading_type: {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_paragraph(text="", bold=False):
    if not text:
        return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}
    return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": text}, "annotations": {"bold": bold}}]}}

def create_bulleted_list(text):
    return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_callout(text, emoji="?’¡", color="default"):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": text}}], "icon": {"type": "emoji", "emoji": emoji}, "color": color}}

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def create_toggle(title, emoji="?“"):
    return {
        "object": "block",
        "type": "toggle",
        "toggle": {
            "rich_text": [{"type": "text", "text": {"content": f"{emoji} {title}"}}],
            "children": []
        }
    }

def create_column_list_2col():
    """2ê°?ì»¬ëŸ¼??ê°€ì§?column_list ?ì„± (ê°?ì»¬ëŸ¼??ë¹?paragraph ?„ìˆ˜)"""
    return {
        "object": "block",
        "type": "column_list",
        "column_list": {
            "children": [
                {
                    "object": "block", 
                    "type": "column", 
                    "column": {
                        "children": [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}]
                    }
                },
                {
                    "object": "block", 
                    "type": "column", 
                    "column": {
                        "children": [{"object": "block", "type": "paragraph", "paragraph": {"rich_text": []}}]
                    }
                }
            ]
        }
    }

def main():
    print("=" * 80)
    print("?“„ ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v9 - 2??ì»¬ëŸ¼ ?ˆì´?„ì›ƒ")
    print("=" * 80)
    
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={"title": {"title": [{"type": "text", "text": {"content": "ì¡°ë½??ê²½ë ¥ê¸°ìˆ ??v9"}}]}},
        icon={"type": "emoji", "emoji": "?‘¨?ğŸ’?},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1517694712202-14dd9538aa97?w=1200"}}
    )
    page_id = new_page["id"]
    print(f"???˜ì´ì§€ ?ì„±: {page_id}\n")
    
    # ===== ?¤ë” (?„ì²´ ?ˆë¹„) =====
    header_blocks = [
        create_heading("ì¡°ë½??(Rakeen Jo)", 1),
        create_callout("Embedded System Engineer | 9?„ì°¨ (2016~?„ì¬) | HW/FW Full-Stack", "??, "blue_background"),
        create_paragraph(""),
        create_callout("?“§ 92lock@kakao.com  |  ?“± 010-7311-0402  |  ?™ github.com/Rakeen-Jo", "?“", "gray_background"),
        create_callout("?”— ?¬íŠ¸?´ë¦¬?? [ë§í¬ ì¶”ê? ?ˆì •]", "?”—", "gray_background"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=header_blocks)
    print("???¤ë”")
    time.sleep(0.3)
    
    # ===== ?µì‹¬ ??Ÿ‰ (?„ì²´ ?ˆë¹„) =====
    summary_blocks = [
        create_heading("?’¼ ?µì‹¬ ??Ÿ‰", 2),
        create_callout("9?„ê°„ STM32 ê¸°ë°˜ ?„ë² ?”ë“œ ?œìŠ¤??Full-Stack ê°œë°œ ?„ë¬¸", "?¯", "yellow_background"),
        create_paragraph(""),
        create_bulleted_list("??8ê°??´ìƒ ë¶„ì„ ?¥ë¹„ ë©”ì¸ë³´ë“œ ?¤ê³„ ë°??‘ì‚° ?„ë£Œ"),
        create_bulleted_list("??HW/FW ?µí•© ê°œë°œë¡??”ë²„ê¹??¨ìœ¨ 40% ?¥ìƒ"),
        create_bulleted_list("??FreeRTOS, PID ?œì–´, Ethernet ?µì‹  ?„ë¬¸"),
        create_bulleted_list("???ë™???ŒìŠ¤??JIG ê°œë°œë¡?ê²€???œê°„ 60% ?¨ì¶•"),
        create_bulleted_list("???¹í—ˆ ë°œëª… 2ê±?ì¶œì›"),
        create_paragraph(""),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=summary_blocks)
    print("???µì‹¬ ??Ÿ‰")
    time.sleep(0.3)
    
    # ===== 2??ì»¬ëŸ¼ ?ˆì´?„ì›ƒ ?ì„± =====
    print("??2??ì»¬ëŸ¼ ?ì„± ì¤?..")
    column_list_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_column_list_2col()]
    )
    column_list_id = column_list_response["results"][0]["id"]
    
    # ì»¬ëŸ¼ ID ê°€?¸ì˜¤ê¸?
    time.sleep(0.3)
    columns = notion.blocks.children.list(block_id=column_list_id)
    left_col_id = columns["results"][0]["id"]
    right_col_id = columns["results"][1]["id"]
    print(f"  ???¼ìª½ ì»¬ëŸ¼ ID: {left_col_id}")
    print(f"  ???¤ë¥¸ìª?ì»¬ëŸ¼ ID: {right_col_id}")
    
    # ===== ?¼ìª½ ì»¬ëŸ¼: ê²½ë ¥ =====
    print("\n?“ ?¼ìª½ ì»¬ëŸ¼: ê²½ë ¥ ì¶”ê? ì¤?..")
    left_content = [
        create_heading("?¢ ê²½ë ¥", 2),
        create_paragraph(""),
        
        # ATIK
        create_callout("ATIK (?„í‹±) - ? ì„ ?°êµ¬??, "?¢", "blue_background"),
        create_paragraph("?“… 2020.08 ~ 2025.01 (4??6ê°œì›”)", bold=True),
        create_paragraph("?’¼ ?°êµ¬??/ ë©”ì¸ë³´ë“œ HW/FW ?¤ê³„", bold=True),
    ]
    notion.blocks.children.append(block_id=left_col_id, children=left_content)
    time.sleep(0.3)
    
    # ATIK ?„ë¡œ?íŠ¸ ? ê?
    atik_toggle = notion.blocks.children.append(
        block_id=left_col_id,
        children=[create_toggle("ì£¼ìš” ?„ë¡œ?íŠ¸ 8ê°?, "?“‚")]
    )
    atik_toggle_id = atik_toggle["results"][0]["id"]
    
    atik_proj = [
        create_bulleted_list("Psi-1000/3000: ?•ë? ê°€???•ë ¥ ?œì–´ ?œìŠ¤??),
        create_bulleted_list("LE_Laser: ?ˆì´?€ ?Œí‹°??ë¶„ì„ê¸?),
        create_bulleted_list("MS: ì§ˆëŸ‰ë¶„ì„ê¸?DRV Board"),
        create_bulleted_list("JIG System: ?ë™??ë³´ë“œ ?ŒìŠ¤???¥ë¹„"),
        create_bulleted_list("L-Titrator: ?ì •???ë„ ë¶„ì„ ?¥ë¹„"),
        create_bulleted_list("Nu-2000: UV+IR ê´‘í•™ ?ë„ ë¶„ì„ê¸?),
        create_bulleted_list("GasQuantrol: ê°€???œì–´ ?¥ë¹„"),
        create_bulleted_list("Sigma/Epsilon/L-LPC ?œë¦¬ì¦?),
    ]
    notion.blocks.children.append(block_id=atik_toggle_id, children=atik_proj)
    print("  ??ATIK")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=left_col_id, children=[create_paragraph("")])
    
    # Chemtronics
    chem_content = [
        create_callout("Chemtronics - ì£¼ì„ ?°êµ¬??, "?¢", "blue_background"),
        create_paragraph("?“… 2019.08 ~ 2020.08 (1??", bold=True),
        create_paragraph("?’¼ ê°€??ê°œë°œ?€ / ?°ì¹˜ ëª¨ë“ˆ", bold=True),
    ]
    notion.blocks.children.append(block_id=left_col_id, children=chem_content)
    time.sleep(0.3)
    
    chem_toggle = notion.blocks.children.append(
        block_id=left_col_id,
        children=[create_toggle("ì£¼ìš” ?„ë¡œ?íŠ¸ 4ê°?, "?“‚")]
    )
    chem_toggle_id = chem_toggle["results"][0]["id"]
    
    chem_proj = [
        create_bulleted_list("LG ?¸ë•???¤ë¸ Touch PBA ?‘ì‚°"),
        create_bulleted_list("?¼ì„± ?ì–´ì»?Touch PBA ?‘ì‚°"),
        create_bulleted_list("Winix ê³µê¸°ì²? •ê¸?Touch PBA"),
        create_bulleted_list("Haatz ?„ë“œ ê·¼ì ‘?¼ì„œ Touch PBA"),
    ]
    notion.blocks.children.append(block_id=chem_toggle_id, children=chem_proj)
    print("  ??Chemtronics")
    time.sleep(0.3)
    
    notion.blocks.children.append(block_id=left_col_id, children=[create_paragraph("")])
    
    # Dreamtech
    dream_content = [
        create_callout("Dreamtech - ì£¼ì„ ?°êµ¬??, "?¢", "blue_background"),
        create_paragraph("?“… 2016.11 ~ 2019.05 (2??7ê°œì›”)", bold=True),
        create_paragraph("?’¼ ? í–‰ ?°êµ¬??/ ì§€ë¬?ëª¨ë“ˆ", bold=True),
    ]
    notion.blocks.children.append(block_id=left_col_id, children=dream_content)
    time.sleep(0.3)
    
    dream_toggle = notion.blocks.children.append(
        block_id=left_col_id,
        children=[create_toggle("ì£¼ìš” ?„ë¡œ?íŠ¸ 4ê°?, "?“‚")]
    )
    dream_toggle_id = dream_toggle["results"][0]["id"]
    
    dream_proj = [
        create_bulleted_list("?„ë?ì°?ì§€ë¬¸ì¸??PBA ?‘ì‚°"),
        create_bulleted_list("ëª¨ë°”??ì´ˆìŒ??ì§€ë¬¸ì¸??PBA"),
        create_bulleted_list("FPCç¤?ì§€ë¬¸ì¸??PBA ODM"),
        create_bulleted_list("3D MID ?„ì¥??ëª¨ë“ˆ (?•ë?ê³¼ì œ)"),
    ]
    notion.blocks.children.append(block_id=dream_toggle_id, children=dream_proj)
    print("  ??Dreamtech")
    time.sleep(0.3)
    
    # ===== ?¤ë¥¸ìª?ì»¬ëŸ¼: ê¸°ìˆ  ?¤íƒ =====
    print("\n?“ ?¤ë¥¸ìª?ì»¬ëŸ¼: ê¸°ìˆ  ?¤íƒ ì¶”ê? ì¤?..")
    right_content = [
        create_heading("?› ï¸?ê¸°ìˆ  ?¤íƒ", 2),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=right_col_id, children=right_content)
    time.sleep(0.3)
    
    # Circuit Design ? ê?
    circuit_toggle = notion.blocks.children.append(
        block_id=right_col_id,
        children=[create_toggle("Circuit Design", "??)]
    )
    circuit_id = circuit_toggle["results"][0]["id"]
    
    circuit_skills = [
        create_bulleted_list("STM32F4/F7 Main Board ?¤ê³„"),
        create_bulleted_list("Switching/Linear Regulator ?„ì›ë¶€"),
        create_bulleted_list("16ë¹„íŠ¸ ADC + OP-Amp/TIA ?¼ì„œ"),
        create_bulleted_list("16ë¹„íŠ¸ DAC ?¸ë? ?¸í„°?˜ì´??),
        create_bulleted_list("Photocoupler ?ˆì—° ?Œë¡œ"),
        create_bulleted_list("PCB Artwork (PADS, Altium)"),
    ]
    notion.blocks.children.append(block_id=circuit_id, children=circuit_skills)
    print("  ??Circuit Design")
    time.sleep(0.3)
    
    # Firmware Design ? ê?
    firmware_toggle = notion.blocks.children.append(
        block_id=right_col_id,
        children=[create_toggle("Firmware Design", "?’»")]
    )
    firmware_id = firmware_toggle["results"][0]["id"]
    
    firmware_skills = [
        create_bulleted_list("STM32 ?Œì›¨??ê°œë°œ (C)"),
        create_bulleted_list("FreeRTOS Task ê´€ë¦?),
        create_bulleted_list("PID ?œì–´ ?Œê³ ë¦¬ì¦˜"),
        create_bulleted_list("Y-modem IAP Bootloader"),
        create_bulleted_list("Ethernet TCP/UDP"),
        create_bulleted_list("UART/SPI/I2C ?µì‹ "),
        create_bulleted_list("Touch LCD UI/UX"),
    ]
    notion.blocks.children.append(block_id=firmware_id, children=firmware_skills)
    print("  ??Firmware Design")
    time.sleep(0.3)
    
    # Tools ? ê?
    tools_toggle = notion.blocks.children.append(
        block_id=right_col_id,
        children=[create_toggle("Tools & Software", "?–¥ï¸?)]
    )
    tools_id = tools_toggle["results"][0]["id"]
    
    tools_list = [
        create_bulleted_list("Circuit: OrCAD, PADS, Altium"),
        create_bulleted_list("Firmware: STM32 IDE, Keil, VS Code"),
        create_bulleted_list("Software: C#, Arduino, ESP32"),
        create_bulleted_list("Version: Git, SVN"),
    ]
    notion.blocks.children.append(block_id=tools_id, children=tools_list)
    print("  ??Tools")
    time.sleep(0.3)
    
    # ?™ë ¥ & ?ê²©
    edu_content = [
        create_paragraph(""),
        create_heading("?“ ?™ë ¥ & ?ê²©", 2),
        create_callout("ëª…ì??€ ?„ìê³µí•™ê³?ì¡¸ì—… (2011~2017)", "?“", "gray_background"),
        create_paragraph(""),
        create_callout("?ê²©: ì»´í“¨?°í™œ?©ëŠ¥??3ê¸? ITQ, TOEIC Speaking 110??, "?“œ", "gray_background"),
    ]
    notion.blocks.children.append(block_id=right_col_id, children=edu_content)
    print("  ???™ë ¥ & ?ê²©")
    time.sleep(0.3)
    
    # ===== ?˜ë‹¨ ë©”ì‹œì§€ (?„ì²´ ?ˆë¹„) =====
    footer_blocks = [
        create_paragraph(""),
        create_divider(),
        create_callout("?“ ?ì„¸ ?„ë¡œ?íŠ¸ ?´ìš© ë°?ì½”ë“œ ?˜í”Œ?€ ?¬íŠ¸?´ë¦¬?¤ë? ì°¸ê³ ?´ì£¼?¸ìš”!", "?”—", "green_background"),
        create_paragraph(""),
    ]
    notion.blocks.children.append(block_id=page_id, children=footer_blocks)
    
    print("\n" + "=" * 80)
    print("?‰ ê²½ë ¥ê¸°ìˆ ??v9 ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print("=" * 80)
    
    return page_id

if __name__ == "__main__":
    main()

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import re
from datetime import datetime

base_path = Path('D:/?°ì°¨ê³?)

data = []

for year_folder in sorted(base_path.iterdir()):
    if not year_folder.is_dir():
        continue
    
    year = year_folder.name
    files = sorted(year_folder.glob('*.xlsx'))
    
    for file in files:
        filename = file.name
        
        # ?°ì°¨/ë°˜ì°¨ êµ¬ë¶„
        if 'ë°˜ì°¨ê³? in filename:
            leave_type = 'ë°˜ì°¨'
            days = 0.5
        elif '?°ì°¨ê³? in filename:
            leave_type = '?°ì°¨'
            # D?«ì ?Œì‹±
            if '_D1' in filename or 'ê²°í˜¼' in filename or 'ê²°í˜¼ê¸°ë…?? in filename:
                days = 1
            elif '_D2' in filename:
                days = 2
            elif '_D3' in filename:
                days = 3
            elif '? í˜¼?¬í–‰' in filename:
                days = 5
            else:
                days = 1
        else:
            continue
        
        # ? ì§œ ?Œì‹± (YYMMDD)
        date_match = re.search(r'_(\d{6})', filename)
        if date_match:
            date_str = date_match.group(1)
            try:
                date_obj = datetime.strptime(date_str, '%y%m%d')
                date_formatted = date_obj.strftime('%Y-%m-%d')
            except:
                date_formatted = date_str
        else:
            date_formatted = ''
        
        # ?¬ìœ  ì¶”ì¶œ (?Œì¼ ?´ìš©?ì„œ)
        reason = ''
        try:
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            
            # ?¼ë°˜?ìœ¼ë¡??°ì°¨ê³??‘ì‹?ì„œ ?¬ìœ ???¹ì • ?€???ˆìŒ
            # ?¬ëŸ¬ ê°€?¥í•œ ?„ì¹˜ ?•ì¸
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # "?¬ìœ ", "?´ê??¬ìœ ", "?´ê?ëª©ì " ?±ì˜ ?¤ì›Œ???¤ìŒ ?€ ì°¾ê¸°
                        if '?¬ìœ ' in cell.value or 'ëª©ì ' in cell.value or '?©ë„' in cell.value:
                            # ?¤ìŒ ?€?´ë‚˜ ê°™ì? ?‰ì˜ ?¤ë¥¸ ?€ ?•ì¸
                            next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            if next_cell.value and str(next_cell.value).strip():
                                reason = str(next_cell.value).strip()
                                break
                            # ?„ë˜ ???•ì¸
                            below_cell = ws.cell(row=cell.row + 1, column=cell.column)
                            if below_cell.value and str(below_cell.value).strip():
                                reason = str(below_cell.value).strip()
                                break
                if reason:
                    break
            
            # ?¹ìˆ˜ ?¬ìœ  ?Œì¼ëª…ì—??ì¶”ì¶œ
            if not reason:
                if 'ê²°í˜¼' in filename:
                    reason = 'ê²°í˜¼'
                elif '? í˜¼?¬í–‰' in filename:
                    reason = '? í˜¼?¬í–‰'
                elif 'ê²°í˜¼ê¸°ë…?? in filename:
                    reason = 'ê²°í˜¼ê¸°ë…??
            
            wb.close()
        except Exception as e:
            # ?Œì¼ëª…ì—???¬ìœ  ì¶”ì¶œ ?œë„
            if 'ê²°í˜¼' in filename:
                reason = 'ê²°í˜¼'
            elif '? í˜¼?¬í–‰' in filename:
                reason = '? í˜¼?¬í–‰'
            elif 'ê²°í˜¼ê¸°ë…?? in filename:
                reason = 'ê²°í˜¼ê¸°ë…??
            else:
                reason = f'?Œì¼ ?½ê¸° ?¤ë¥˜: {str(e)[:30]}'
        
        data.append({
            '?°ë„': year,
            '? ì§œ': date_formatted,
            'êµ¬ë¶„': leave_type,
            '?¬ìš©?¼ìˆ˜': days,
            '?¬ìœ ': reason,
            '?Œì¼ëª?: filename
        })

# DataFrame ?ì„±
df = pd.DataFrame(data)

# ? ì§œ???•ë ¬
df = df.sort_values(['?°ë„', '? ì§œ'])

# ?‘ì?ë¡??€??
output_file = 'D:/Portfolio_Professional/?°ì°¨?¬ìš©?´ì—­_?•ë¦¬.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='?„ì²´?´ì—­', index=False)
    
    # ?°ë„ë³??œíŠ¸ ì¶”ê?
    for year in sorted(df['?°ë„'].unique()):
        year_df = df[df['?°ë„'] == year]
        year_df.to_excel(writer, sheet_name=f'{year}??, index=False)
    
    # ?µê³„ ?œíŠ¸
    stats_data = []
    for year in sorted(df['?°ë„'].unique()):
        year_df = df[df['?°ë„'] == year]
        annual_count = len(year_df[year_df['êµ¬ë¶„'] == '?°ì°¨'])
        annual_days = year_df[year_df['êµ¬ë¶„'] == '?°ì°¨']['?¬ìš©?¼ìˆ˜'].sum()
        half_count = len(year_df[year_df['êµ¬ë¶„'] == 'ë°˜ì°¨'])
        half_days = year_df[year_df['êµ¬ë¶„'] == 'ë°˜ì°¨']['?¬ìš©?¼ìˆ˜'].sum()
        total = annual_days + half_days
        
        stats_data.append({
            '?°ë„': year,
            '?°ì°¨ê±´ìˆ˜': annual_count,
            '?°ì°¨?¼ìˆ˜': annual_days,
            'ë°˜ì°¨ê±´ìˆ˜': half_count,
            'ë°˜ì°¨?¼ìˆ˜': half_days,
            'ì´ì‚¬?©ì¼??: total
        })
    
    stats_df = pd.DataFrame(stats_data)
    stats_df.to_excel(writer, sheet_name='?°ë„ë³„í†µê³?, index=False)

print(f'???°ì°¨ ?¬ìš© ?´ì—­???•ë¦¬?˜ì—ˆ?µë‹ˆ?? {output_file}')
print(f'\nì´?{len(df)}ê±´ì˜ ?°ì°¨/ë°˜ì°¨ ê¸°ë¡???•ë¦¬?˜ì—ˆ?µë‹ˆ??')
print(f'\n?œíŠ¸ êµ¬ì„±:')
print(f'  - ?„ì²´?´ì—­: ëª¨ë“  ?°ì°¨/ë°˜ì°¨ ê¸°ë¡')
print(f'  - 2021??2025?? ?°ë„ë³??ì„¸ ?´ì—­')
print(f'  - ?°ë„ë³„í†µê³? ?°ë„ë³??¬ìš© ?µê³„')

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import re
from datetime import datetime

base_path = Path('D:/?°ì°¨ê³?)

data = []

for year_folder in sorted(base_path.iterdir()):
    if not year_folder.is_dir():
        continue
    
    year = year_folder.name
    files = sorted(year_folder.glob('*.xlsx'))
    
    for file in files:
        filename = file.name
        
        # ?°ì°¨/ë°˜ì°¨ êµ¬ë¶„
        if 'ë°˜ì°¨ê³? in filename:
            leave_type = 'ë°˜ì°¨'
            days = 0.5
        elif '?°ì°¨ê³? in filename:
            leave_type = '?°ì°¨'
            days = 1  # ê¸°ë³¸ê°?
        else:
            continue
        
        # ? ì§œ ?Œì‹± (YYMMDD)
        date_match = re.search(r'_(\d{6})', filename)
        if date_match:
            date_str = date_match.group(1)
            try:
                date_obj = datetime.strptime(date_str, '%y%m%d')
                date_formatted = date_obj.strftime('%Y-%m-%d')
            except:
                date_formatted = date_str
        else:
            date_formatted = ''
        
        # ?¬ìœ ?€ ?¼ìˆ˜ë¥??Œì¼ ?´ìš©?ì„œ ì¶”ì¶œ
        reason = ''
        try:
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            
            days_found = False
            reason_found = False
            
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                for i, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        # ?¼ìˆ˜ ì¶”ì¶œ (?°ì°¨ë§??´ë‹¹)
                        if leave_type == '?°ì°¨' and not days_found:
                            match = re.search(r'(\d+\.?\d*)\s*??s*ê°?, cell)
                            if match:
                                days = float(match.group(1))
                                days_found = True
                        
                        # ?¬ìœ  ì¶”ì¶œ
                        if not reason_found:
                            if '?¬ìœ ' in cell:
                                # ?¤ìŒ ?€ ?•ì¸
                                if i + 1 < len(row) and row[i + 1]:
                                    reason_text = str(row[i + 1]).strip()
                                    if reason_text and len(reason_text) < 100:
                                        reason = reason_text
                                        reason_found = True
            
            wb.close()
        except Exception as e:
            pass
        
        # ?¬ìœ ê°€ ?†ìœ¼ë©??Œì¼ëª…ì—??ì¶”ì¶œ
        if not reason or reason == 'None':
            if 'ê²°í˜¼ê¸°ë…?? in filename:
                reason = 'ê²°í˜¼ê¸°ë…??
            elif '? í˜¼?¬í–‰' in filename:
                reason = '? í˜¼?¬í–‰'
            elif 'ê²°í˜¼' in filename:
                reason = 'ê²°í˜¼'
            else:
                reason = 'ê°œì¸ ?¬ìœ '
        
        data.append({
            '?°ë„': year,
            '? ì§œ': date_formatted,
            'êµ¬ë¶„': leave_type,
            '?¬ìš©?¼ìˆ˜': days,
            '?¬ìœ ': reason,
            '?Œì¼ëª?: filename
        })

# DataFrame ?ì„±
df = pd.DataFrame(data)

# ? ì§œ???•ë ¬
df = df.sort_values(['?°ë„', '? ì§œ'])

# ?‘ì?ë¡??€??
output_file = 'D:/Portfolio_Professional/?°ì°¨?¬ìš©?´ì—­_?•ë¦¬.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='?„ì²´?´ì—­', index=False)
    
    # ?°ë„ë³??œíŠ¸ ì¶”ê?
    for year in sorted(df['?°ë„'].unique()):
        year_df = df[df['?°ë„'] == year].reset_index(drop=True)
        year_total = year_df['?¬ìš©?¼ìˆ˜'].sum()
        
        # ?°ë„ë³??°ì´??
        year_df.to_excel(writer, sheet_name=f'{year}??, index=False)
        
        # ê°??œíŠ¸???©ê³„ ì¶”ê?
        workbook = writer.book
        worksheet = writer.sheets[f'{year}??]
        last_row = len(year_df) + 2
        worksheet.cell(row=last_row, column=1, value='?©ê³„')
        worksheet.cell(row=last_row, column=4, value=year_total)
    
    # ?µê³„ ?œíŠ¸
    stats_data = []
    for year in sorted(df['?°ë„'].unique()):
        year_df = df[df['?°ë„'] == year]
        annual_count = len(year_df[year_df['êµ¬ë¶„'] == '?°ì°¨'])
        annual_days = year_df[year_df['êµ¬ë¶„'] == '?°ì°¨']['?¬ìš©?¼ìˆ˜'].sum()
        half_count = len(year_df[year_df['êµ¬ë¶„'] == 'ë°˜ì°¨'])
        half_days = year_df[year_df['êµ¬ë¶„'] == 'ë°˜ì°¨']['?¬ìš©?¼ìˆ˜'].sum()
        total = annual_days + half_days
        
        stats_data.append({
            '?°ë„': year,
            '?°ì°¨ê±´ìˆ˜': annual_count,
            '?°ì°¨?¼ìˆ˜': annual_days,
            'ë°˜ì°¨ê±´ìˆ˜': half_count,
            'ë°˜ì°¨?¼ìˆ˜': half_days,
            'ì´ì‚¬?©ì¼??: total
        })
    
    # ì´í•©ê³?ì¶”ê?
    total_stats = {
        '?°ë„': 'ì´í•©ê³?,
        '?°ì°¨ê±´ìˆ˜': sum(s['?°ì°¨ê±´ìˆ˜'] for s in stats_data),
        '?°ì°¨?¼ìˆ˜': sum(s['?°ì°¨?¼ìˆ˜'] for s in stats_data),
        'ë°˜ì°¨ê±´ìˆ˜': sum(s['ë°˜ì°¨ê±´ìˆ˜'] for s in stats_data),
        'ë°˜ì°¨?¼ìˆ˜': sum(s['ë°˜ì°¨?¼ìˆ˜'] for s in stats_data),
        'ì´ì‚¬?©ì¼??: sum(s['ì´ì‚¬?©ì¼??] for s in stats_data)
    }
    stats_data.append(total_stats)
    
    stats_df = pd.DataFrame(stats_data)
    stats_df.to_excel(writer, sheet_name='?°ë„ë³„í†µê³?, index=False)

print(f'???°ì°¨ ?¬ìš© ?´ì—­???•ë¦¬?˜ì—ˆ?µë‹ˆ?? {output_file}')
print(f'\nì´?{len(df)}ê±´ì˜ ?°ì°¨/ë°˜ì°¨ ê¸°ë¡???•ë¦¬?˜ì—ˆ?µë‹ˆ??')
print(f'\n?°ë„ë³??¬ìš© ?¼ìˆ˜:')
for year in sorted(df['?°ë„'].unique()):
    year_total = df[df['?°ë„'] == year]['?¬ìš©?¼ìˆ˜'].sum()
    print(f'  {year}?? {year_total}??)
print(f'\nì´??¬ìš© ?¼ìˆ˜: {df["?¬ìš©?¼ìˆ˜"].sum()}??)

# -*- coding: utf-8 -*-
from notion_client import Client
import json
import sys
import io

# UTF-8 ì¶œë ¥ ?¤ì •
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# TODO: ?˜ê²½ ë³€?˜ì—??? í° ?½ê¸°
notion = Client(auth='YOUR_NOTION_TOKEN_HERE')
v5_page_id = '2e8f0746-9821-818f-82a0-d0421f8d53ea'

def extract_text(rich_text_array):
    return ''.join([text['plain_text'] for text in rich_text_array])

def print_block(block, indent=0):
    prefix = "  " * indent
    block_type = block['type']
    
    if block_type == 'heading_1':
        print(f"{prefix}# {extract_text(block['heading_1']['rich_text'])}")
    elif block_type == 'heading_2':
        print(f"{prefix}## {extract_text(block['heading_2']['rich_text'])}")
    elif block_type == 'heading_3':
        print(f"{prefix}### {extract_text(block['heading_3']['rich_text'])}")
    elif block_type == 'paragraph':
        text = extract_text(block['paragraph']['rich_text'])
        if text:
            print(f"{prefix}{text}")
    elif block_type == 'bulleted_list_item':
        print(f"{prefix}- {extract_text(block['bulleted_list_item']['rich_text'])}")
    elif block_type == 'toggle':
        print(f"{prefix}[? ê?] {extract_text(block['toggle']['rich_text'])}")
        if block['toggle'].get('children'):
            for child in block['toggle']['children']:
                print_block(child, indent + 1)
    elif block_type == 'callout':
        icon = block['callout'].get('icon')
        emoji = icon.get('emoji', '?’¡') if icon else '?’¡'
        print(f"{prefix}{emoji} {extract_text(block['callout']['rich_text'])}")
    elif block_type == 'code':
        lang = block['code']['language']
        code = extract_text(block['code']['rich_text'])
        print(f"{prefix}```{lang}")
        print(f"{code}")
        print(f"{prefix}```")
    elif block_type == 'divider':
        print(f"{prefix}---")

print("=== ?¸ì…˜ ?¬íŠ¸?´ë¦¬??v5 - Final ?´ìš© ===\n")

# ?˜ì´ì§€??ëª¨ë“  ë¸”ë¡ ê°€?¸ì˜¤ê¸?
blocks = notion.blocks.children.list(block_id=v5_page_id, page_size=100)

for block in blocks['results']:
    # ? ê? ë¸”ë¡?€ ?ì‹??ê°€?¸ì?????
    if block['type'] == 'toggle' and block['has_children']:
        children = notion.blocks.children.list(block_id=block['id'], page_size=100)
        block['toggle']['children'] = children['results']
    
    print_block(block)
    print()

# -*- coding: utf-8 -*-
from notion_client import Client
import json

notion = Client(auth='YOUR_NOTION_TOKEN_HERE')
v5_page_id = '2e8f0746-9821-818f-82a0-d0421f8d53ea'

def extract_text(rich_text_array):
    if not rich_text_array:
        return ""
    return ''.join([text['plain_text'] for text in rich_text_array])

def process_block(block, indent=0):
    prefix = "  " * indent
    block_type = block['type']
    result = []
    
    if block_type == 'heading_1':
        result.append(f"{prefix}# {extract_text(block['heading_1']['rich_text'])}")
    elif block_type == 'heading_2':
        result.append(f"{prefix}## {extract_text(block['heading_2']['rich_text'])}")
    elif block_type == 'heading_3':
        result.append(f"{prefix}### {extract_text(block['heading_3']['rich_text'])}")
    elif block_type == 'paragraph':
        text = extract_text(block['paragraph']['rich_text'])
        if text:
            result.append(f"{prefix}{text}")
    elif block_type == 'bulleted_list_item':
        text = extract_text(block['bulleted_list_item']['rich_text'])
        result.append(f"{prefix}- {text}")
    elif block_type == 'callout':
        icon = block['callout'].get('icon')
        emoji = icon.get('emoji', '?’¡') if icon else '?’¡'
        text = extract_text(block['callout']['rich_text'])
        result.append(f"{prefix}{emoji} {text}")
    elif block_type == 'code':
        lang = block['code']['language']
        code = extract_text(block['code']['rich_text'])
        result.append(f"{prefix}```{lang}")
        result.append(code)
        result.append(f"{prefix}```")
    elif block_type == 'divider':
        result.append(f"{prefix}---")
    elif block_type == 'toggle':
        text = extract_text(block['toggle']['rich_text'])
        result.append(f"\n{'='*80}")
        result.append(f"{prefix}[? ê?] {text}")
        result.append(f"{'='*80}")
    
    return result

# ?˜ì´ì§€??ëª¨ë“  ë¸”ë¡ ê°€?¸ì˜¤ê¸?
print("?¸ì…˜ v5 ?„ì²´ ?´ìš©??ê°€?¸ì˜¤??ì¤?..")
blocks = notion.blocks.children.list(block_id=v5_page_id, page_size=100)

all_content = []
all_content.append("=== ?¸ì…˜ ?¬íŠ¸?´ë¦¬??v5 - Final ?„ì²´ ?´ìš© ===\n")

for block in blocks['results']:
    lines = process_block(block)
    all_content.extend(lines)
    
    # ? ê? ë¸”ë¡?´ë©´ ?ì‹??ê°€?¸ì˜¤ê¸?
    if block['type'] == 'toggle' and block['has_children']:
        children = notion.blocks.children.list(block_id=block['id'], page_size=100)
        for child in children['results']:
            child_lines = process_block(child, indent=1)
            all_content.extend(child_lines)
        all_content.append("")

# ?Œì¼ë¡??€??
output_file = 'notion_v5_full_content.txt'
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('\n'.join(all_content))

print(f"???€???„ë£Œ: {output_file}")
print(f"??ì´?{len(all_content)}ì¤?)

# Psi-1000 ë¶€ë¶„ë§Œ ì¶”ì¶œ?˜ì—¬ ì¶œë ¥
psi_start = False
psi_content = []
for line in all_content:
    if 'Psi-1000' in line and '? ê?' in line:
        psi_start = True
    if psi_start:
        psi_content.append(line)
        if line.startswith('===') and len(psi_content) > 5:
            break

print("\n" + "="*80)
print("?“Š Psi-1000 ?„ë¡œ?íŠ¸ ?´ìš©:")
print("="*80)
for line in psi_content[:50]:  # ì²˜ìŒ 50ì¤„ë§Œ ì¶œë ¥
    print(line)

# -*- coding: utf-8 -*-
"""
?¸ì…˜?ì„œ "ì¡°ë½??Rakeen Jo" ?˜ì´ì§€ ì°¾ê¸° ë°??´ìš© ê°€?¸ì˜¤ê¸?
"""

from notion_client import Client

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

print("=" * 60)
print("ë¶€ëª??˜ì´ì§€ ??ëª¨ë“  ?ì‹ ?˜ì´ì§€ ê²€??ì¤?..")
print("=" * 60)

# ë¶€ëª??˜ì´ì§€???ì‹ ë¸”ë¡ ê°€?¸ì˜¤ê¸?
children = notion.blocks.children.list(block_id=PARENT_PAGE_ID)

print(f"\nì´?{len(children['results'])}ê°œì˜ ?ì‹ ë¸”ë¡ ë°œê²¬\n")

target_page_id = None

for block in children['results']:
    if block['type'] == 'child_page':
        title = block['child_page']['title']
        page_id = block['id']
        print(f"?“„ ?˜ì´ì§€: {title}")
        print(f"   ID: {page_id}\n")
        
        if 'ì¡°ë½?? in title or 'Rakeen' in title:
            target_page_id = page_id
            print(f"???€ê²??˜ì´ì§€ ë°œê²¬: {title}\n")

if target_page_id:
    print("=" * 60)
    print("?˜ì´ì§€ ?´ìš© ê°€?¸ì˜¤??ì¤?..")
    print("=" * 60)
    
    # ?˜ì´ì§€ ?´ìš© ê°€?¸ì˜¤ê¸?
    page_blocks = notion.blocks.children.list(block_id=target_page_id, page_size=100)
    
    print(f"\nì´?{len(page_blocks['results'])}ê°œì˜ ë¸”ë¡\n")
    
    for idx, block in enumerate(page_blocks['results'], 1):
        block_type = block['type']
        block_id = block['id']
        
        print(f"\n[{idx}] ë¸”ë¡ ?€?? {block_type} (ID: {block_id})")
        
        # ?ìŠ¤??ì¶”ì¶œ
        if block_type == 'paragraph':
            texts = block['paragraph']['rich_text']
            if texts:
                content = ''.join([t['plain_text'] for t in texts])
                print(f"    ?´ìš©: {content}")
        
        elif block_type.startswith('heading_'):
            level = block_type.split('_')[1]
            texts = block[block_type]['rich_text']
            if texts:
                content = ''.join([t['plain_text'] for t in texts])
                print(f"    ?œëª©(H{level}): {content}")
        
        elif block_type == 'bulleted_list_item':
            texts = block['bulleted_list_item']['rich_text']
            if texts:
                content = ''.join([t['plain_text'] for t in texts])
                print(f"    ??{content}")
        
        elif block_type == 'numbered_list_item':
            texts = block['numbered_list_item']['rich_text']
            if texts:
                content = ''.join([t['plain_text'] for t in texts])
                print(f"    {idx}. {content}")
        
        elif block_type == 'code':
            texts = block['code']['rich_text']
            language = block['code']['language']
            if texts:
                content = ''.join([t['plain_text'] for t in texts])
                print(f"    ì½”ë“œ({language}):")
                print(f"    {content[:200]}...")  # ì²˜ìŒ 200?ë§Œ
        
        elif block_type == 'callout':
            texts = block['callout']['rich_text']
            icon = block['callout'].get('icon', {})
            emoji = icon.get('emoji', '?’¡') if icon.get('type') == 'emoji' else '?’¡'
            if texts:
                content = ''.join([t['plain_text'] for t in texts])
                print(f"    {emoji} {content}")
        
        elif block_type == 'toggle':
            texts = block['toggle']['rich_text']
            if texts:
                content = ''.join([t['plain_text'] for t in texts])
                print(f"    ?”½ ? ê?: {content}")
        
        elif block_type == 'divider':
            print(f"    ?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€?€")
        
        elif block_type == 'table':
            print(f"    ?“Š ?Œì´ë¸?(?ˆë¹„: {block['table']['table_width']})")
        
        else:
            print(f"    (ê¸°í? ë¸”ë¡: {block_type})")
    
    # has_moreê°€ ?ˆìœ¼ë©?ì¶”ê? ë¸”ë¡ ê°€?¸ì˜¤ê¸?
    if page_blocks.get('has_more'):
        print("\n? ï¸ ??ë§ì? ë¸”ë¡???ˆìŠµ?ˆë‹¤. ?„ì²´ ?´ìš©??ë³´ë ¤ë©??˜ì´ì§€?¤ì´???„ìš”")
    
else:
    print("\n??'ì¡°ë½??Rakeen Jo' ?˜ì´ì§€ë¥?ì°¾ì„ ???†ìŠµ?ˆë‹¤.")

# -*- coding: utf-8 -*-
"""
"ì¡°ë½?? Rakeen Jo" ?˜ì´ì§€ ?„ì „ ë¶„ì„ (ëª¨ë“  ?ì‹ ë¸”ë¡ ?¬í•¨)
"""

from notion_client import Client
import json

# TODO: ?˜ê²½ ë³€?˜ì—??? í° ?½ê¸°
NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
TARGET_PAGE_ID = "b920a0b4-9159-47ab-8efc-71b7b5a155bf"
notion = Client(auth=NOTION_TOKEN)

print("=" * 100)
print("?“„ 'ì¡°ë½?? Rakeen Jo' ?˜ì´ì§€ ?„ì „ ë¶„ì„ (ëª¨ë“  ?ì‹ ë¸”ë¡ ?¬ê? ?ìƒ‰)")
print("=" * 100)

def extract_text(rich_text_array):
    if not rich_text_array:
        return ""
    return ''.join([t['plain_text'] for t in rich_text_array])

def get_all_children(block_id, depth=0):
    """?¬ê??ìœ¼ë¡?ëª¨ë“  ?ì‹ ë¸”ë¡ ê°€?¸ì˜¤ê¸?""
    try:
        children = notion.blocks.children.list(block_id=block_id, page_size=100)
        return children['results']
    except:
        return []

def print_block(block, depth=0):
    """ë¸”ë¡???¤ì—¬?°ê¸°?€ ?¨ê»˜ ì¶œë ¥"""
    indent = "  " * depth
    block_type = block['type']
    block_id = block['id']
    
    # ë¸”ë¡ ?€??ì¶œë ¥
    print(f"\n{indent}[{block_type.upper()}] (depth={depth})")
    
    # ?´ìš© ì¶”ì¶œ
    content = ""
    
    if block_type == 'paragraph':
        content = extract_text(block['paragraph']['rich_text'])
    
    elif block_type.startswith('heading_'):
        content = extract_text(block[block_type]['rich_text'])
    
    elif block_type == 'bulleted_list_item':
        content = "??" + extract_text(block['bulleted_list_item']['rich_text'])
    
    elif block_type == 'numbered_list_item':
        content = extract_text(block['numbered_list_item']['rich_text'])
    
    elif block_type == 'callout':
        icon = block['callout'].get('icon', {})
        emoji = icon.get('emoji', '?’¡') if icon.get('type') == 'emoji' else '?’¡'
        content = f"{emoji} " + extract_text(block['callout']['rich_text'])
    
    elif block_type == 'toggle':
        content = "?”½ " + extract_text(block['toggle']['rich_text'])
    
    elif block_type == 'quote':
        content = "> " + extract_text(block['quote']['rich_text'])
    
    elif block_type == 'code':
        language = block['code']['language']
        code_text = extract_text(block['code']['rich_text'])
        content = f"[{language}]\n{indent}{code_text[:200]}"
    
    elif block_type == 'divider':
        content = "?€" * 60
    
    elif block_type == 'table':
        width = block['table']['table_width']
        content = f"?“Š ?Œì´ë¸?({width}??"
    
    elif block_type == 'table_row':
        cells = block['table_row']['cells']
        content = " | ".join([extract_text(cell) for cell in cells])
    
    elif block_type == 'to_do':
        checked = block['to_do']['checked']
        text = extract_text(block['to_do']['rich_text'])
        content = f"{'?? if checked else '??} {text}"
    
    elif block_type == 'image':
        img_type = block['image']['type']
        if img_type == 'external':
            content = f"?–¼ï¸??¸ë? ?´ë?ì§€: {block['image']['external']['url']}"
        elif img_type == 'file':
            content = f"?–¼ï¸??Œì¼ ?´ë?ì§€: {block['image']['file']['url']}"
    
    elif block_type == 'video':
        vid_type = block['video']['type']
        if vid_type == 'external':
            content = f"?¥ ?¸ë? ë¹„ë””?? {block['video']['external']['url']}"
    
    elif block_type == 'embed':
        content = f"?”— ?„ë² ?? {block['embed']['url']}"
    
    # ?´ìš© ì¶œë ¥
    if content:
        for line in content.split('\n'):
            print(f"{indent}  {line}")
    
    # ?ì‹ ë¸”ë¡ ?¬ê? ?ìƒ‰
    if block.get('has_children'):
        children = get_all_children(block_id)
        for child in children:
            print_block(child, depth + 1)

# ë©”ì¸ ?˜ì´ì§€ ë¸”ë¡??ê°€?¸ì˜¤ê¸?
main_blocks = get_all_children(TARGET_PAGE_ID)

print(f"\nì´?{len(main_blocks)}ê°œì˜ ìµœìƒ??ë¸”ë¡\n")
print("=" * 100)

# ëª¨ë“  ë¸”ë¡ ?¬ê? ì¶œë ¥
for idx, block in enumerate(main_blocks, 1):
    print(f"\n{'='*100}")
    print(f"ë¸”ë¡ #{idx}")
    print_block(block, depth=0)

print(f"\n{'='*100}")
print("ë¶„ì„ ?„ë£Œ!")
print("=" * 100)

"""
Á¶¶ôÇö Æ÷Æ®Æú¸®¿À ³ë¼Ç ÀÚµ¿ »ı¼º ½ºÅ©¸³Æ®
"""

from notion_client import Client
import json

# Notion API ¼³Á¤
NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"

# Å¬¶óÀÌ¾ğÆ® ÃÊ±âÈ­
notion = Client(auth=NOTION_TOKEN)

def search_page(page_name):
    """ÆäÀÌÁö °Ë»ö"""
    try:
        response = notion.search(
            query=page_name,
            filter={"property": "object", "value": "page"}
        )
        if response["results"]:
            for result in response["results"]:
                title = ""
                if "title" in result.get("properties", {}):
                    title_prop = result["properties"]["title"]
                    if title_prop.get("title"):
                        title = title_prop["title"][0]["plain_text"]
                elif "properties" in result:
                    for prop in result["properties"].values():
                        if prop.get("type") == "title" and prop.get("title"):
                            title = prop["title"][0]["plain_text"]
                            break
                
                # child_pageÀÎ °æ¿ì
                if result.get("type") == "child_page":
                    title = result.get("child_page", {}).get("title", "")
                
                print(f"  ¹ß°ß: {title} (ID: {result['id']})")
                if page_name in title or title in page_name:
                    return result["id"]
            
            # Ã¹ ¹øÂ° °á°ú ¹İÈ¯
            return response["results"][0]["id"]
        return None
    except Exception as e:
        print(f"°Ë»ö ¿À·ù: {e}")
        return None

def create_portfolio_content(parent_id):
    """Æ÷Æ®Æú¸®¿À ÄÜÅÙÃ÷ »ı¼º"""
    
    # ¸ŞÀÎ ÆäÀÌÁö »ı¼º
    print("\n?? Æ÷Æ®Æú¸®¿À ÆäÀÌÁö »ı¼º Áß...")
    
    portfolio_page = notion.pages.create(
        parent={"page_id": parent_id},
        icon={"type": "emoji", "emoji": "?????"},
        cover={
            "type": "external",
            "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1200"}
        },
        properties={
            "title": {
                "title": [{"text": {"content": "Á¶¶ôÇö | Senior Embedded Systems Engineer"}}]
            }
        },
        children=[
            # Çì´õ ¼½¼Ç
            {
                "object": "block",
                "type": "callout",
                "callout": {
                    "icon": {"type": "emoji", "emoji": "??"},
                    "color": "blue_background",
                    "rich_text": [{"type": "text", "text": {"content": "9³â °æ·Â | Hardware & Firmware Full-Stack °³¹ß | STM32 Expert | FreeRTOS | PID Control"}}]
                }
            },
            # ¿¬¶ôÃ³
            {
                "object": "block",
                "type": "paragraph",
                "paragraph": {
                    "rich_text": [
                        {"type": "text", "text": {"content": "?? 92lock@kakao.com | ?? 010-7311-0402 | "}, "annotations": {"color": "gray"}},
                        {"type": "text", "text": {"content": "GitHub", "link": {"url": "https://github.com/gari210404"}}, "annotations": {"color": "blue"}}
                    ]
                }
            },
            {"object": "block", "type": "divider", "divider": {}},
            
            # Executive Summary
            {
                "object": "block",
                "type": "heading_1",
                "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Executive Summary"}}]}
            },
            {
                "object": "block",
                "type": "paragraph",
                "paragraph": {
                    "rich_text": [{"type": "text", "text": {"content": "9³â °æ·ÂÀÇ ½Ã´Ï¾î ÀÓº£µğµå ½Ã½ºÅÛ ¿£Áö´Ï¾î·Î, ÇÏµå¿ş¾î ¼³°èºÎÅÍ Æß¿ş¾î °³¹ß±îÁö Full-Stack Embedded °³¹ß Àü¹®°¡ÀÔ´Ï´Ù. ¹İµµÃ¼ °øÁ¤ Àåºñ, ±¤ÇĞ ºĞ¼® ±â±â, »ê¾÷¿ë Á¦¾î ½Ã½ºÅÛÀÇ ¼³°è ¹× °³¹ß¿¡¼­ °ËÁõµÈ ½ÇÀûÀ» º¸À¯ÇÏ°í ÀÖ½À´Ï´Ù."}}]
                }
            },
            {"object": "block", "type": "divider", "divider": {}},
            
            # Core Competencies
            {
                "object": "block",
                "type": "heading_1",
                "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Core Competencies"}}]}
            },
            {
                "object": "block",
                "type": "column_list",
                "column_list": {"children": []}
            },
        ]
    )
    
    page_id = portfolio_page["id"]
    print(f"  ? ¸ŞÀÎ ÆäÀÌÁö »ı¼º ¿Ï·á: {page_id}")
    
    # ÇÙ½É ¿ª·® Å×ÀÌºí Ãß°¡
    notion.blocks.children.append(
        block_id=page_id,
        children=[
            {
                "object": "block",
                "type": "table",
                "table": {
                    "table_width": 3,
                    "has_column_header": True,
                    "has_row_header": False,
                    "children": [
                        {
                            "type": "table_row",
                            "table_row": {
                                "cells": [
                                    [{"type": "text", "text": {"content": "ºĞ¾ß"}}],
                                    [{"type": "text", "text": {"content": "±â¼ú"}}],
                                    [{"type": "text", "text": {"content": "°æ·Â"}}]
                                ]
                            }
                        },
                        {
                            "type": "table_row",
                            "table_row": {
                                "cells": [
                                    [{"type": "text", "text": {"content": "Firmware"}}],
                                    [{"type": "text", "text": {"content": "STM32, FreeRTOS, HAL/LL, Bootloader"}}],
                                    [{"type": "text", "text": {"content": "9³â"}}]
                                ]
                            }
                        },
                        {
                            "type": "table_row",
                            "table_row": {
                                "cells": [
                                    [{"type": "text", "text": {"content": "Hardware"}}],
                                    [{"type": "text", "text": {"content": "Analog/Digital Circuit, PCB Layout"}}],
                                    [{"type": "text", "text": {"content": "9³â"}}]
                                ]
                            }
                        },
                        {
                            "type": "table_row",
                            "table_row": {
                                "cells": [
                                    [{"type": "text", "text": {"content": "Communication"}}],
                                    [{"type": "text", "text": {"content": "Modbus RTU/ASCII, TCP/IP, CAN"}}],
                                    [{"type": "text", "text": {"content": "8³â"}}]
                                ]
                            }
                        },
                        {
                            "type": "table_row",
                            "table_row": {
                                "cells": [
                                    [{"type": "text", "text": {"content": "Control"}}],
                                    [{"type": "text", "text": {"content": "PID, Auto-tuning, State Machine"}}],
                                    [{"type": "text", "text": {"content": "6³â"}}]
                                ]
                            }
                        }
                    ]
                }
            }
        ]
    )
    
    # ÇÁ·ÎÁ§Æ® ¼½¼Ç
    notion.blocks.children.append(
        block_id=page_id,
        children=[
            {"object": "block", "type": "divider", "divider": {}},
            {
                "object": "block",
                "type": "heading_1",
                "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Major Projects"}}]}
            }
        ]
    )
    
    # ÇÁ·ÎÁ§Æ® 1: L-LPC
    create_project_toggle(page_id, 
        "1?? L-LPC (H-Sensor) System | 2022-2024 | ? ´ëÇ¥ ÇÁ·ÎÁ§Æ®",
        [
            "¹İµµÃ¼ Àú¾Ğ Ã¨¹ö Á¦¾î ½Ã½ºÅÛ - Hardware & Firmware Full-Stack °³¹ß",
            "",
            "?? ÇÏµå¿ş¾î ¼³°è:",
            "? Main Board: STM32H743 (480MHz, 2MB Flash)",
            "? Analog Frontend: °íÁ¤¹Ğ TIA (S/N > 60dB)",
            "? FPGA Interface: Xilinx Artix-7 ¿¬µ¿",
            "? Power: Multi-rail SMPS (È¿À² > 90%)",
            "",
            "?? Æß¿ş¾î °³¹ß:",
            "? FreeRTOS ±â¹İ ¸ÖÆ¼ÅÂ½ºÅ© ±¸Á¶ (7°³ Task)",
            "? DMA ±â¹İ °í¼Ó ADC µ¥ÀÌÅÍ ¼öÁı",
            "? TCP/IP, Modbus RTU Åë½Å ±¸Çö",
            "",
            "?? µğ¹ö±ë °æÇè:",
            "? ADC ³ëÀÌÁî ÀÌ½´: ¡¾500mV ¡æ ¡¾5mV (100¹è °³¼±)",
            "? FreeRTOS µ¥µå¶ô ÇØ°á: Mutex ¼ø¼­ ÅëÀÏ",
            "",
            "?? ¼º°ú: ¾ç»ê 100+ units, ÇÊµå ºÒ·®·ü < 0.2%"
        ]
    )
    
    # ÇÁ·ÎÁ§Æ® 2: Psi-1000
    create_project_toggle(page_id,
        "2?? Psi-1000/3000 Pressure Controller | 2019-2023 | ? PID Àü¹®",
        [
            "Á¤¹Ğ Áø°ø ¾Ğ·Â Á¦¾î ½Ã½ºÅÛ - PID ¾Ë°í¸®Áò °³¹ß ¹× ÃÖÀûÈ­",
            "",
            "?? PID Á¦¾î ¾Ë°í¸®Áò:",
            "? Anti-windup, Derivative filtering ±¸Çö",
            "? Auto-tuning (Relay Feedback, Z-N Method)",
            "? Cascade control (Pressure + Temperature)",
            "",
            "?? Åë½Å ÇÁ·ÎÅäÄİ:",
            "? Modbus RTU/ASCII Master/Slave",
            "? Ethernet (Modbus TCP)",
            "? Beckhoff PLC ¿¬µ¿",
            "",
            "?? ¼º°ú:",
            "? ¡¾0.1% ¾Ğ·Â Á¦¾î Á¤È®µµ ´Ş¼º",
            "? < 2ÃÊ Á¤Âø ½Ã°£ (Step Response)",
            "? µ¿¾Æ´ëÇĞ±³ ¿¬±¸ Çù·Â"
        ]
    )
    
    # ÇÁ·ÎÁ§Æ® 3: Nu-2000
    create_project_toggle(page_id,
        "3?? Nu-2000 Optical Analysis System | 2021-2022",
        [
            "±¤ÇĞ Èí¼ö ºĞ±¤ ½Ã½ºÅÛ - Multi-channel Æß¿ş¾î °³¹ß",
            "",
            "?? Æß¿ş¾î Æ¯Â¡:",
            "? FreeRTOS 5°³ Task ±¸Á¶",
            "? 4Ã¤³Î LED Driver PWM Á¦¾î",
            "? 4Ã¤³Î Photodiode °í¼Ó ADC (250kSPS)",
            "? IAP Bootloader °³¹ß (UART/USB)",
            "",
            "?? ÇÏµå¿ş¾î Áö¿ø:",
            "? ±¤ÇĞ ¸ğµâ Wiki Optics ¿ÜÁÖ °ü¸®",
            "? Alpha/Beta ¹öÀü ¼ö¶ô ½ÃÇè"
        ]
    )
    
    # ÇÁ·ÎÁ§Æ® 4: MS
    create_project_toggle(page_id,
        "4?? MS (Mass Spectrometer) System | 2024-2025 | ?? ÃÖ½Å",
        [
            "¹İµµÃ¼ °øÁ¤ °¡½º ºĞ¼® Mass Spectrometer - ½Ã½ºÅÛ Æß¿ş¾î °³¹ß",
            "",
            "? µ¥ÀÌÅÍ ¼öÁı Æß¿ş¾î °³¹ß",
            "? RF Á¦¾î ÀÎÅÍÆäÀÌ½º",
            "? Åë½Å ÇÁ·ÎÅäÄİ ±¸Çö",
            "? ÇöÀç °³¹ß ÁøÇà Áß"
        ]
    )
    
    # ÇÁ·ÎÁ§Æ® 5: LE_Laser
    create_project_toggle(page_id,
        "5?? LE_Laser (Mantis SSC) | 2025 | ?? ÃÖ½Å",
        [
            "·¹ÀÌÀú ±â¹İ Á¤¹Ğ ¼¾¼­ ½Ã½ºÅÛ",
            "",
            "? STM32G071RB (ÃÖ½Å Cortex-M0+)",
            "? 22-bit ADC °íºĞÇØ´É ÃøÁ¤",
            "? Compact 2-board design",
            "? Low-power operation"
        ]
    )
    
    # ÄÚµå »ùÇÃ ¼½¼Ç
    notion.blocks.children.append(
        block_id=page_id,
        children=[
            {"object": "block", "type": "divider", "divider": {}},
            {
                "object": "block",
                "type": "heading_1",
                "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Code Samples"}}]}
            },
            {
                "object": "block",
                "type": "heading_3",
                "heading_3": {"rich_text": [{"type": "text", "text": {"content": "PID Controller Implementation"}}]}
            },
            {
                "object": "block",
                "type": "code",
                "code": {
                    "language": "c",
                    "rich_text": [{"type": "text", "text": {"content": """// PID Update Function
float PID_Update(PID_Controller_t* pid, float setpoint, float measurement) {
    float error = setpoint - measurement;
    float dt = 0.001f;  // 1ms sampling
    
    // Proportional
    float P = pid->Kp * error;
    
    // Integral with Anti-windup
    pid->integral += (pid->Ki * dt) * error;
    pid->integral = CLAMP(pid->integral, -pid->integral_max, pid->integral_max);
    
    // Derivative with Low-pass Filter
    float derivative = (error - pid->prev_error) / dt;
    pid->prev_derivative = 0.1f * derivative + 0.9f * pid->prev_derivative;
    float D = pid->Kd * pid->prev_derivative;
    
    pid->prev_error = error;
    return CLAMP(P + pid->integral + D, pid->output_min, pid->output_max);
}"""}}]
                }
            },
            {
                "object": "block",
                "type": "heading_3",
                "heading_3": {"rich_text": [{"type": "text", "text": {"content": "FreeRTOS Task with Queue"}}]}
            },
            {
                "object": "block",
                "type": "code",
                "code": {
                    "language": "c",
                    "rich_text": [{"type": "text", "text": {"content": """// ADC Data Acquisition Task
void ADC_Task(void *argument) {
    uint16_t adc_data[4];
    
    while(1) {
        // DMA ±â¹İ ADC º¯È¯
        HAL_ADC_Start_DMA(&hadc1, (uint32_t*)adc_data, 4);
        
        // ¿Ï·á ´ë±â
        if (xSemaphoreTake(ADC_Sem, pdMS_TO_TICKS(100)) == pdTRUE) {
            // Queue·Î µ¥ÀÌÅÍ Àü¼Û
            xQueueSend(ADC_DataQueue, &adc_data, 0);
        }
        
        osDelay(10);  // 100Hz
    }
}"""}}]
                }
            }
        ]
    )
    
    # ±â¼ú ½ºÅ³ ¼½¼Ç
    notion.blocks.children.append(
        block_id=page_id,
        children=[
            {"object": "block", "type": "divider", "divider": {}},
            {
                "object": "block",
                "type": "heading_1",
                "heading_1": {"rich_text": [{"type": "text", "text": {"content": "??? Technical Skills"}}]}
            },
            {
                "object": "block",
                "type": "column_list",
                "column_list": {"children": []}
            }
        ]
    )
    
    # ½ºÅ³ Å×ÀÌºí
    notion.blocks.children.append(
        block_id=page_id,
        children=[
            {
                "object": "block",
                "type": "table",
                "table": {
                    "table_width": 2,
                    "has_column_header": True,
                    "has_row_header": False,
                    "children": [
                        {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Category"}}], [{"type": "text", "text": {"content": "Technologies"}}]]}},
                        {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "MCU"}}], [{"type": "text", "text": {"content": "STM32 F1/F4/F7/H7/G0, ARM Cortex-M"}}]]}},
                        {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "RTOS"}}], [{"type": "text", "text": {"content": "FreeRTOS (Task, Queue, Semaphore, Mutex)"}}]]}},
                        {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Languages"}}], [{"type": "text", "text": {"content": "C (Expert), C++, Python"}}]]}},
                        {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Protocols"}}], [{"type": "text", "text": {"content": "Modbus RTU/ASCII/TCP, UART, SPI, I2C, CAN"}}]]}},
                        {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Tools"}}], [{"type": "text", "text": {"content": "STM32CubeIDE, Altium Designer, Git, Logic Analyzer"}}]]}},
                        {"type": "table_row", "table_row": {"cells": [[{"type": "text", "text": {"content": "Control"}}], [{"type": "text", "text": {"content": "PID, Auto-tuning, Kalman Filter, State Machine"}}]]}}
                    ]
                }
            }
        ]
    )
    
    # ¼º°ú ¼½¼Ç
    notion.blocks.children.append(
        block_id=page_id,
        children=[
            {"object": "block", "type": "divider", "divider": {}},
            {
                "object": "block",
                "type": "heading_1",
                "heading_1": {"rich_text": [{"type": "text", "text": {"content": "?? Key Achievements"}}]}
            },
            {
                "object": "block",
                "type": "bulleted_list_item",
                "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "12+ ¾ç»ê Æß¿ş¾î ¹èÆ÷, Ä¡¸íÀû ¹ö±× Zero"}}]}
            },
            {
                "object": "block",
                "type": "bulleted_list_item",
                "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "99.8%+ ½Ã½ºÅÛ °¡µ¿·ü ´Ş¼º"}}]}
            },
            {
                "object": "block",
                "type": "bulleted_list_item",
                "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "PID ¶óÀÌºê·¯¸®, IAP Bootloader - ÇÁ·ÎÁ§Æ® Àü¹İ Àç»ç¿ë"}}]}
            },
            {
                "object": "block",
                "type": "bulleted_list_item",
                "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": "20+ ±â¼ú ¹®¼­, 15+ »ç¿ëÀÚ ¸Å´º¾ó ÀÛ¼º"}}]}
            }
        ]
    )
    
    print(f"\n? Æ÷Æ®Æú¸®¿À »ı¼º ¿Ï·á!")
    print(f"?? ÆäÀÌÁö URL: https://notion.so/{page_id.replace('-', '')}")
    
    return page_id


def create_project_toggle(page_id, title, content_lines):
    """ÇÁ·ÎÁ§Æ® Åä±Û ºí·Ï »ı¼º"""
    # Toggle ºí·Ï »ı¼º
    toggle_block = notion.blocks.children.append(
        block_id=page_id,
        children=[
            {
                "object": "block",
                "type": "toggle",
                "toggle": {
                    "rich_text": [{"type": "text", "text": {"content": title}, "annotations": {"bold": True}}],
                    "color": "default",
                    "children": []
                }
            }
        ]
    )
    
    # Toggle ³»ºÎ¿¡ ÄÜÅÙÃ÷ Ãß°¡
    toggle_id = toggle_block["results"][0]["id"]
    
    children = []
    for line in content_lines:
        if line.startswith("?"):
            children.append({
                "object": "block",
                "type": "bulleted_list_item",
                "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": line[2:]}}]}
            })
        elif line.startswith("??") or line.startswith("??") or line.startswith("??") or line.startswith("??") or line.startswith("??"):
            children.append({
                "object": "block",
                "type": "paragraph",
                "paragraph": {"rich_text": [{"type": "text", "text": {"content": line}, "annotations": {"bold": True}}]}
            })
        elif line == "":
            continue
        else:
            children.append({
                "object": "block",
                "type": "paragraph",
                "paragraph": {"rich_text": [{"type": "text", "text": {"content": line}}]}
            })
    
    if children:
        notion.blocks.children.append(block_id=toggle_id, children=children)


def main():
    print("=" * 50)
    print("?? Á¶¶ôÇö Æ÷Æ®Æú¸®¿À ³ë¼Ç ÀÚµ¿ »ı¼º")
    print("=" * 50)
    
    # ÆäÀÌÁö °Ë»ö
    print("\n?? '»õ ÆäÀÌÁö' °Ë»ö Áß...")
    parent_id = search_page("»õ ÆäÀÌÁö")
    
    if parent_id:
        print(f"  ? ÆäÀÌÁö ¹ß°ß: {parent_id}")
        
        # Æ÷Æ®Æú¸®¿À »ı¼º
        create_portfolio_content(parent_id)
    else:
        print("  ? '»õ ÆäÀÌÁö'¸¦ Ã£À» ¼ö ¾ø½À´Ï´Ù.")
        print("\n?? ÇØ°á ¹æ¹ı:")
        print("  1. ³ë¼Ç¿¡¼­ '»õ ÆäÀÌÁö' ¿­±â")
        print("  2. ¿ìÃø »ó´Ü '...' ¡æ 'Connections' ¡æ Integration Ãß°¡")
        print("  3. ½ºÅ©¸³Æ® ´Ù½Ã ½ÇÇà")


if __name__ == "__main__":
    main()

from notion_client import Client

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
notion = Client(auth=NOTION_TOKEN)

print("=" * 50)
print("Á¢±Ù °¡´ÉÇÑ ¸ğµç ÆäÀÌÁö °Ë»ö")
print("=" * 50)

response = notion.search(query="", filter={"property": "object", "value": "page"})
results = response.get("results", [])
print(f"\nÁ¢±Ù °¡´ÉÇÑ ÆäÀÌÁö: {len(results)}°³\n")

for r in results[:15]:
    title = "N/A"
    if r.get("properties"):
        for v in r["properties"].values():
            if v.get("title") and v["title"]:
                title = v["title"][0]["plain_text"]
                break
    if r.get("child_page"):
        title = r["child_page"].get("title", title)
    
    page_id = r["id"]
    print(f"  - {title}")
    print(f"    ID: {page_id}")
    print()

# -*- coding: utf-8 -*-
"""
?¸ì…˜ ?„ì²´ ê²€?‰ìœ¼ë¡?"ì¡°ë½??Rakeen Jo" ?˜ì´ì§€ ì°¾ê¸°
"""

from notion_client import Client

# TODO: ?˜ê²½ ë³€?˜ì—??? í° ?½ê¸°
NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
notion = Client(auth=NOTION_TOKEN)

print("=" * 60)
print("?¸ì…˜ ?„ì²´ ê²€?? 'ì¡°ë½?? ?ëŠ” 'Rakeen' ?¬í•¨ ?˜ì´ì§€")
print("=" * 60)

# ê²€??
response = notion.search(query="ì¡°ë½??, filter={"property": "object", "value": "page"})
results = response.get("results", [])

print(f"\n'ì¡°ë½?? ê²€??ê²°ê³¼: {len(results)}ê°?n")

target_pages = []

for r in results:
    title = "N/A"
    if r.get("properties"):
        for v in r["properties"].values():
            if v.get("title") and v["title"]:
                title = v["title"][0]["plain_text"]
                break
    if r.get("child_page"):
        title = r["child_page"].get("title", title)
    
    page_id = r["id"]
    print(f"?“„ {title}")
    print(f"   ID: {page_id}")
    
    if 'Rakeen' in title or 'ì¡°ë½?? in title:
        target_pages.append((page_id, title))
    print()

# Rakeen?¼ë¡œ??ê²€??
response2 = notion.search(query="Rakeen", filter={"property": "object", "value": "page"})
results2 = response2.get("results", [])

print(f"\n'Rakeen' ê²€??ê²°ê³¼: {len(results2)}ê°?n")

for r in results2:
    title = "N/A"
    if r.get("properties"):
        for v in r["properties"].values():
            if v.get("title") and v["title"]:
                title = v["title"][0]["plain_text"]
                break
    if r.get("child_page"):
        title = r["child_page"].get("title", title)
    
    page_id = r["id"]
    
    # ì¤‘ë³µ ì²´í¬
    if page_id not in [p[0] for p in target_pages]:
        print(f"?“„ {title}")
        print(f"   ID: {page_id}")
        
        if 'Rakeen' in title or 'ì¡°ë½?? in title:
            target_pages.append((page_id, title))
        print()

print("=" * 60)
print(f"ì´?{len(target_pages)}ê°œì˜ ?€ê²??˜ì´ì§€ ë°œê²¬")
print("=" * 60)

for page_id, title in target_pages:
    print(f"\n?“Œ ?˜ì´ì§€: {title}")
    print(f"   ID: {page_id}")
    print(f"   URL: https://notion.so/{page_id.replace('-', '')}")

# ì²?ë²ˆì§¸ ?€ê²??˜ì´ì§€???ì„¸ ?´ìš© ê°€?¸ì˜¤ê¸?
if target_pages:
    target_id, target_title = target_pages[0]
    print(f"\n{'='*60}")
    print(f"'{target_title}' ?˜ì´ì§€ ?´ìš© ë¶„ì„")
    print("="*60)
    
    page_blocks = notion.blocks.children.list(block_id=target_id, page_size=100)
    
    print(f"\nì´?{len(page_blocks['results'])}ê°œì˜ ë¸”ë¡\n")
    
    for idx, block in enumerate(page_blocks['results'], 1):
        block_type = block['type']
        
        print(f"\n[{idx}] {block_type.upper()}")
        
        # ?ìŠ¤??ì¶”ì¶œ
        text_content = ""
        if block_type == 'paragraph':
            texts = block['paragraph']['rich_text']
            text_content = ''.join([t['plain_text'] for t in texts]) if texts else ""
        
        elif block_type.startswith('heading_'):
            texts = block[block_type]['rich_text']
            text_content = ''.join([t['plain_text'] for t in texts]) if texts else ""
        
        elif block_type == 'bulleted_list_item':
            texts = block['bulleted_list_item']['rich_text']
            text_content = '??' + ''.join([t['plain_text'] for t in texts]) if texts else ""
        
        elif block_type == 'numbered_list_item':
            texts = block['numbered_list_item']['rich_text']
            text_content = ''.join([t['plain_text'] for t in texts]) if texts else ""
        
        elif block_type == 'callout':
            texts = block['callout']['rich_text']
            icon = block['callout'].get('icon', {})
            emoji = icon.get('emoji', '?’¡') if icon.get('type') == 'emoji' else '?’¡'
            text_content = f"{emoji} " + ''.join([t['plain_text'] for t in texts]) if texts else ""
        
        elif block_type == 'toggle':
            texts = block['toggle']['rich_text']
            text_content = '?”½ ' + ''.join([t['plain_text'] for t in texts]) if texts else ""
        
        elif block_type == 'code':
            texts = block['code']['rich_text']
            language = block['code']['language']
            code_text = ''.join([t['plain_text'] for t in texts]) if texts else ""
            text_content = f"[{language}]\n{code_text[:300]}..."
        
        elif block_type == 'divider':
            text_content = "?€" * 50
        
        elif block_type == 'table':
            text_content = f"?“Š ?Œì´ë¸?({block['table']['table_width']}??"
        
        if text_content:
            print(f"    {text_content}")
    
    if page_blocks.get('has_more'):
        print("\n? ï¸ ??ë§ì? ë¸”ë¡ ?ˆìŒ (?˜ì´ì§€?¤ì´???„ìš”)")

"""
Notion API Å×½ºÆ® ½ºÅ©¸³Æ®
"""
from notion_client import Client

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
notion = Client(auth=NOTION_TOKEN)

print("=" * 50)
print("Notion API ¿¬°á Å×½ºÆ®")
print("=" * 50)

try:
    response = notion.search(query="»õ ÆäÀÌÁö")
    print(f"\n°Ë»ö °á°ú: {len(response.get('results', []))}°³")
    
    for i, r in enumerate(response.get('results', [])):
        print(f"\n[{i+1}] ID: {r['id']}")
        print(f"    Object: {r.get('object')}")
        print(f"    Type: {r.get('type', 'N/A')}")
        
        # Á¦¸ñ ÃßÃâ ½Ãµµ
        if r.get('properties'):
            for k, v in r['properties'].items():
                if v.get('type') == 'title' and v.get('title'):
                    title_text = v['title'][0]['plain_text'] if v['title'] else 'N/A'
                    print(f"    Title: {title_text}")
        
        # child_pageÀÎ °æ¿ì
        if r.get('child_page'):
            print(f"    Child Page Title: {r['child_page'].get('title', 'N/A')}")
            
except Exception as e:
    print(f"\n¿À·ù ¹ß»ı: {type(e).__name__}")
    print(f"¸Ş½ÃÁö: {e}")
    
    if "Could not find" in str(e) or "unauthorized" in str(e).lower():
        print("\nÇØ°á ¹æ¹ı:")
        print("1. ³ë¼Ç¿¡¼­ '»õ ÆäÀÌÁö' ¿­±â")
        print("2. ¿ìÃø »ó´Ü '...' Å¬¸¯")
        print("3. 'Connections' ¶Ç´Â '¿¬°á' Å¬¸¯")
        print("4. Integration °Ë»öÇÏ¿© Ãß°¡")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ì¡°ë½???¬íŠ¸?´ë¦¬??- ?¸ì…˜ ?…ë°?´íŠ¸ ?¤í¬ë¦½íŠ¸ v2
?„ë¡œ?íŠ¸ ë³´ì™„: Titrator, JIG Board, SSC ì¶”ê?
FPGA, LE_Laser ?œì™¸
Lux = Nu-2000 ?µí•©
"""

import os
from notion_client import Client

# Notion API ?¤ì •
# TODO: ?˜ê²½ ë³€??NOTION_TOKEN ?¤ì • ?„ìš”
NOTION_TOKEN = os.environ.get("NOTION_TOKEN", "YOUR_NOTION_TOKEN_HERE")
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"

def create_text_block(text, bold=False, color="default"):
    """?ìŠ¤??ë¦¬ì¹˜ ?ìŠ¤???ì„±"""
    annotations = {"bold": bold, "italic": False, "strikethrough": False, "underline": False, "code": False, "color": color}
    return {"type": "text", "text": {"content": text}, "annotations": annotations}

def create_heading1(text):
    return {"type": "heading_1", "heading_1": {"rich_text": [create_text_block(text)], "color": "default"}}

def create_heading2(text):
    return {"type": "heading_2", "heading_2": {"rich_text": [create_text_block(text)], "color": "default"}}

def create_heading3(text):
    return {"type": "heading_3", "heading_3": {"rich_text": [create_text_block(text)], "color": "default"}}

def create_paragraph(text_list):
    rich_texts = []
    for item in text_list:
        if isinstance(item, str):
            rich_texts.append(create_text_block(item))
        elif isinstance(item, dict):
            rich_texts.append(create_text_block(item.get("text", ""), item.get("bold", False), item.get("color", "default")))
    return {"type": "paragraph", "paragraph": {"rich_text": rich_texts, "color": "default"}}

def create_bullet(text_list):
    rich_texts = []
    for item in text_list:
        if isinstance(item, str):
            rich_texts.append(create_text_block(item))
        elif isinstance(item, dict):
            rich_texts.append(create_text_block(item.get("text", ""), item.get("bold", False)))
    return {"type": "bulleted_list_item", "bulleted_list_item": {"rich_text": rich_texts, "color": "default"}}

def create_divider():
    return {"type": "divider", "divider": {}}

def create_callout(text, emoji="?’¡"):
    return {"type": "callout", "callout": {"rich_text": [create_text_block(text)], "icon": {"type": "emoji", "emoji": emoji}, "color": "blue_background"}}

def create_code_block(code, language="c"):
    return {"type": "code", "code": {"rich_text": [create_text_block(code)], "language": language}}

def create_table_row(cells):
    """?Œì´ë¸????ì„±"""
    return {"type": "table_row", "table_row": {"cells": [[create_text_block(cell)] for cell in cells]}}

def create_toggle(title, children):
    """? ê? ë¸”ë¡ ?ì„±"""
    return {"type": "toggle", "toggle": {"rich_text": [create_text_block(title)], "color": "default", "children": children}}

def main():
    notion = Client(auth=NOTION_TOKEN)
    
    # ë©”ì¸ ?˜ì´ì§€ ?ì„±
    print("?? ì¡°ë½???¬íŠ¸?´ë¦¬???¥ìƒ ë²„ì „ v2 ?ì„± ?œì‘...")
    
    page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={
            "title": {"title": [{"text": {"content": "?“˜ ì¡°ë½??- Senior Embedded Engineer Portfolio (Enhanced v2)"}}]}
        },
        icon={"type": "emoji", "emoji": "?‘¨?ğŸ’?},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1200"}}
    )
    page_id = page["id"]
    print(f"??ë©”ì¸ ?˜ì´ì§€ ?ì„±: {page_id}")
    
    blocks = []
    
    # ============ ?„ë¡œ???¹ì…˜ ============
    blocks.append(create_callout("9?„ì°¨ ?Œë¡œ?¤ê³„/?Œì›¨???”ì??ˆì–´ | ë¶„ì„ê¸°ê¸° ?„ë¬¸ | STM32/ARM ?„ë² ?”ë“œ ?œìŠ¤??, "?¯"))
    blocks.append(create_paragraph([]))
    
    # ?°ë½ì²??•ë³´
    blocks.append(create_heading2("?“ Contact Information"))
    blocks.append(create_bullet(["?“§ Email: rakguard@gmail.com"]))
    blocks.append(create_bullet(["?“± Phone: 010-3233-5365"]))
    blocks.append(create_bullet(["?”— GitHub: github.com/gari210404"]))
    blocks.append(create_divider())
    
    # ============ ê¸°ìˆ  ?¤íƒ ?¹ì…˜ ============
    blocks.append(create_heading1("?› ï¸?Technical Skills"))
    blocks.append(create_paragraph([]))
    
    # ?„ë¡œê·¸ë˜ë°??¸ì–´
    blocks.append(create_heading3("?’» Programming Languages"))
    blocks.append(create_bullet([{"text": "C/C++ ", "bold": True}, "- ?„ë² ?”ë“œ ?Œì›¨??ê°œë°œ ?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "Python ", "bold": True}, "- ?ë™???¤í¬ë¦½íŠ¸, ?ŒìŠ¤???„êµ¬ ?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "Assembly ", "bold": True}, "- ARM Cortex-M ìµœì ???…â˜…?…â˜†??]))
    blocks.append(create_paragraph([]))
    
    # MCU/ê°œë°œ?˜ê²½
    blocks.append(create_heading3("?”§ MCU & Development Tools"))
    blocks.append(create_bullet([{"text": "STM32 Series ", "bold": True}, "- F4, F7, G0 ?œë¦¬ì¦??…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "STM32CubeMX/IDE ", "bold": True}, "- HAL ?œë¼?´ë²„, ì½”ë“œ ?ì„± ?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "Renesas RA ", "bold": True}, "- e2 studio, FSP ?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "FreeRTOS ", "bold": True}, "- ë©€?°íƒœ?¤í¬ ?œìŠ¤???¤ê³„ ?…â˜…?…â˜…??]))
    blocks.append(create_paragraph([]))
    
    # ?µì‹  ?„ë¡œ? ì½œ
    blocks.append(create_heading3("?“¡ Communication Protocols"))
    blocks.append(create_bullet([{"text": "UART/RS232/RS485 ", "bold": True}, "?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "SPI/I2C ", "bold": True}, "- ?¼ì„œ, ?”ìŠ¤?Œë ˆ???¸í„°?˜ì´???…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "Ethernet/TCP/UDP ", "bold": True}, "- LwIP ?¤íƒ, ?ê²© ?œì–´ ?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "Modbus RTU/TCP ", "bold": True}, "- ?°ì—…???„ë¡œ? ì½œ ?…â˜…?…â˜…??]))
    blocks.append(create_paragraph([]))
    
    # ?Œë¡œ ?¤ê³„
    blocks.append(create_heading3("??Hardware Design"))
    blocks.append(create_bullet([{"text": "OrCAD/Cadence ", "bold": True}, "- ?Œë¡œ???¤ê³„ ?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "PADS ", "bold": True}, "- PCB ?„íŠ¸?Œí¬, ?¤ì¸µ ë³´ë“œ ?…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "?„ë‚ ë¡œê·¸ ?Œë¡œ ", "bold": True}, "- Op-Amp, ADC, ?¼ì„œ ?¸í„°?˜ì´???…â˜…?…â˜…??]))
    blocks.append(create_bullet([{"text": "?„ì› ?¤ê³„ ", "bold": True}, "- SMPS, LDO, ?¸ì´ì¦??„í„°ë§??…â˜…?…â˜…??]))
    
    blocks.append(create_divider())
    
    # ============ ?„ë¡œ?íŠ¸ ?¹ì…˜ ============
    blocks.append(create_heading1("?? Major Projects"))
    blocks.append(create_paragraph([]))
    
    # ==================== 1. L-LPC ?„ë¡œ?íŠ¸ ====================
    blocks.append(create_heading2("1. L-LPC (Laser Particle Counter) System"))
    blocks.append(create_callout("?ˆì´?€ ?Œí‹°??ì¹´ìš´??- ë°˜ë„ì²??´ë¦°ë£??…ì ì¸¡ì • ?¥ë¹„ (HW + FW ?„ë‹´)", "?”¬"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“‹ ?„ë¡œ?íŠ¸ ê°œìš”"))
    blocks.append(create_bullet(["ê¸°ê°„: 2022.03 ~ 2024.12 (??2??10ê°œì›”)"]))
    blocks.append(create_bullet(["??• : ?˜ë“œ?¨ì–´ ?¤ê³„ ë°??Œì›¨??ê°œë°œ ?„ë‹´"]))
    blocks.append(create_bullet(["MCU: STM32F407VGT6 (168MHz, ARM Cortex-M4)"]))
    blocks.append(create_bullet(["RTOS: FreeRTOS ê¸°ë°˜ ë©€?°íƒœ?¤í¬ ?¤ê³„"]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?™ï¸ ?˜ë“œ?¨ì–´ ?„í‚¤?ì²˜"))
    blocks.append(create_bullet([{"text": "Main Board: ", "bold": True}, "STM32F407 + Ethernet PHY (LAN8720)"]))
    blocks.append(create_bullet([{"text": "ADC Module: ", "bold": True}, "24-bit ê³ ì •ë°€ ADC (ADS1256) - ê´‘ì„¼??? í˜¸ ì²˜ë¦¬"]))
    blocks.append(create_bullet([{"text": "DAC Module: ", "bold": True}, "16-bit DAC (DAC8568) - ?ˆì´?€ ì¶œë ¥ ?œì–´"]))
    blocks.append(create_bullet([{"text": "Power: ", "bold": True}, "Multi-rail SMPS (24V??2V??V??.3V), ?€?¸ì´ì¦??¤ê³„"]))
    blocks.append(create_bullet([{"text": "Interface: ", "bold": True}, "RS232, RS485, Ethernet, USB"]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?’» ?Œì›¨??êµ¬í˜„ ?´ìš©"))
    blocks.append(create_bullet([{"text": "?¤ì‹œê°??…ì ê³„ìˆ˜: ", "bold": True}, "APD ? í˜¸ ì²˜ë¦¬, ?¤ì±„???¬ì´ì¦?ë¶„ë¥˜"]))
    blocks.append(create_bullet([{"text": "TCP/IP ?µì‹ : ", "bold": True}, "LwIP ?¤íƒ, ?ê²© ?œì–´ ë°??°ì´???„ì†¡"]))
    blocks.append(create_bullet([{"text": "Modbus RTU/TCP: ", "bold": True}, "?°ì—… ?œì? ?„ë¡œ? ì½œ êµ¬í˜„"]))
    blocks.append(create_bullet([{"text": "?°ì´??ë¡œê¹…: ", "bold": True}, "Flash ?€?? ?´ë ¥ ê´€ë¦?]))
    blocks.append(create_bullet([{"text": "?ê? ì§„ë‹¨: ", "bold": True}, "?¼ì„œ ?íƒœ ëª¨ë‹ˆ?°ë§, ?ëŸ¬ ?¸ë“¤ë§?]))
    blocks.append(create_paragraph([]))
    
    # L-LPC ì½”ë“œ ?˜í”Œ
    blocks.append(create_heading3("?“ ì½”ë“œ ?˜í”Œ - ?´ë”???µì‹  ì´ˆê¸°??))
    lpc_code = '''// L-LPC Ethernet Initialization (LwIP + FreeRTOS)
void init_ethernet(void)
{
    // Ethernet PHY Reset & Init
    HAL_ETH_Init(&heth);
    
    // LwIP Stack Init with RTOS
    tcpip_init(NULL, NULL);
    
    // Network Interface Setup
    IP4_ADDR(&ipaddr, 192, 168, 1, 100);
    IP4_ADDR(&netmask, 255, 255, 255, 0);
    IP4_ADDR(&gw, 192, 168, 1, 1);
    
    netif_add(&netif, &ipaddr, &netmask, &gw, 
              NULL, &ethernetif_init, &tcpip_input);
    netif_set_default(&netif);
    netif_set_up(&netif);
    
    // TCP Server for Remote Control
    tcp_server_init(TCP_PORT_CMD);
    
    // UDP Server for Measurement Data
    udp_server_init(UDP_PORT_MEAS);
}'''
    blocks.append(create_code_block(lpc_code, "c"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?”§ ?”ë²„ê¹?ê²½í—˜"))
    blocks.append(create_bullet([{"text": "ADC ?¸ì´ì¦??´ìŠˆ: ", "bold": True}, "?„ì› ë¶„ë¦¬ ?¤ê³„, ?”ì????„ë‚ ë¡œê·¸ GND ë¶„ë¦¬ë¡??´ê²°"]))
    blocks.append(create_bullet([{"text": "Ethernet ?¨í‚· ?ì‹¤: ", "bold": True}, "DMA ë²„í¼ ?¬ê¸° ìµœì ?? ?¸í„°?½íŠ¸ ?°ì„ ?œìœ„ ì¡°ì •"]))
    blocks.append(create_bullet([{"text": "FreeRTOS ?¤íƒ ?¤ë²„?Œë¡œ?? ", "bold": True}, "?œìŠ¤?¬ë³„ ?¤íƒ ?¬ìš©??ë¶„ì„ ë°??¬í• ??]))
    
    blocks.append(create_divider())
    
    # ==================== 2. L-Titrator ?„ë¡œ?íŠ¸ ====================
    blocks.append(create_heading2("2. L-Titrator (?ë™ ?ì •ê¸?"))
    blocks.append(create_callout("?”í•™ ë¶„ì„???ë™ ?ì •ê¸?- ?„ì „ ? ê·œ ê°œë°œ (HW + FW ?¤ê³„)", "?§ª"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“‹ ?„ë¡œ?íŠ¸ ê°œìš”"))
    blocks.append(create_bullet(["ê¸°ê°„: 2021.10 ~ 2022.09 (??1??"]))
    blocks.append(create_bullet(["??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ ?„ë‹´"]))
    blocks.append(create_bullet(["MCU: STM32F407VGT6"]))
    blocks.append(create_bullet(["ë³´ë“œ êµ¬ì„±: Main Board + MCU Module + Relay Module"]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?™ï¸ ?˜ë“œ?¨ì–´ ?¤ê³„"))
    blocks.append(create_bullet([{"text": "Main Board: ", "bold": True}, "ëª¨í„° ?œë¼?´ë²„ (L6203), ë°¸ë¸Œ ?œì–´, ?¼ì„œ ?¸í„°?˜ì´??]))
    blocks.append(create_bullet([{"text": "MCU Module: ", "bold": True}, "STM32F407 + Ethernet PHY + SD Card"]))
    blocks.append(create_bullet([{"text": "Relay Module: ", "bold": True}, "8ì±„ë„ ë¦´ë ˆ?? ?¤í† ?˜í”Œ???œì–´"]))
    blocks.append(create_bullet([{"text": "Syringe Pump: ", "bold": True}, "?¤í…Œ??ëª¨í„° ?•ë? ?œì–´"]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?’» ?Œì›¨??êµ¬í˜„"))
    blocks.append(create_bullet([{"text": "?œë¦°ì§€ ?Œí”„ ?œì–´: ", "bold": True}, "?¤í…Œ??ëª¨í„° ê°€ê°ì† ?„ë¡œ?Œì¼, Î¼L ?¨ìœ„ ?•ë???]))
    blocks.append(create_bullet([{"text": "pH ì¸¡ì •: ", "bold": True}, "ê³ ì •ë°€ ADC ? í˜¸ ì²˜ë¦¬, ?¨ë„ ë³´ìƒ"]))
    blocks.append(create_bullet([{"text": "?ë™ ?ì • ?Œê³ ë¦¬ì¦˜: ", "bold": True}, "ì¢…ì  ê²€ì¶? ?¤ë‹¨ê³??ì •"]))
    blocks.append(create_bullet([{"text": "FreeRTOS ?œìŠ¤?? ", "bold": True}, "?Œí”„ ?œì–´, ?¼ì„œ ëª¨ë‹ˆ?°ë§, ?µì‹  ë¶„ë¦¬"]))
    blocks.append(create_paragraph([]))
    
    # Titrator ì½”ë“œ ?˜í”Œ
    blocks.append(create_heading3("?“ ì½”ë“œ ?˜í”Œ - ?œë¦°ì§€ ?Œí”„ ?œì–´"))
    titrator_code = '''// L-Titrator Syringe Pump Motor Control
void syr_pump_control(SRG_PUMP_STR* pump, uint16_t target_ul)
{
    uint32_t steps = (uint32_t)(target_ul * STEPS_PER_UL);
    
    // Acceleration profile
    for(uint32_t i = 0; i < steps; i++) {
        if(i < ACCEL_STEPS) {
            // Acceleration phase
            pump->delay_us = DELAY_MAX - (i * DELAY_STEP);
        } else if(i > (steps - ACCEL_STEPS)) {
            // Deceleration phase  
            pump->delay_us = DELAY_MIN + ((steps - i) * DELAY_STEP);
        } else {
            // Constant speed
            pump->delay_us = DELAY_MIN;
        }
        
        // Step pulse generation
        HAL_GPIO_WritePin(pump->step_port, pump->step_pin, GPIO_PIN_SET);
        delay_us(1);
        HAL_GPIO_WritePin(pump->step_port, pump->step_pin, GPIO_PIN_RESET);
        delay_us(pump->delay_us);
    }
}'''
    blocks.append(create_code_block(titrator_code, "c"))
    
    blocks.append(create_divider())
    
    # ==================== 3. Nu-2000 / Lux ?„ë¡œ?íŠ¸ ====================
    blocks.append(create_heading2("3. Nu-2000 / Lux (ë¶„ê´‘ê´‘ë„ê³?"))
    blocks.append(create_callout("UV-Vis ë¶„ê´‘ê´‘ë„ê³??œìŠ¤??- Lux??Nu-2000???„ì† ë²„ì „", "?”¬"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“‹ ?„ë¡œ?íŠ¸ ê°œìš”"))
    blocks.append(create_bullet(["ê¸°ê°„: 2020.01 ~ 2023.12 (Nu-2000 ??Lux ì§„í™”)"]))
    blocks.append(create_bullet(["??• : ?Œì›¨??ê°œë°œ ë°?? ì?ë³´ìˆ˜"]))
    blocks.append(create_bullet(["MCU: STM32F407 + TouchGFX ê¸°ë°˜ GUI"]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?’» ?Œì›¨??êµ¬í˜„"))
    blocks.append(create_bullet([{"text": "ê´‘í•™ ?œìŠ¤???œì–´: ", "bold": True}, "?¤í…Œ??ëª¨í„° (?Œì ˆê²©ì), ê´‘ì› ?„í™˜"]))
    blocks.append(create_bullet([{"text": "ê´?ê²€ì¶? ", "bold": True}, "Photodiode ? í˜¸ ì¦í­, ADC ?°ì´??ì²˜ë¦¬"]))
    blocks.append(create_bullet([{"text": "?¤í™?¸ëŸ¼ ?¤ìº”: ", "bold": True}, "?Œì¥ë³??¡ê´‘??ì¸¡ì •, ?¼í¬ ê²€ì¶?]))
    blocks.append(create_bullet([{"text": "TouchGFX GUI: ", "bold": True}, "?¤ì‹œê°?ê·¸ë˜?? ?¬ìš©???¸í„°?˜ì´??]))
    blocks.append(create_bullet([{"text": "?°ì´???€?? ", "bold": True}, "USB, SD Card ?´ë³´?´ê¸°"]))
    
    blocks.append(create_divider())
    
    # ==================== 4. Psi-1000/3000 ?„ë¡œ?íŠ¸ ====================
    blocks.append(create_heading2("4. Psi-1000 / Psi-3000 (??˜¨ì¡??œìŠ¤??"))
    blocks.append(create_callout("?•ë? ?¨ë„ ?œì–´ ?œìŠ¤??- PID ?Œê³ ë¦¬ì¦˜ ê¸°ë°˜", "?Œ¡ï¸?))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“‹ ?„ë¡œ?íŠ¸ ê°œìš”"))
    blocks.append(create_bullet(["ê¸°ê°„: 2019.06 ~ 2021.12"]))
    blocks.append(create_bullet(["??• : ?Œì›¨??ê°œë°œ"]))
    blocks.append(create_bullet(["MCU: STM32F407VGT6"]))
    blocks.append(create_bullet(["?¹ì§•: Â±0.01Â°C ?•ë??? ?€??ì±„ë„ ?œì–´"]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?’» PID ?¨ë„ ?œì–´ ?Œê³ ë¦¬ì¦˜"))
    blocks.append(create_bullet([{"text": "PID ?Œë¼ë¯¸í„°: ", "bold": True}, "Kp, Ki, Kd ?ë™ ?œë‹ ê¸°ëŠ¥"]))
    blocks.append(create_bullet([{"text": "?¨ë„ ?¼ì„œ: ", "bold": True}, "PT100 RTD, ê³ ì •ë°€ ADC (24-bit)"]))
    blocks.append(create_bullet([{"text": "ì¶œë ¥ ?œì–´: ", "bold": True}, "PWM Heater + Peltier Cooler"]))
    blocks.append(create_bullet([{"text": "?ˆì „ ê¸°ëŠ¥: ", "bold": True}, "ê³¼ì—´ ë³´í˜¸, ?¼ì„œ ?¨ì„  ê°ì?"]))
    blocks.append(create_paragraph([]))
    
    # PID ì½”ë“œ ?˜í”Œ
    blocks.append(create_heading3("?“ ì½”ë“œ ?˜í”Œ - PID ?¨ë„ ?œì–´"))
    pid_code = '''// Psi-1000/3000 PID Temperature Control
typedef struct {
    float Kp, Ki, Kd;           // PID gains
    float setpoint;             // Target temperature
    float integral;             // Integral accumulator
    float prev_error;           // Previous error
    float output;               // Control output (0~100%)
} PID_CTRL_STR;

float pid_calculate(PID_CTRL_STR* pid, float current_temp)
{
    float error = pid->setpoint - current_temp;
    
    // Proportional
    float P = pid->Kp * error;
    
    // Integral with anti-windup
    pid->integral += error * DT;
    if(pid->integral > INTEGRAL_MAX) pid->integral = INTEGRAL_MAX;
    if(pid->integral < INTEGRAL_MIN) pid->integral = INTEGRAL_MIN;
    float I = pid->Ki * pid->integral;
    
    // Derivative
    float D = pid->Kd * (error - pid->prev_error) / DT;
    pid->prev_error = error;
    
    // Output limiting
    pid->output = P + I + D;
    if(pid->output > 100.0f) pid->output = 100.0f;
    if(pid->output < 0.0f) pid->output = 0.0f;
    
    return pid->output;
}'''
    blocks.append(create_code_block(pid_code, "c"))
    
    blocks.append(create_divider())
    
    # ==================== 5. MS (Mass Spectrometer) ?„ë¡œ?íŠ¸ ====================
    blocks.append(create_heading2("5. MS (Mass Spectrometer) - Aston ?œë¦¬ì¦?))
    blocks.append(create_callout("ì§ˆëŸ‰ ë¶„ì„ê¸??œìŠ¤??- ?¤ìˆ˜ ë³´ë“œ ê°œë°œ (Main, Interface, Sensor ??", "?—ï¸"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“‹ ?„ë¡œ?íŠ¸ ê°œìš”"))
    blocks.append(create_bullet(["ê¸°ê°„: 2019.01 ~ 2024.12 (ì§€??ê°œë°œ)"]))
    blocks.append(create_bullet(["??• : ?˜ë“œ?¨ì–´ ?¤ê³„ (?¤ìˆ˜ ëª¨ë“ˆ ë³´ë“œ)"]))
    blocks.append(create_bullet(["ë³´ë“œ: Main Board, Interface Board, HV Power, Sensor Board ??]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?™ï¸ ê°œë°œ ë³´ë“œ ëª©ë¡"))
    blocks.append(create_bullet([{"text": "Main Board: ", "bold": True}, "?œìŠ¤???œì–´, ?µì‹  ?¸í„°?˜ì´??]))
    blocks.append(create_bullet([{"text": "Interface Board: ", "bold": True}, "?¸ë? ?¥ì¹˜ ?°ê²°, ? í˜¸ ì»¨ë””?”ë‹"]))
    blocks.append(create_bullet([{"text": "HV Power Board: ", "bold": True}, "ê³ ì „???„ì› (??kV), ?ˆì „ ?Œë¡œ"]))
    blocks.append(create_bullet([{"text": "Sensor Board: ", "bold": True}, "?´ì˜¨ ê²€ì¶œê¸° ?¸í„°?˜ì´?? ë¯¸ì„¸? í˜¸ ì¦í­"]))
    blocks.append(create_bullet([{"text": "Vacuum Gauge: ", "bold": True}, "ì§„ê³µ??ì¸¡ì •, ?Œí”„ ?œì–´"]))
    
    blocks.append(create_divider())
    
    # ==================== 6. SSC (Mantis) ?„ë¡œ?íŠ¸ ====================
    blocks.append(create_heading2("6. SSC - Mantis (?Œí˜• ë¶„ê´‘ê³?"))
    blocks.append(create_callout("?Œí˜• ë¶„ê´‘ê³??œìŠ¤??- ? ê·œ ê°œë°œ ì¤?(2024~)", "?“Š"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“‹ ?„ë¡œ?íŠ¸ ê°œìš”"))
    blocks.append(create_bullet(["ê¸°ê°„: 2024.03 ~ ?„ì¬ (ì§„í–‰ ì¤?"]))
    blocks.append(create_bullet(["??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ"]))
    blocks.append(create_bullet(["MCU: STM32G0 ?œë¦¬ì¦?]))
    blocks.append(create_bullet(["?¹ì§•: ?€?„ë ¥, ?Œí˜•?? RS485 Modbus"]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?™ï¸ ?œìŠ¤??êµ¬ì„±"))
    blocks.append(create_bullet([{"text": "Main Board: ", "bold": True}, "STM32G0, ?„ì› ê´€ë¦? ?µì‹ "]))
    blocks.append(create_bullet([{"text": "Sensor Board: ", "bold": True}, "ê´??¼ì„œ ?´ë ˆ?? ? í˜¸ ì²˜ë¦¬"]))
    blocks.append(create_bullet([{"text": "?µì‹ : ", "bold": True}, "RS485 Modbus RTU, UART Debug"]))
    blocks.append(create_paragraph([]))
    
    # SSC ì½”ë“œ ?˜í”Œ
    blocks.append(create_heading3("?“ ì½”ë“œ ?˜í”Œ - UART ?¸í„°?½íŠ¸ ?¸ë“¤??))
    ssc_code = '''// SSC UART Rx Interrupt Handler
void my_uart_rx_irq_handler(UART_HandleTypeDef *huart)
{
    TERM_PORT   tpid;
    UART_STR*   pUART;
    uint32_t    isrflags = READ_REG(huart->Instance->ISR);

    // Find UART Rx source
    if (huart->Instance == USART1) {
        tpid  = TPID_RS485;
        pUART = &uart[TPID_RS485];
    }
    else if (huart->Instance == USART2) {
        tpid  = TPID_DEBUG;
        pUART = &uart[TPID_DEBUG];
    }
    else return;

    // Check RXNE flag
    if (isrflags & USART_ISR_RXNE) {
        uint8_t data = (uint8_t)(huart->Instance->RDR & 0xFF);
        ring_buffer_put(&pUART->rx_buf, data);
        
        // Modbus frame detection
        if (tpid == TPID_RS485) {
            modbus_rx_handler(data);
        }
    }
}'''
    blocks.append(create_code_block(ssc_code, "c"))
    
    blocks.append(create_divider())
    
    # ==================== 7. ATIK JIG Board ?„ë¡œ?íŠ¸ ====================
    blocks.append(create_heading2("7. ATIK JIG Board (?ì‚° ?ŒìŠ¤???¥ë¹„)"))
    blocks.append(create_callout("?œí’ˆ ?‘ì‚° ?ŒìŠ¤?¸ìš© JIG ë³´ë“œ ?¤ê³„ - Sigma, Titrator ??, "?”Œ"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“‹ ?„ë¡œ?íŠ¸ ê°œìš”"))
    blocks.append(create_bullet(["ê¸°ê°„: 2022.01 ~ 2022.12"]))
    blocks.append(create_bullet(["??• : JIG ë³´ë“œ ?Œë¡œ ?¤ê³„"]))
    blocks.append(create_bullet(["ëª©ì : ?‘ì‚° ?œí’ˆ??ê¸°ëŠ¥ ?ŒìŠ¤???ë™??]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?™ï¸ JIG ë³´ë“œ ê¸°ëŠ¥"))
    blocks.append(create_bullet([{"text": "Power Test: ", "bold": True}, "?„ì› ?ˆì¼ ?ë™ ì¸¡ì •"]))
    blocks.append(create_bullet([{"text": "Communication Test: ", "bold": True}, "UART, SPI, I2C ë£¨í”„ë°??ŒìŠ¤??]))
    blocks.append(create_bullet([{"text": "GPIO Test: ", "bold": True}, "I/O ?€ ?íƒœ ?•ì¸"]))
    blocks.append(create_bullet([{"text": "ê²°ê³¼ ì¶œë ¥: ", "bold": True}, "Pass/Fail LED, ?œë¦¬??ë¡œê·¸"]))
    
    blocks.append(create_divider())
    
    # ============ ê²½ë ¥ ?”ì•½ ?¹ì…˜ ============
    blocks.append(create_heading1("?“ˆ Career Summary"))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?¢ ATIK (2016.03 ~ ?„ì¬) - 9?„ì°¨"))
    blocks.append(create_bullet(["ë¶„ì„ê¸°ê¸° ?„ë¬¸ ?Œì‚¬"]))
    blocks.append(create_bullet(["?˜ë“œ?¨ì–´ ?¤ê³„ + ?Œì›¨??ê°œë°œ ?´ë‹¹"]))
    blocks.append(create_bullet(["ì£¼ìš” ?œí’ˆ: ë¶„ê´‘ê´‘ë„ê³? ?Œí‹°??ì¹´ìš´?? ?ì •ê¸? ì§ˆëŸ‰ë¶„ì„ê¸???]))
    blocks.append(create_paragraph([]))
    
    blocks.append(create_heading3("?“Š ?„ë¡œ?íŠ¸ ì°¸ì—¬ ?µê³„"))
    blocks.append(create_bullet(["HW ?¤ê³„ ?„ë¡œ?íŠ¸: 15+ ì¢?]))
    blocks.append(create_bullet(["FW ê°œë°œ ?„ë¡œ?íŠ¸: 10+ ì¢?]))
    blocks.append(create_bullet(["?‘ì‚° ?ìš© ?œí’ˆ: 8+ ì¢?]))
    
    blocks.append(create_divider())
    
    # ============ GitHub ?¹ì…˜ ============
    blocks.append(create_heading1("?”— GitHub Portfolio"))
    blocks.append(create_paragraph([{"text": "?“‚ Repository: ", "bold": True}, "github.com/gari210404/Portfolio_Professional"]))
    blocks.append(create_paragraph([]))
    blocks.append(create_bullet(["?„ë¡œ?íŠ¸ë³??Œë¡œ??PCB ?Œì¼"]))
    blocks.append(create_bullet(["ë¸”ë¡ ?¤ì´?´ê·¸???´ë?ì§€"]))
    blocks.append(create_bullet(["?Œì›¨??ì½”ë“œ ?˜í”Œ"]))
    blocks.append(create_bullet(["ê¸°ìˆ  ë¬¸ì„œ"]))
    
    # ë¸”ë¡ ì¶”ê?
    print(f"?“ ë¸”ë¡ ì¶”ê? ì¤?.. ({len(blocks)}ê°?")
    
    # 100ê°œì”© ?˜ëˆ ??ì¶”ê?
    for i in range(0, len(blocks), 100):
        chunk = blocks[i:i+100]
        notion.blocks.children.append(block_id=page_id, children=chunk)
        print(f"  ??ë¸”ë¡ {i+1}~{min(i+100, len(blocks))} ì¶”ê? ?„ë£Œ")
    
    print(f"\n???¬íŠ¸?´ë¦¬???ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ URL: https://notion.so/{page_id.replace('-', '')}")
    
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Notion Portfolio Update v3 - With Code Samples
- Psi ?œë¦¬ì¦? ?•ë? ê°€?¤ì œ???œìŠ¤??(ì§„ê³µ ê²Œì´ì§€ + ê³ ì† ?œì–´)
- MS(Aston): DRV Boardë§??´ë‹¹ (?Œë¡œ + ?Œì›¨??
- ì½”ë“œ ?˜í”Œ ì¶”ê?: Lux, JIG Board, MS DRV Board
"""

from notion_client import Client
import time

# Notion API ?¤ì •
# TODO: ?˜ê²½ ë³€?˜ì—??? í° ?½ê¸°
NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"

notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    """?¤ë”© ë¸”ë¡ ?ì„±"""
    heading_type = f"heading_{level}"
    return {
        "object": "block",
        "type": heading_type,
        heading_type: {
            "rich_text": [{"type": "text", "text": {"content": text}}]
        }
    }

def create_paragraph(text, bold=False):
    """ë¬¸ë‹¨ ë¸”ë¡ ?ì„±"""
    return {
        "object": "block",
        "type": "paragraph",
        "paragraph": {
            "rich_text": [{
                "type": "text",
                "text": {"content": text},
                "annotations": {"bold": bold}
            }]
        }
    }

def create_code_block(code, language="c"):
    """ì½”ë“œ ë¸”ë¡ ?ì„±"""
    return {
        "object": "block",
        "type": "code",
        "code": {
            "rich_text": [{"type": "text", "text": {"content": code}}],
            "language": language
        }
    }

def create_bulleted_list(text):
    """ê¸€ë¨¸ë¦¬ ê¸°í˜¸ ë¦¬ìŠ¤???ì„±"""
    return {
        "object": "block",
        "type": "bulleted_list_item",
        "bulleted_list_item": {
            "rich_text": [{"type": "text", "text": {"content": text}}]
        }
    }

def create_divider():
    """êµ¬ë¶„???ì„±"""
    return {"object": "block", "type": "divider", "divider": {}}

def create_callout(text, emoji="?’¡"):
    """ì½œì•„??ë¸”ë¡ ?ì„±"""
    return {
        "object": "block",
        "type": "callout",
        "callout": {
            "rich_text": [{"type": "text", "text": {"content": text}}],
            "icon": {"type": "emoji", "emoji": emoji}
        }
    }

def main():
    print("=== Notion Portfolio v3 ?…ë°?´íŠ¸ ?œì‘ ===\n")
    
    # 1. ???˜ì´ì§€ ?ì„±
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={
            "title": {
                "title": [{"type": "text", "text": {"content": "ì¡°ë½???¬íŠ¸?´ë¦¬??v3 - ì½”ë“œ ?˜í”Œ ?¬í•¨"}}]
            }
        },
        icon={"type": "emoji", "emoji": "??"},
        cover={
            "type": "external",
            "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1200"}
        }
    )
    page_id = new_page["id"]
    print(f"?????˜ì´ì§€ ?ì„±: {page_id}")
    
    # 2. ?„ë¡œ???¹ì…˜
    blocks = [
        create_heading("?‘¤ ì¡°ë½??(Rakhyun Cho)", 1),
        create_callout("Senior Embedded Engineer | 9?„ì°¨ | HW/FW Full-Stack ê°œë°œ??, "??),
        create_paragraph(""),
        create_heading("?“‹ ?µì‹¬ ??Ÿ‰", 2),
        create_bulleted_list("?„ë² ?”ë“œ ?œìŠ¤?? STM32, ARM Cortex-M ê¸°ë°˜ ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("?˜ë“œ?¨ì–´ ?¤ê³„: ?Œë¡œ ?¤ê³„ (Altium Designer), PCB Layout"),
        create_bulleted_list("?¤ì‹œê°??œì–´: RTOS (FreeRTOS), PID ?œì–´, ëª¨í„° ?œì–´"),
        create_bulleted_list("?µì‹  ?„ë¡œ? ì½œ: UART, SPI, I2C, Ethernet (LwIP), CAN"),
        create_bulleted_list("?¼ì„œ ?¸í„°?˜ì´?? ADC (AD7682), DAC, RTD, Thermocouple"),
        create_divider(),
    ]
    
    notion.blocks.children.append(block_id=page_id, children=blocks)
    print("???„ë¡œ???¹ì…˜ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?„ë¡œ?íŠ¸ 1: Psi-1000/3000 (?˜ì •???¤ëª…) =====
    print("\n?“ Psi-1000/3000 ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    
    psi_code = '''// PrsCtrl.c - ?•ë? ê°€???•ë ¥ ?œì–´ ?Œê³ ë¦¬ì¦˜
// ì§„ê³µ ê²Œì´ì§€(P1)ë¥?ëª¨ë‹ˆ?°ë§?˜ë©° ê³ ì† ?¼ë“œë°??œì–´ ?˜í–‰

/* ?•ë ¥ ?…ë ¥ ë³€??*/
INTERNAL float32_t sglPrsCtrlFilteredPressureP1;  // ?„í„°ë§ëœ ?…ë ¥ ?•ë ¥ P1 (torr)
INTERNAL float32_t sglPrsCtrlDifferentialPressureP1;  // ë¯¸ë¶„ ?•ë ¥ (torr/sec)

/* ?œì–´ ì£¼ê¸°: 250msë¡?ê³ ì† ?¼ë“œë°?*/
INTERNAL float32_t sglPrsCtrlControlPeriod = 0.25f;

/* PID ê°€???•ë ¥ ?œì–´ */
void pid_htr_compute(PID_STR *pHTR) {
    float error = pHTR->sp - pHTR->pv;  // ëª©í‘œ?•ë ¥ - ?„ì¬?•ë ¥
    
    // P (ë¹„ë?) ?œì–´
    pHTR->p_term = pHTR->kp * error;
    
    // I (?ë¶„) ?œì–´ - ?„ì  ?¤ì°¨ ë³´ì •
    pHTR->i_term += pHTR->ki * error;
    if (pHTR->i_term > pHTR->i_max) pHTR->i_term = pHTR->i_max;
    
    // D (ë¯¸ë¶„) ?œì–´ - ê¸‰ê²©??ë³€???µì œ
    pHTR->d_term = pHTR->kd * (error - pHTR->prev_error);
    
    // ìµœì¢… ì¶œë ¥ = PWM ?€???¬ì´??
    pHTR->co = pHTR->p_term + pHTR->i_term + pHTR->d_term;
}'''
    
    psi_blocks = [
        create_heading("?”¬ Psi-1000/3000 ?•ë? ê°€?¤ì œ???œìŠ¤??, 2),
        create_callout("ì§„ê³µ ê²Œì´ì§€ë¡?ì§„ê³µ ?íƒœë¥?ëª¨ë‹ˆ?°ë§?˜ë©° ê³ ì† ê°€???•ë ¥ ?œì–´", "?¯"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ë°˜ë„ì²??”ìŠ¤?Œë ˆ??ê³µì •???•ë? ê°€???œì–´ ?œìŠ¤??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4, 168MHz)"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("ì§„ê³µ ê²Œì´ì§€(Vacuum Gauge) ?¸í„°?˜ì´?? Pirani/Capacitance Gauge ?°ë™"),
        create_bulleted_list("ê³ ì† PID ?¼ë“œë°??œì–´: 250ms ì£¼ê¸°ë¡??•ë ¥ ?•ë? ?œì–´"),
        create_bulleted_list("FreeRTOS ê¸°ë°˜ ë©€?°íƒœ?¤í‚¹: ?•ë ¥?œì–´, ?¼ì„œ?½ê¸°, ?µì‹  ë³‘ë ¬ ì²˜ë¦¬"),
        create_bulleted_list("Ethernet (LwIP) TCP/UDP ?µì‹ : PC ?œì–´ ?¸í„°?˜ì´??),
        create_bulleted_list("PID Auto-Tuning: ìµœì  ?œì–´ ?Œë¼ë¯¸í„° ?ë™ ?ìƒ‰"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ê°€???•ë ¥ PID ?œì–´", 3),
        create_code_block(psi_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=psi_blocks)
    print("  ??Psi ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?„ë¡œ?íŠ¸ 2: MS ì§ˆëŸ‰ë¶„ì„ê¸?(DRV Boardë§? =====
    print("\n?“ MS ì§ˆëŸ‰ë¶„ì„ê¸?(DRV Board) ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    
    ms_code = '''// mfc.c - Mass Flow Controller 4-20mA ?¸í„°?˜ì´??
// MS DRV Board?ì„œ MFCë¥??µí•´ ê°€??? ëŸ‰???•ë? ?œì–´

/*
 * McMillan U803 MFC ?¬ì–‘:
 * - ? ëŸ‰ ë²”ìœ„: 10mL/min ~ 50mL/min
 * - ?¸í„°?˜ì´?? 4-20mA ?„ë¥˜ ë£¨í”„
 * 
 * ?˜ì‹: Flow Rate = 3.125 Ã— Current(mA) - 12.5
 */

void mfc_out_svc(void) {
    // 1. MFC ?œì„±???íƒœ ?•ì¸
    if (bd.env.working[OUT_420_EN] & (0x01 << 2))
        mfc.onoff = 1;
    else
        mfc.onoff = 0;
    
    // 2. ëª©í‘œ ? ëŸ‰ ê³„ì‚° (mL/min)
    mfc.out_ml = x2_float(bd.env.working[MASS_FLOW_SPEED]);
    
    // 3. ?¡ì²´ ë³´ì • ê³„ìˆ˜ ?ìš©
    if (bd.env.working[MASS_FLOW_FACTOR] == 0)
        mfc.liq_factor = 1.7;  // ê¸°ë³¸ DSP ?¡ì²´ ê³„ìˆ˜
    else
        mfc.liq_factor = x2_float(bd.env.working[MASS_FLOW_FACTOR]);
    
    mfc.out_ml_factor = mfc.out_ml * mfc.liq_factor;
    if (mfc.out_ml_factor > 50.0) mfc.out_ml_factor = 50.0;  // MFC ìµœë? ?œê³„
    
    // 4. ? ëŸ‰ ???„ë¥˜(mA) ë³€??
    mfc.out_cur = (mfc.out_ml_factor + 12.5) / 3.125;
    
    // 5. ?„ë¥˜ ??DAC ê°?ë³€??(êµì •ê°??ìš©)
    mfc.out_dac = (bd.env.working[CAL_420_OUT2_20m] - bd.env.working[CAL_420_OUT2_4m]) 
                  / 16 * mfc.out_cur;
}'''
    
    ms_blocks = [
        create_heading("?§ª MS ì§ˆëŸ‰ë¶„ì„ê¸?(Aston) - DRV Board", 2),
        create_callout("Drive Board ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ ?´ë‹¹", "?”§"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?˜ì§ˆë¶„ì„??ì§ˆëŸ‰ë¶„ì„ê¸?(Mass Spectrometer)"),
        create_bulleted_list("?´ë‹¹ ë³´ë“œ: DRV Board (Drive Board) - ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("DRV Board ?µì‹¬ ê¸°ëŠ¥", 3),
        create_bulleted_list("MFC(Mass Flow Controller) ?œì–´: 4-20mA ?„ë¥˜ ë£¨í”„ ?¸í„°?˜ì´??),
        create_bulleted_list("?¤ì±„??ADC (AD7682): 16ë¹„íŠ¸ ê³ ì •ë°€ ?¼ì„œ ? í˜¸ ì¸¡ì •"),
        create_bulleted_list("DAC ì¶œë ¥: ?„ë‚ ë¡œê·¸ ?œì–´ ? í˜¸ ?ì„±"),
        create_bulleted_list("PID ?¨ë„ ?œì–´: ?¼ì„œ ?ˆí„°/ì¿¨ëŸ¬ ?œì–´"),
        create_bulleted_list("SDì¹´ë“œ ?°ì´??ë¡œê¹…: FATFS ?Œì¼ ?œìŠ¤??),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - MFC ? ëŸ‰ ?œì–´", 3),
        create_code_block(ms_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=ms_blocks)
    print("  ??MS ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?„ë¡œ?íŠ¸ 3: JIG Board =====
    print("\n?“ JIG Board ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    
    jig_code = '''// ATIK_JIG_main.c - ?¤ì¤‘ ?œí’ˆ ?ë™???ŒìŠ¤???œìŠ¤??
// ?¬ëŸ¬ ?œí’ˆ(L-Titrator, Nu-2000, Sigma, Psi ?????˜ë‚˜??JIGë¡??ŒìŠ¤??

/* ì§€???œí’ˆ ëª©ë¡ */
const char hw_model_str[16][30] = {
    "1.L-Titrator",
    "2.Nu-2000",
    "3.Sigma-3000/4000",
    "4.Psi-3000",
    "5.Psi-1000",
    // ...
    "15.ATIK_JIG"
};

/* ì´ˆê¸°???œì–´ - ë¦´ë ˆ??MUX ?œì–´ */
void init_enable_ctrl(void) {
    uint16_t wr_buf_16;
    
    wr_buf_16 = GPIOF->ODR;
    wr_buf_16 &= ~0xff;
    GPIOF->ODR = wr_buf_16;
    
    // MUX ë¹„í™œ?±í™”
    HAL_GPIO_WritePin(MUX_ENB_GPIO_Port, MUX_ENB_Pin, 0);
    
    // TC&RTD ë¦´ë ˆ??OFF
    HAL_GPIO_WritePin(TC_CTRL_ENB_GPIO_Port, TC_CTRL_ENB_Pin, 0);
    HAL_GPIO_WritePin(TC_CTRL_CLK_GPIO_Port, TC_CTRL_CLK_Pin, 1);
    asm("nop"); asm("nop"); asm("nop");
    HAL_GPIO_WritePin(TC_CTRL_CLK_GPIO_Port, TC_CTRL_CLK_Pin, 0);
    
    // 24V ì¶œë ¥ OFF
    HAL_GPIO_WritePin(OUT1_24V_ENB_GPIO_Port, OUT1_24V_ENB_Pin, 0);
}

/* ?ŒìŠ¤???íƒœ ë¨¸ì‹  */
uint32_t tick_1ms = 0;
uint32_t tick_10ms = 0;
uint16_t target_board_rx_flag = 0;
uint16_t target_board_reset_flag = 0;'''
    
    jig_blocks = [
        create_heading("?”§ JIG Board (?ë™???ŒìŠ¤???œìŠ¤??", 2),
        create_callout("?¤ì¤‘ ?œí’ˆ ?ˆì§ˆ ê²€ì¦ìš© ?ë™???ŒìŠ¤???¥ë¹„ ê°œë°œ", "?¯"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ì‚°?¼ì¸ ?ˆì§ˆê²€ì¦ìš© JIG (Test Fixture)"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?¤ì¤‘ ?œí’ˆ ì§€?? L-Titrator, Nu-2000, Sigma, Psi ?œë¦¬ì¦???),
        create_bulleted_list("ë©€??UART ?¸ë“¤ë§? ?¬ëŸ¬ ?¬íŠ¸ ?™ì‹œ ?µì‹ "),
        create_bulleted_list("ADC/DAC ?ë™ ?ŒìŠ¤?? ?„ë‚ ë¡œê·¸ ?Œë¡œ ê²€ì¦?),
        create_bulleted_list("ë¦´ë ˆ??MUX ?œì–´: ?ŒìŠ¤???¬ì¸??? íƒ"),
        create_bulleted_list("?íƒœ ë¨¸ì‹  ê¸°ë°˜ ?ŒìŠ¤???œí€€?? ?ë™?”ëœ ?ŒìŠ¤???ˆì°¨"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ë©€???œí’ˆ ?ŒìŠ¤???œìŠ¤??, 3),
        create_code_block(jig_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=jig_blocks)
    print("  ??JIG ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?„ë¡œ?íŠ¸ 4: Nu-2000 (Lux) =====
    print("\n?“ Nu-2000 (Lux) ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    
    lux_code = '''// motor.c - A3977 ?¤í…Œ??ëª¨í„° ?œë¼?´ë²„ ?œì–´
// ë§ˆì´?¬ë¡œ ?¤í… ëª¨ë“œë¡??•ë? ?„ì¹˜ ?œì–´ êµ¬í˜„

/* GPIO ?€ ?•ì˜ - A3977 ?œë¼?´ë²„ ?¸í„°?˜ì´??*/
#define STEP_MOTOR_GPIO     GPIOI
#define PIN_MODE2           GPIO_PIN_0   // ë§ˆì´?¬ë¡œ?¤í… ëª¨ë“œ[2]
#define PIN_MODE1           GPIO_PIN_1   // ë§ˆì´?¬ë¡œ?¤í… ëª¨ë“œ[1]
#define PIN_MODE0           GPIO_PIN_2   // ë§ˆì´?¬ë¡œ?¤í… ëª¨ë“œ[0]
#define PIN_STEP            GPIO_PIN_3   // ?¤í… ?„ìŠ¤
#define PIN_DIR             GPIO_PIN_4   // ë°©í–¥ ?œì–´
#define PIN_ENABLE          GPIO_PIN_5   // ì¶œë ¥ ?œì„±??
#define PIN_SLEEP           GPIO_PIN_6   // ?ˆì „ ëª¨ë“œ
#define PIN_RESET           GPIO_PIN_7   // ë¦¬ì…‹

/* ë§ˆì´?¬ë¡œ ?¤í… ëª¨ë“œ */
#define A3977_FULL_STEP     0   // ?„ì²´ ?¤í…
#define A3977_HALF_STEP     1   // 1/2 ?¤í…
#define A3977_QUARTER_STEP  2   // 1/4 ?¤í…
#define A3977_EIGHTH_STEP   3   // 1/8 ?¤í…

/* A3977 ì´ˆê¸°??*/
void init_A3977() {
    A3977_Reset();
    A3977_Output_Disable();
    A3977_Sleep_Disable();
    
    A3977_Set_Step(micro_step);      // ë§ˆì´?¬ë¡œ ?¤í… ëª¨ë“œ ?¤ì •
    A3977_Set_Direction(motor_dir);  // ë°©í–¥ ?¤ì •
    A3977_Set_SR(A3977_SR_ACTIVE);   // Slew Rate ?œì„±??
    
    A3977_Step_Off();
    A3977_Output_Enable();
}

/* ?¤í… ëª¨í„° ë¦¬ì…‹ */
void A3977_Reset() {
    HAL_GPIO_WritePin(STEP_MOTOR_GPIO, PIN_RESET, 0);
    HAL_Delay(10);
    HAL_GPIO_WritePin(STEP_MOTOR_GPIO, PIN_RESET, 1);
}'''
    
    lux_blocks = [
        create_heading("?’§ Nu-2000 (Lux) ?ë™?ì • ?œìŠ¤??, 2),
        create_callout("?˜ì§ˆë¶„ì„???ë™?ì •?¥ì¹˜ - ?œë¦°ì§€ ?Œí”„ ?•ë? ?œì–´", "?—ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?˜ì§ˆ/?í’ˆ ë¶„ì„???ë™?ì •?¥ì¹˜"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("A3977 ?¤í…Œ??ëª¨í„° ?œë¼?´ë²„: ë§ˆì´?¬ë¡œ?¤í… ?œì–´ (1/8 ?¤í…)"),
        create_bulleted_list("?œë¦°ì§€ ?Œí”„ ?•ë? ?œì–´: 0.1Î¼L ?¨ìœ„ ? ì¶œ???œì–´"),
        create_bulleted_list("?¤ì±„??ADC: pH, ?„ë„?? ?¨ë„ ?¼ì„œ ?™ì‹œ ì¸¡ì •"),
        create_bulleted_list("Nextion HMI LCD: ?°ì¹˜?¤í¬ë¦?UI êµ¬í˜„"),
        create_bulleted_list("Ethernet/USB ?µì‹ : PC ?°ë™ ë°??°ì´???„ì†¡"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - A3977 ?¤í…Œ??ëª¨í„° ?œì–´", 3),
        create_code_block(lux_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=lux_blocks)
    print("  ??Nu-2000 ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?„ë¡œ?íŠ¸ 5: L-Titrator =====
    print("\n?“ L-Titrator ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    
    titrator_blocks = [
        create_heading("?§ª L-Titrator ?ë™?ì • ?œìŠ¤??, 2),
        create_callout("?´ë????„ì¥ ?˜ì§ˆë¶„ì„ ?ë™?ì •?¥ì¹˜", "?“±"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?„ì¥???´ë????ë™?ì •?¥ì¹˜"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?œë¦°ì§€ ?Œí”„ ?œì–´: BLDC/?¤í…Œ??ëª¨í„° ?•ë? ?œì–´"),
        create_bulleted_list("ë°°í„°ë¦?ê´€ë¦??œìŠ¤?? ì¶©ì „/ë°©ì „ ?íƒœ ëª¨ë‹ˆ?°ë§"),
        create_bulleted_list("?€?„ë ¥ ?¤ê³„: ?´ë???ê¸°ê¸° ìµœì ??),
        create_bulleted_list("?°ì¹˜?¤í¬ë¦?UI: Nextion HMI LCD"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=titrator_blocks)
    print("  ??L-Titrator ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?„ë¡œ?íŠ¸ 6: Sigma ?œë¦¬ì¦?=====
    print("\n?“ Sigma ?œë¦¬ì¦??„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    
    sigma_blocks = [
        create_heading("?—ï¸ Sigma-1000/3000/4000 ë¶„ì„ ?œìŠ¤??, 2),
        create_callout("COD/BOD/TOC ?˜ì§ˆë¶„ì„ ?„ìš© ?¥ë¹„", "?“Š"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: COD(?”í•™???°ì†Œ?”êµ¬?? ?ë™ë¶„ì„ ?œìŠ¤??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("UV-IR LED ê´‘í•™ ì¸¡ì •: ?¡ê´‘??ê¸°ë°˜ ?ë„ ì¸¡ì •"),
        create_bulleted_list("?¤ì¤‘ ì±„ë„ ?œë¦°ì§€ ?Œí”„: ?œì•½/?˜í”Œ ?ë™ ì£¼ì…"),
        create_bulleted_list("?¨ë„ ?œì–´: ë°˜ì‘ì¡?PID ?¨ë„ ?œì–´"),
        create_bulleted_list("SDì¹´ë“œ ?°ì´??ë¡œê¹…: ì¸¡ì • ê²°ê³¼ ?€??),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=sigma_blocks)
    print("  ??Sigma ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?„ë¡œ?íŠ¸ 7: SSC ?¨ë„?œì–´ =====
    print("\n?“ SSC ?¨ë„?œì–´ ?œìŠ¤???„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    
    ssc_blocks = [
        create_heading("?Œ¡ï¸?SSC ??˜¨ì¡??¨ë„?œì–´ ?œìŠ¤??, 2),
        create_callout("ê³ ì •ë°€ ??˜¨ì¡?PID ?¨ë„?œì–´ ?œìŠ¤??, "?¨ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?¤í—˜??ê³ ì •ë°€ ??˜¨ì¡?(Â±0.01Â°C)"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("RTD/Thermocouple ?¨ë„ ?¼ì„œ: ê³ ì •ë°€ ?¨ë„ ì¸¡ì •"),
        create_bulleted_list("PID ?¨ë„ ?œì–´: ?ˆí„°/ì¿¨ëŸ¬ PWM ?œì–´"),
        create_bulleted_list("PID Auto-Tuning: Ziegler-Nichols ë°©ì‹ ?ë™ ?œë‹"),
        create_bulleted_list("Ethernet ?µì‹ : ?ê²© ëª¨ë‹ˆ?°ë§ ë°??œì–´"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=ssc_blocks)
    print("  ??SSC ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ê¸°ìˆ  ?¤íƒ =====
    print("\n?“ ê¸°ìˆ  ?¤íƒ ?¹ì…˜ ì¶”ê? ì¤?..")
    
    tech_blocks = [
        create_heading("?› ï¸?ê¸°ìˆ  ?¤íƒ", 2),
        create_paragraph(""),
        create_heading("?„ë¡œê·¸ë˜ë°??¸ì–´", 3),
        create_bulleted_list("C/C++ (Embedded), Python, Verilog"),
        create_paragraph(""),
        create_heading("MCU/?„ë¡œ?¸ì„œ", 3),
        create_bulleted_list("STM32 ?œë¦¬ì¦?(F1, F4, H7), Renesas RA6M4"),
        create_bulleted_list("Xilinx Zynq (FPGA + ARM)"),
        create_paragraph(""),
        create_heading("ê°œë°œ ?„êµ¬", 3),
        create_bulleted_list("IDE: STM32CubeIDE, VS Code, Keil"),
        create_bulleted_list("EDA: Altium Designer (?Œë¡œ?¤ê³„/PCB)"),
        create_bulleted_list("FPGA: Vivado, Vitis"),
        create_bulleted_list("ë²„ì „ê´€ë¦? Git, GitHub"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=tech_blocks)
    print("  ??ê¸°ìˆ  ?¤íƒ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)
    
    # ===== ?°ë½ì²?=====
    contact_blocks = [
        create_heading("?“ ?°ë½ì²?, 2),
        create_bulleted_list("?“§ Email: gari210@naver.com"),
        create_bulleted_list("?™ GitHub: github.com/gari210404"),
        create_paragraph(""),
        create_callout("Thank you for reviewing my portfolio!", "?™"),
    ]
    notion.blocks.children.append(block_id=page_id, children=contact_blocks)
    print("  ???°ë½ì²?ì¶”ê? ?„ë£Œ")
    
    print(f"\n{'='*50}")
    print("?‰ ?¸ì…˜ ?¬íŠ¸?´ë¦¬??v3 ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print(f"{'='*50}")
    
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Notion Portfolio v4 - ?•í™•???œí’ˆ ?¤ëª… + ë¸”ë¡ ?¤ì´?´ê·¸???ë¦¬ + ì½”ë“œ ?˜í”Œ
"""

from notion_client import Client
import time

# Notion API ?¤ì •
NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"

notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {
        "object": "block",
        "type": heading_type,
        heading_type: {
            "rich_text": [{"type": "text", "text": {"content": text}}]
        }
    }

def create_paragraph(text, bold=False):
    return {
        "object": "block",
        "type": "paragraph",
        "paragraph": {
            "rich_text": [{
                "type": "text",
                "text": {"content": text},
                "annotations": {"bold": bold}
            }] if text else []
        }
    }

def create_code_block(code, language="c"):
    return {
        "object": "block",
        "type": "code",
        "code": {
            "rich_text": [{"type": "text", "text": {"content": code}}],
            "language": language
        }
    }

def create_bulleted_list(text):
    return {
        "object": "block",
        "type": "bulleted_list_item",
        "bulleted_list_item": {
            "rich_text": [{"type": "text", "text": {"content": text}}]
        }
    }

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def create_callout(text, emoji="?’¡"):
    return {
        "object": "block",
        "type": "callout",
        "callout": {
            "rich_text": [{"type": "text", "text": {"content": text}}],
            "icon": {"type": "emoji", "emoji": emoji}
        }
    }

def create_image_placeholder(caption="[ë¸”ë¡ ?¤ì´?´ê·¸???½ì… ?„ì¹˜]"):
    """ë¸”ë¡ ?¤ì´?´ê·¸???½ì… ?„ì¹˜ ?œì‹œ"""
    return {
        "object": "block",
        "type": "callout",
        "callout": {
            "rich_text": [{"type": "text", "text": {"content": caption}}],
            "icon": {"type": "emoji", "emoji": "?“Š"},
            "color": "gray_background"
        }
    }

def main():
    print("=== Notion Portfolio v4 ?…ë°?´íŠ¸ ?œì‘ ===\n")
    
    # ???˜ì´ì§€ ?ì„±
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={
            "title": {
                "title": [{"type": "text", "text": {"content": "ì¡°ë½???¬íŠ¸?´ë¦¬??v4 - Final"}}]
            }
        },
        icon={"type": "emoji", "emoji": "??"},
        cover={
            "type": "external",
            "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1200"}
        }
    )
    page_id = new_page["id"]
    print(f"?????˜ì´ì§€ ?ì„±: {page_id}")
    
    # ===== ?„ë¡œ???¹ì…˜ =====
    blocks = [
        create_heading("?‘¤ ì¡°ë½??(Rakhyun Cho)", 1),
        create_callout("Senior Embedded Engineer | 9?„ì°¨ | HW/FW Full-Stack ê°œë°œ??, "??),
        create_paragraph(""),
        create_heading("?“‹ ?µì‹¬ ??Ÿ‰", 2),
        create_bulleted_list("?„ë² ?”ë“œ ?œìŠ¤?? STM32 (F1, F4, H7), ARM Cortex-M ê¸°ë°˜ ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("?˜ë“œ?¨ì–´ ?¤ê³„: Altium Designer ê¸°ë°˜ ?Œë¡œ ?¤ê³„ ë°?PCB Layout"),
        create_bulleted_list("?¤ì‹œê°??œì–´: FreeRTOS, PID ?œì–´, ëª¨í„° ?œì–´ (Stepper, BLDC)"),
        create_bulleted_list("ê´‘í•™ ?œìŠ¤?? UV/IR LED, Photo Diode Array, ?ˆì´?€ ê´‘í•™ê³??¸í„°?˜ì´??),
        create_bulleted_list("?µì‹  ?„ë¡œ? ì½œ: UART, SPI, I2C, Ethernet (LwIP), USB"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=blocks)
    print("???„ë¡œ???¹ì…˜ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 1. Nu-2000 (Lux) =====
    print("\n?“ Nu-2000 (Lux) ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    lux_code = '''// motor.c - A3977 ?¤í…Œ??ëª¨í„° ?œë¼?´ë²„ ?œì–´
// ?œë¦°ì§€ ?Œí”„??ë§ˆì´?¬ë¡œ ?¤í… ?•ë? ?œì–´

#define STEP_MOTOR_GPIO     GPIOI
#define PIN_STEP            GPIO_PIN_3   // ?¤í… ?„ìŠ¤
#define PIN_DIR             GPIO_PIN_4   // ë°©í–¥ ?œì–´
#define PIN_ENABLE          GPIO_PIN_5   // ì¶œë ¥ ?œì„±??

/* ë§ˆì´?¬ë¡œ ?¤í… ëª¨ë“œ */
#define A3977_FULL_STEP     0   // ?„ì²´ ?¤í…
#define A3977_HALF_STEP     1   // 1/2 ?¤í…
#define A3977_QUARTER_STEP  2   // 1/4 ?¤í…
#define A3977_EIGHTH_STEP   3   // 1/8 ?¤í…

void init_A3977() {
    A3977_Reset();
    A3977_Output_Disable();
    A3977_Sleep_Disable();
    
    A3977_Set_Step(micro_step);      // ë§ˆì´?¬ë¡œ ?¤í… ëª¨ë“œ ?¤ì •
    A3977_Set_Direction(motor_dir);  // ë°©í–¥ ?¤ì •
    A3977_Set_SR(A3977_SR_ACTIVE);   // Slew Rate ?œì„±??
    
    A3977_Step_Off();
    A3977_Output_Enable();
}'''

    lux_blocks = [
        create_heading("?’§ Nu-2000 (Lux) - UV+IR ?©ì•¡ ?ë„ ë¶„ì„ê¸?, 2),
        create_callout("UV+IR ê´‘í•™???´ìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„", "?”¬"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: UV+IR(ê´‘í•™)ë¥??´ìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4, 168MHz)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Nu-2000 ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("UV/IR LED ê´‘í•™ ì¸¡ì •: ?¡ê´‘??ê¸°ë°˜ ?ë„ ë¶„ì„"),
        create_bulleted_list("A3977 ?¤í…Œ??ëª¨í„°: ?œë¦°ì§€ ?Œí”„ ë§ˆì´?¬ë¡œ?¤í… ?œì–´ (1/8 ?¤í…)"),
        create_bulleted_list("?¤ì±„??ADC (AD7682): 16ë¹„íŠ¸ ê³ ì •ë°€ ê´‘ì„¼??? í˜¸ ì¸¡ì •"),
        create_bulleted_list("Nextion HMI LCD: ?°ì¹˜?¤í¬ë¦?UI"),
        create_bulleted_list("Ethernet/USB ?µì‹ : PC ?°ë™ ë°??°ì´???„ì†¡"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ?¤í…Œ??ëª¨í„° ?œì–´", 3),
        create_code_block(lux_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=lux_blocks)
    print("  ??Nu-2000 ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 2. L-Titrator =====
    print("\n?“ L-Titrator ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    titrator_blocks = [
        create_heading("?§ª L-Titrator - ?ì •???©ì•¡ ?ë„ ë¶„ì„ê¸?, 2),
        create_callout("?ì •(Titration) ë°©ì‹???œìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„", "?—ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ì • ë°©ì‹???œìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[L-Titrator ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?œë¦°ì§€ ?Œí”„ ?œì–´: BLDC/?¤í…Œ??ëª¨í„° ?•ë? ? ì¶œ ?œì–´"),
        create_bulleted_list("pH/?„ë„???¼ì„œ: ?ì • ì¢…ë§??ê²€ì¶?),
        create_bulleted_list("ë°°í„°ë¦?ê´€ë¦??œìŠ¤?? ?´ë???ê¸°ê¸° ì¶©ë°©??ê´€ë¦?),
        create_bulleted_list("?€?„ë ¥ ?¤ê³„: ?„ì¥ ?´ë???ìµœì ??),
        create_bulleted_list("?°ì¹˜?¤í¬ë¦?UI: Nextion HMI LCD"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=titrator_blocks)
    print("  ??L-Titrator ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 3. Psi-1000/3000 =====
    print("\n?“ Psi ?œë¦¬ì¦??„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    psi_code = '''// PrsCtrl.c - ?•ë? ê°€???•ë ¥ ?œì–´ ?Œê³ ë¦¬ì¦˜
// ì§„ê³µ ê²Œì´ì§€ ëª¨ë‹ˆ?°ë§ + ê³ ì† PID ?¼ë“œë°??œì–´

/* ?•ë ¥ ?œì–´ ë³€??*/
INTERNAL float32_t sglPrsCtrlFilteredPressureP1;  // ?…ë ¥ ?•ë ¥ (torr)
INTERNAL float32_t sglPrsCtrlControlPeriod = 0.25f;  // 250ms ?œì–´ ì£¼ê¸°

/* PID ?•ë ¥ ?œì–´ */
void pid_pressure_compute(PID_STR *pPID) {
    float error = pPID->sp - pPID->pv;  // ëª©í‘œ?•ë ¥ - ?„ì¬?•ë ¥
    
    pPID->p_term = pPID->kp * error;           // ë¹„ë?
    pPID->i_term += pPID->ki * error;          // ?ë¶„
    pPID->d_term = pPID->kd * (error - pPID->prev_error);  // ë¯¸ë¶„
    
    // ìµœì¢… ì¶œë ¥ (ë°¸ë¸Œ PWM)
    pPID->co = pPID->p_term + pPID->i_term + pPID->d_term;
}'''

    psi_blocks = [
        create_heading("?”¬ Psi-1000/3000 - ?•ë? ê°€???œì–´ ?œìŠ¤??, 2),
        create_callout("ì§„ê³µ ê²Œì´ì§€ ëª¨ë‹ˆ?°ë§ ê¸°ë°˜ ?•ë? ê°€???œì–´ ?œìŠ¤??, "?¯"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?•ë? ê°€???œì–´ ?œìŠ¤??(ë°˜ë„ì²??”ìŠ¤?Œë ˆ??ê³µì •??"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4, 168MHz)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Psi ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("ì§„ê³µ ê²Œì´ì§€ ?¸í„°?˜ì´?? Pirani/Capacitance Gauge ?°ë™"),
        create_bulleted_list("ê³ ì† PID ?¼ë“œë°??œì–´: 250ms ì£¼ê¸° ?•ë ¥ ?œì–´"),
        create_bulleted_list("FreeRTOS ë©€?°íƒœ?¤í‚¹: ?•ë ¥?œì–´, ?¼ì„œ, ?µì‹  ë³‘ë ¬ ì²˜ë¦¬"),
        create_bulleted_list("Ethernet (LwIP): TCP/UDP PC ?œì–´ ?¸í„°?˜ì´??),
        create_bulleted_list("PID Auto-Tuning: ìµœì  ?Œë¼ë¯¸í„° ?ë™ ?ìƒ‰"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ê°€???•ë ¥ PID ?œì–´", 3),
        create_code_block(psi_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=psi_blocks)
    print("  ??Psi ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 4. LPC =====
    print("\n?“ LPC ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    lpc_blocks = [
        create_heading("?”´ LPC - ?ˆì´?€ ?Œí‹°??ë¶„ì„ê¸?, 2),
        create_callout("?ˆì´?€(ê´‘í•™)ë¥??´ìš©???¬ëŸ¬ë¦??Œí‹°??ë¶„ì„ ?¥ë¹„", "?’¡"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ˆì´?€ ê´‘í•™ ê¸°ë°˜ ?¬ëŸ¬ë¦??Œí‹°??ì¹´ìš´??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[LPC ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?ˆì´?€ ê´‘í•™ê³? ?°ë?ê´?ê²€ì¶?ê¸°ë°˜ ?Œí‹°??ê³„ìˆ˜"),
        create_bulleted_list("ê³ ì† ADC ?˜í”Œë§? ?¤ì‹œê°??Œí‹°??? í˜¸ ê²€ì¶?),
        create_bulleted_list("? í˜¸ ì²˜ë¦¬: ?”ì????„í„°ë§?ë°??¼í¬ ê²€ì¶?),
        create_bulleted_list("?°ì´??ë¡œê¹…: SDì¹´ë“œ ì¸¡ì • ê²°ê³¼ ?€??),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=lpc_blocks)
    print("  ??LPC ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 5. SSC =====
    print("\n?“ SSC ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    ssc_blocks = [
        create_heading("?“Š SSC - Photo Diode Array ?¬ëŸ¬ë¦?ë¶„ì„ê¸?, 2),
        create_callout("Photo Diode Arrayë¥??¬ìš©??ê´‘í•™???¬ëŸ¬ë¦?ë¶„ì„ ?¥ë¹„", "?“ˆ"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: PDA(Photo Diode Array) ê¸°ë°˜ ê´‘í•™???¬ëŸ¬ë¦?ë¶„ì„ê¸?),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[SSC ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("Photo Diode Array: ?¤ì±„???™ì‹œ ê´‘í•™ ì¸¡ì •"),
        create_bulleted_list("ë¶„ê´‘ ë¶„ì„: ?Œì¥ë³??¡ê´‘??ì¸¡ì •"),
        create_bulleted_list("ê³ ì† ?°ì´???˜ì§‘: ë©€?°ì±„??ADC ?™ê¸°??),
        create_bulleted_list("?¤ì‹œê°?ë¶„ì„: ?¬ëŸ¬ë¦??ë„/?…ë„ ë¶„ì„"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=ssc_blocks)
    print("  ??SSC ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 6. MS (Aston) =====
    print("\n?“ MS ì§ˆëŸ‰ë¶„ì„ê¸??„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    ms_code = '''// mfc.c - Mass Flow Controller 4-20mA ?¸í„°?˜ì´??
// DRV Board?ì„œ ?Œí”„/?¼ì„œ ?¥ë¹„ ?œì–´

/* McMillan U803 MFC: 10~50 mL/min, 4-20mA */
void mfc_out_svc(void) {
    // 1. MFC ?œì„±???íƒœ ?•ì¸
    if (bd.env.working[OUT_420_EN] & (0x01 << 2))
        mfc.onoff = 1;
    
    // 2. ëª©í‘œ ? ëŸ‰ (mL/min)
    mfc.out_ml = x2_float(bd.env.working[MASS_FLOW_SPEED]);
    
    // 3. ?¡ì²´ ë³´ì • ê³„ìˆ˜ ?ìš©
    mfc.liq_factor = x2_float(bd.env.working[MASS_FLOW_FACTOR]);
    mfc.out_ml_factor = mfc.out_ml * mfc.liq_factor;
    if (mfc.out_ml_factor > 50.0) mfc.out_ml_factor = 50.0;
    
    // 4. ? ëŸ‰ ???„ë¥˜(mA) ??DAC ë³€??
    mfc.out_cur = (mfc.out_ml_factor + 12.5) / 3.125;
    mfc.out_dac = (bd.env.working[CAL_420_OUT2_20m] - bd.env.working[CAL_420_OUT2_4m]) 
                  / 16 * mfc.out_cur;
}'''

    ms_blocks = [
        create_heading("?§ª MS (Aston) - ì§ˆëŸ‰ë¶„ì„ê¸?DRV Board", 2),
        create_callout("ì§ˆëŸ‰ë¶„ì„ê¸?Drive Board - ?Œí”„, ?¼ì„œ ???¥ë¹„ ?œì–´", "?™ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?˜ì§ˆë¶„ì„??ì§ˆëŸ‰ë¶„ì„ê¸?(Mass Spectrometer)"),
        create_bulleted_list("?´ë‹¹: DRV Board (Drive Board) ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[MS DRV Board ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("DRV Board ?µì‹¬ ê¸°ëŠ¥", 3),
        create_bulleted_list("MFC(Mass Flow Controller) ?œì–´: 4-20mA ?„ë¥˜ ë£¨í”„"),
        create_bulleted_list("?Œí”„ ?œì–´: ?œë£Œ/?œì•½ ì£¼ì… ?Œí”„ êµ¬ë™"),
        create_bulleted_list("?¼ì„œ ?¸í„°?˜ì´?? ?¤ì±„??ADC ?¼ì„œ ? í˜¸ ì¸¡ì •"),
        create_bulleted_list("PID ?¨ë„ ?œì–´: ?¼ì„œ ?ˆí„°/ì¿¨ëŸ¬ ?œì–´"),
        create_bulleted_list("SDì¹´ë“œ ë¡œê¹…: FATFS ?°ì´???€??),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - MFC ? ëŸ‰ ?œì–´", 3),
        create_code_block(ms_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=ms_blocks)
    print("  ??MS ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 7. ATIK JIG Board =====
    print("\n?“ ATIK JIG Board ?„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    jig_code = '''// ATIK_JIG_main.c - ë³´ë“œ ?…ê³  ?ŒìŠ¤???œìŠ¤??
// ?¤ì¤‘ ?œí’ˆ ë³´ë“œ ?ë™???ŒìŠ¤??ì§€??

const char hw_model_str[16][30] = {
    "1.L-Titrator",
    "2.Nu-2000",
    "3.Sigma-3000/4000",
    "4.Psi-3000",
    "5.Psi-1000",
    "15.ATIK_JIG"
};

void init_enable_ctrl(void) {
    // MUX ë¹„í™œ?±í™”
    HAL_GPIO_WritePin(MUX_ENB_GPIO_Port, MUX_ENB_Pin, 0);
    
    // TC&RTD ë¦´ë ˆ??OFF
    HAL_GPIO_WritePin(TC_CTRL_ENB_GPIO_Port, TC_CTRL_ENB_Pin, 0);
    
    // 24V ì¶œë ¥ OFF
    HAL_GPIO_WritePin(OUT1_24V_ENB_GPIO_Port, OUT1_24V_ENB_Pin, 0);
}'''

    jig_blocks = [
        create_heading("?”§ ATIK JIG Board - ë³´ë“œ ?…ê³  ?ŒìŠ¤???¥ë¹„", 2),
        create_callout("ë³´ë“œ ?…ê³  ?ŒìŠ¤?¸ë? ?„í•œ ?ë™???ŒìŠ¤???¥ë¹„", "?› ï¸?),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ì‚°?¼ì¸ ë³´ë“œ ?…ê³  ?ŒìŠ¤?¸ìš© JIG"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[JIG Board ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?¤ì¤‘ ?œí’ˆ ì§€?? L-Titrator, Nu-2000, Sigma, Psi ??),
        create_bulleted_list("?ë™???ŒìŠ¤?? ADC/DAC, ?µì‹ , GPIO ?ë™ ê²€ì¦?),
        create_bulleted_list("ë©€??UART ?¸ë“¤ë§? ?€ê²?ë³´ë“œ ?µì‹  ?ŒìŠ¤??),
        create_bulleted_list("ë¦´ë ˆ??MUX ?œì–´: ?ŒìŠ¤???¬ì¸???ë™ ? íƒ"),
        create_bulleted_list("?ŒìŠ¤??ê²°ê³¼ ë¦¬í¬?? PASS/FAIL ?ì • ë°?ë¡œê¹…"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ë©€???œí’ˆ ?ŒìŠ¤??, 3),
        create_code_block(jig_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=jig_blocks)
    print("  ??JIG Board ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 8. Sigma ?œë¦¬ì¦?=====
    print("\n?“ Sigma ?œë¦¬ì¦??„ë¡œ?íŠ¸ ì¶”ê? ì¤?..")
    sigma_blocks = [
        create_heading("?—ï¸ Sigma-1000/3000/4000 - COD ë¶„ì„ ?œìŠ¤??, 2),
        create_callout("COD/BOD/TOC ?˜ì§ˆë¶„ì„ ?„ìš© ?¥ë¹„", "?“Š"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: COD(?”í•™???°ì†Œ?”êµ¬?? ?ë™ë¶„ì„ ?œìŠ¤??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Sigma ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("UV-IR LED ê´‘í•™ ì¸¡ì •: ?¡ê´‘??ê¸°ë°˜ ?ë„ ì¸¡ì •"),
        create_bulleted_list("?¤ì¤‘ ì±„ë„ ?œë¦°ì§€ ?Œí”„: ?œì•½/?˜í”Œ ?ë™ ì£¼ì…"),
        create_bulleted_list("ë°˜ì‘ì¡??¨ë„ ?œì–´: PID ê¸°ë°˜ ?•ë? ?¨ë„ ?œì–´"),
        create_bulleted_list("SDì¹´ë“œ ?°ì´??ë¡œê¹…: ì¸¡ì • ê²°ê³¼ ?€??),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=sigma_blocks)
    print("  ??Sigma ?„ë¡œ?íŠ¸ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== ê¸°ìˆ  ?¤íƒ =====
    print("\n?“ ê¸°ìˆ  ?¤íƒ ?¹ì…˜ ì¶”ê? ì¤?..")
    tech_blocks = [
        create_heading("?› ï¸?ê¸°ìˆ  ?¤íƒ", 2),
        create_paragraph(""),
        create_heading("?„ë¡œê·¸ë˜ë°??¸ì–´", 3),
        create_bulleted_list("C/C++ (Embedded), Python, Verilog"),
        create_heading("MCU/?„ë¡œ?¸ì„œ", 3),
        create_bulleted_list("STM32 ?œë¦¬ì¦?(F1, F4, H7), Renesas RA6M4"),
        create_bulleted_list("Xilinx Zynq (FPGA + ARM)"),
        create_heading("ê°œë°œ ?„êµ¬", 3),
        create_bulleted_list("IDE: STM32CubeIDE, VS Code, Keil"),
        create_bulleted_list("EDA: Altium Designer (?Œë¡œ?¤ê³„/PCB)"),
        create_bulleted_list("FPGA: Vivado, Vitis"),
        create_bulleted_list("ë²„ì „ê´€ë¦? Git, GitHub"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=tech_blocks)
    print("  ??ê¸°ìˆ  ?¤íƒ ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== ?°ë½ì²?=====
    contact_blocks = [
        create_heading("?“ ?°ë½ì²?, 2),
        create_bulleted_list("?“§ Email: gari210@naver.com"),
        create_bulleted_list("?™ GitHub: github.com/gari210404"),
        create_paragraph(""),
        create_callout("Thank you for reviewing my portfolio!", "?™"),
    ]
    notion.blocks.children.append(block_id=page_id, children=contact_blocks)
    print("  ???°ë½ì²?ì¶”ê? ?„ë£Œ")

    print(f"\n{'='*50}")
    print("?‰ ?¸ì…˜ ?¬íŠ¸?´ë¦¬??v4 ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print(f"{'='*50}")
    
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Notion Portfolio v5 - ?•í™•???œí’ˆ ?¤ëª… + ?ì ˆ??ì½”ë“œ ?˜í”Œ + ë¸”ë¡ ?¤ì´?´ê·¸???ë¦¬
"""

from notion_client import Client
import time

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {"object": "block", "type": heading_type, heading_type: {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_paragraph(text, bold=False):
    return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": text}, "annotations": {"bold": bold}}] if text else []}}

def create_code_block(code, language="c"):
    return {"object": "block", "type": "code", "code": {"rich_text": [{"type": "text", "text": {"content": code}}], "language": language}}

def create_bulleted_list(text):
    return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def create_callout(text, emoji="?’¡"):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": text}}], "icon": {"type": "emoji", "emoji": emoji}}}

def create_image_placeholder(caption):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": caption}}], "icon": {"type": "emoji", "emoji": "?“Š"}, "color": "gray_background"}}

def main():
    print("=== Notion Portfolio v5 ?…ë°?´íŠ¸ ?œì‘ ===\n")
    
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={"title": {"title": [{"type": "text", "text": {"content": "ì¡°ë½???¬íŠ¸?´ë¦¬??v5 - Final"}}]}},
        icon={"type": "emoji", "emoji": "??"},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1200"}}
    )
    page_id = new_page["id"]
    print(f"?????˜ì´ì§€ ?ì„±: {page_id}")
    
    # ===== ?„ë¡œ??=====
    blocks = [
        create_heading("?‘¤ ì¡°ë½??(Rakhyun Cho)", 1),
        create_callout("Senior Embedded Engineer | 9?„ì°¨ | HW/FW Full-Stack ê°œë°œ??, "??),
        create_paragraph(""),
        create_heading("?“‹ ?µì‹¬ ??Ÿ‰", 2),
        create_bulleted_list("?„ë² ?”ë“œ ?œìŠ¤?? STM32 (F1, F4, H7), ARM Cortex-M ê¸°ë°˜ ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("?˜ë“œ?¨ì–´ ?¤ê³„: Altium Designer ê¸°ë°˜ ?Œë¡œ ?¤ê³„ ë°?PCB Layout"),
        create_bulleted_list("?¤ì‹œê°??œì–´: FreeRTOS, PID ?œì–´, ëª¨í„° ?œì–´ (Stepper, BLDC)"),
        create_bulleted_list("ê´‘í•™ ?œìŠ¤?? UV/IR LED, Photo Diode Array, ?ˆì´?€ ê´‘í•™ê³?),
        create_bulleted_list("?µì‹  ?„ë¡œ? ì½œ: UART, SPI, I2C, Ethernet (LwIP), TCP/UDP"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=blocks)
    print("???„ë¡œ???¹ì…˜ ì¶”ê?")
    time.sleep(0.5)

    # ===== 1. Nu-2000 (Lux) - I2C ?¼ì„œ + ?‰ê· ê°?=====
    print("\n?“ Nu-2000 ì¶”ê? ì¤?..")
    nu2000_code = '''// ADS1115.c - I2C ADC ?¼ì„œ ?°ì´???˜ì§‘ ë°??´ë™?‰ê·  ê³„ì‚°
// UV/IR Photo Diode ? í˜¸ë¥?16ë¹„íŠ¸ ADCë¡?ì¸¡ì •

/* I2Cë¥??µí•œ ADC ?°ì´???½ê¸° */
int ads1115_read(uint16_t addr, uint8_t *pdata) {
    uint16_t i2c_addr = (addr | 0x01);  // Read operation
    int result = HAL_I2C_Master_Receive(&hi2c3, i2c_addr, pdata, 2, 10);
    ads1115_i2c_err_check(addr, result);
    return result;
}

/* ?´ë™?‰ê·  ê³„ì‚° - ?¸ì´ì¦??œê±° ë°??ˆì •??ì¸¡ì • */
void adc_pd_svc_beta(uint8_t ch, int sig_dark, uint16_t val) {
    uint32_t sum;
    int i, cnt;
    
    // ë²„í¼ ?¬ì¸??ì¦ê? (?œí™˜ ë²„í¼)
    cnt = pPD_STR->ma_cnt;
    if (++cnt >= ((bd.env.working[FILTER] & 0x00FFFFFF) * 5))
        cnt = 0;
    pPD_STR->ma_cnt = cnt;
    
    // ADC ê°??€??(?‘ìˆ˜ë§??ˆìš©)
    pPD_STR->raw[cnt] = val;
    if (val < 0x8000) {
        pPD_STR->raw_pos[cnt] = val;
        pPD_STR->vtg_raw = (float)val * 0.000125;  // 125uV/step
    } else {
        pPD_STR->raw_pos[cnt] = 0;
    }
    
    // ?´ë™?‰ê·  ê³„ì‚° (Filter * 5???˜í”Œ)
    sum = 0;
    for (i = 0; i < ((bd.env.working[FILTER] & 0x00FFFFFF) * 5); i++) 
        sum += pPD_STR->raw_pos[i];
    
    pPD_STR->raw_mavg = (uint16_t)((float)sum / (float)((bd.env.working[FILTER] & 0x00FFFFFF) * 5));
    pPD_STR->vtg_mavg = (float)pPD_STR->raw_mavg * 0.000125;  // ?„ì•• ë³€??
}'''

    nu2000_blocks = [
        create_heading("?’§ Nu-2000 (Lux) - UV+IR ?©ì•¡ ?ë„ ë¶„ì„ê¸?, 2),
        create_callout("UV+IR ê´‘í•™???´ìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„", "?”¬"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: UV+IR(ê´‘í•™)ë¥??´ìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4, 168MHz)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Nu-2000 ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("UV/IR LED ê´‘í•™ ì¸¡ì •: ?¡ê´‘??ê¸°ë°˜ ?ë„ ë¶„ì„"),
        create_bulleted_list("ADS1115 16ë¹„íŠ¸ ADC: I2C ?¸í„°?˜ì´?? Delta-Sigma ë°©ì‹"),
        create_bulleted_list("?´ë™?‰ê·  ?„í„°: ?¸ì´ì¦??œê±° ë°?ì¸¡ì • ?ˆì •??),
        create_bulleted_list("?œë¦°ì§€ ?Œí”„: A3977 ?¤í…Œ??ëª¨í„° ë§ˆì´?¬ë¡œ?¤í… ?œì–´"),
        create_bulleted_list("Nextion HMI LCD: ?°ì¹˜?¤í¬ë¦?UI"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - I2C ?¼ì„œ ?°ì´???´ë™?‰ê· ", 3),
        create_code_block(nu2000_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=nu2000_blocks)
    print("  ??Nu-2000 ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 2. L-Titrator - Syringe ?Œí”„ ?œì–´ =====
    print("\n?“ L-Titrator ì¶”ê? ì¤?..")
    titrator_code = '''// syringe_pump_ctrl.c - Hamilton ?œë¦°ì§€ ?Œí”„ ?íƒœë¨¸ì‹  ?œì–´
// RS485 ?µì‹ ?¼ë¡œ ?œë¦°ì§€ ?Œí”„ ëª…ë ¹/?‘ë‹µ ì²˜ë¦¬

/* ?œë¦°ì§€ ?Œí”„ ?íƒœ ë¨¸ì‹  */
void syr_pump_st10_svc(void) {
    int data_pos;
    
    switch (syr_pump[syr_pump_ch].st) {
        case 0x10:  // ?œë¦°ì§€ ?íƒœ ì¡°íšŒ
            syr_pump_query_cmd(QUERY_SRG_STATUS);
            syr_pump[syr_pump_ch].st++;
            syr_pump[syr_pump_ch].resp_cnt = 0;
            break;
            
        case 0x11:  // ?‘ë‹µ ?€ê¸?ë°?ì²˜ë¦¬
            if (++syr_pump[syr_pump_ch].resp_cnt >= SRG_PUMP_RESP_TIMEOUT) {
                // ?€?„ì•„??- ì´ˆê¸° ?íƒœë¡?ë³µê?
                syr_pump[syr_pump_ch].st = SRG_PUMP_ST00_DETECT;
                trace_printf(TPID_DEBUG, "syr_pump[%d] : no response\\n", syr_pump_ch);
            } else {
                if (syr_pump_rx_flag == 1) {
                    syr_pump_rx_flag = 0;
                    data_pos = is_Q_ack_packet(uart[TPID_RS485].command, uart[TPID_RS485].cmd_index);
                    
                    if (data_pos > 0) {
                        syr_pump[syr_pump_ch].syr_status = syr_pump[syr_pump_ch].temp;
                        if (syr_pump[syr_pump_ch].syr_status == 0) {
                            trace_printf(TPID_DEBUG, "syr_pump[%d] : initialized\\n", syr_pump_ch);
                            syr_pump[syr_pump_ch].st++;
                        }
                    }
                }
            }
            break;
    }
}'''

    titrator_blocks = [
        create_heading("?§ª L-Titrator - ?ì •???©ì•¡ ?ë„ ë¶„ì„ê¸?, 2),
        create_callout("?ì •(Titration) ë°©ì‹???œìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„", "?—ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ì • ë°©ì‹???œìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[L-Titrator ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("Hamilton ?œë¦°ì§€ ?Œí”„: RS485 ?„ë¡œ? ì½œ ?œì–´"),
        create_bulleted_list("?íƒœë¨¸ì‹  ê¸°ë°˜ ?œì–´: ?ˆì •?ì¸ ?Œí”„ ?™ì‘ ê´€ë¦?),
        create_bulleted_list("pH/?„ë„???¼ì„œ: ?ì • ì¢…ë§???ë™ ê²€ì¶?),
        create_bulleted_list("ë°°í„°ë¦?ê´€ë¦? ?´ë???ê¸°ê¸° ?€?„ë ¥ ?¤ê³„"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ?œë¦°ì§€ ?Œí”„ ?íƒœë¨¸ì‹ ", 3),
        create_code_block(titrator_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=titrator_blocks)
    print("  ??L-Titrator ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 3. Psi - ?•ë? ê°€???œì–´ =====
    print("\n?“ Psi ì¶”ê? ì¤?..")
    psi_code = '''// PrsCtrl.c - ?•ë? ê°€???•ë ¥ PID ?œì–´
// ì§„ê³µ ê²Œì´ì§€ ëª¨ë‹ˆ?°ë§ + 250ms ì£¼ê¸° ê³ ì† ?¼ë“œë°?

/* ?•ë ¥ ?œì–´ ë³€??*/
INTERNAL float32_t sglPrsCtrlFilteredPressureP1;   // ?…ë ¥ ?•ë ¥ (torr)
INTERNAL float32_t sglPrsCtrlControlPeriod = 0.25f; // 250ms ?œì–´ ì£¼ê¸°

/* PID ?•ë ¥ ?œì–´ ?Œê³ ë¦¬ì¦˜ */
void pid_pressure_compute(PID_STR *pPID) {
    float error = pPID->sp - pPID->pv;  // ëª©í‘œ?•ë ¥ - ?„ì¬?•ë ¥
    
    // P (ë¹„ë??? - ì¦‰ê°?ì¸ ?¤ì°¨ ë°˜ì‘
    pPID->p_term = pPID->kp * error;
    
    // I (?ë¶„?? - ?„ì  ?¤ì°¨ ë³´ì •
    pPID->i_term += pPID->ki * error;
    if (pPID->i_term > pPID->i_max) pPID->i_term = pPID->i_max;
    if (pPID->i_term < pPID->i_min) pPID->i_term = pPID->i_min;
    
    // D (ë¯¸ë¶„?? - ê¸‰ê²©??ë³€???µì œ
    pPID->d_term = pPID->kd * (error - pPID->prev_error);
    pPID->prev_error = error;
    
    // ìµœì¢… ì¶œë ¥ = ë°¸ë¸Œ PWM ?€??
    pPID->co = pPID->p_term + pPID->i_term + pPID->d_term;
}'''

    psi_blocks = [
        create_heading("?”¬ Psi-1000/3000 - ?•ë? ê°€???œì–´ ?œìŠ¤??, 2),
        create_callout("ì§„ê³µ ê²Œì´ì§€ ëª¨ë‹ˆ?°ë§ ê¸°ë°˜ ?•ë? ê°€???œì–´ ?œìŠ¤??, "?¯"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?•ë? ê°€???œì–´ ?œìŠ¤??(ë°˜ë„ì²??”ìŠ¤?Œë ˆ??ê³µì •??"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4, 168MHz)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Psi ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("ì§„ê³µ ê²Œì´ì§€ ?¸í„°?˜ì´?? Pirani/Capacitance Gauge ?°ë™"),
        create_bulleted_list("ê³ ì† PID ?¼ë“œë°??œì–´: 250ms ì£¼ê¸° ?•ë ¥ ?œì–´"),
        create_bulleted_list("FreeRTOS ë©€?°íƒœ?¤í‚¹: ?•ë ¥?œì–´, ?¼ì„œ, ?µì‹  ë³‘ë ¬ ì²˜ë¦¬"),
        create_bulleted_list("Ethernet (LwIP): TCP/UDP PC ?œì–´ ?¸í„°?˜ì´??),
        create_bulleted_list("PID Auto-Tuning: ìµœì  ?Œë¼ë¯¸í„° ?ë™ ?ìƒ‰"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ê°€???•ë ¥ PID ?œì–´", 3),
        create_code_block(psi_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=psi_blocks)
    print("  ??Psi ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 4. LPC - TCP ?µì‹  =====
    print("\n?“ LPC ì¶”ê? ì¤?..")
    lpc_code = '''// ethernet.c - TCP ?œë²„ ?µì‹  (LwIP netconn API)
// FreeRTOS ?œìŠ¤?¬ë¡œ TCP ?´ë¼?´ì–¸???°ê²° ê´€ë¦?

void TcpTask(void *arg) {
    struct netconn *conn, *newconn;
    err_t err, recv_err;
    
    // 1. TCP ?°ê²° ?ì„± ë°?ë°”ì¸??
    conn = netconn_new(NETCONN_TCP);
    err = netconn_bind(conn, NULL, TELNET_TCP_PORT);  // Port 23
    netconn_listen(conn);
    
    while (1) {
        task_cnt_tcp++;
        osDelay(1);
        
        // 2. ?´ë¼?´ì–¸???°ê²° ?˜ë½ (1ms ?€?„ì•„??
        conn->recv_timeout = 1;
        err = netconn_accept(conn, &newconn);
        
        if (err == ERR_OK) {
            bd.eth.tcp.connected = 1;
            netconn_getaddr(newconn, &bd.eth.tcp.remote_ip, &bd.eth.tcp.remote_port, 0);
            
            for (;;) {
                // 3. ?°ì´???˜ì‹  ì²˜ë¦¬
                newconn->recv_timeout = 1;
                recv_err = netconn_recv(newconn, &buf);
                
                if (recv_err == ERR_OK) {
                    netbuf_data(buf, &data, &len);
                    for (i = 0; i < len; i++)
                        debug_command_process(TPID_TCP, ((char *)data)[i]);
                    netbuf_delete(buf);
                } else if (recv_err == ERR_CLSD) {
                    netconn_close(newconn);
                    break;  // ?°ê²° ì¢…ë£Œ
                }
            }
        }
    }
}'''

    lpc_blocks = [
        create_heading("?”´ LPC - ?ˆì´?€ ?Œí‹°??ë¶„ì„ê¸?, 2),
        create_callout("?ˆì´?€(ê´‘í•™)ë¥??´ìš©???¬ëŸ¬ë¦??Œí‹°??ë¶„ì„ ?¥ë¹„", "?’¡"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ˆì´?€ ê´‘í•™ ê¸°ë°˜ ?¬ëŸ¬ë¦??Œí‹°??ì¹´ìš´??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[LPC ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?ˆì´?€ ê´‘í•™ê³? ?°ë?ê´?ê²€ì¶?ê¸°ë°˜ ?Œí‹°??ê³„ìˆ˜"),
        create_bulleted_list("TCP/IP ?µì‹  (LwIP): PC ?ê²© ?œì–´ ë°??°ì´???„ì†¡"),
        create_bulleted_list("FreeRTOS: ë©€?°íƒœ?¤í‚¹ ê¸°ë°˜ ?¤ì‹œê°?ì²˜ë¦¬"),
        create_bulleted_list("ê³ ì† ADC ?˜í”Œë§? ?¤ì‹œê°??Œí‹°??? í˜¸ ê²€ì¶?),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - TCP ?œë²„ ?µì‹ ", 3),
        create_code_block(lpc_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=lpc_blocks)
    print("  ??LPC ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 5. SSC - ?¼ì„œ ?°ì´??ì·¨ë“ =====
    print("\n?“ SSC ì¶”ê? ì¤?..")
    ssc_code = '''// SSC_main.c - UART ?˜ì‹  ?¸í„°?½íŠ¸ ë°???ì²˜ë¦¬
// Photo Diode Array ?¼ì„œ ?°ì´???˜ì§‘

/* UART ?˜ì‹  ?¸í„°?½íŠ¸ ?¸ë“¤??*/
void my_uart_rx_irq_handler(UART_HandleTypeDef *huart) {
    TERM_PORT   tpid;
    UART_STR*   pUART;
    uint32_t    isrflags = READ_REG(huart->Instance->ISR);
    
    // 1. UART ?ŒìŠ¤ ?ë³„
    if (huart->Instance == USART1) {
        tpid  = TPID_DEBUG;
        pUART = &uart[TPID_DEBUG];
    } else if (huart->Instance == USART2) {
        tpid  = TPID_RS485;
        pUART = &uart[TPID_RS485];
    } else {
        return;
    }
    
    // 2. ?˜ì‹  ?°ì´??ë°??ëŸ¬ ?Œë˜ê·??½ê¸°
    pUART->rx_err_flag = (isrflags & (USART_ISR_PE | USART_ISR_FE | USART_ISR_ORE | USART_ISR_NE));
    pUART->rx_data = (uint8_t)(huart->Instance->RDR & 0x00FF);
    
    // 3. ?˜ì‹  ?ì— ?€??
    rxQ_write(tpid, pUART->rx_data);
    
    // 4. ?ëŸ¬ ë°œìƒ ??ì¶œë ¥
    if (pUART->rx_err_flag)
        term_printf(TPID_DEBUG, "UART[%d] Rx Err: 0x%08x\\n", tpid, pUART->rx_err_flag);
}

/* ?€?´ë¨¸ ?¸í„°?½íŠ¸ (1ms/10ms ???ì„±) */
void HAL_IncTick(void) {
    uwTick += uwTickFreq;
    tick_1ms++;
    if ((tick_1ms % 10) == 0) tick_10ms++;
}'''

    ssc_blocks = [
        create_heading("?“Š SSC - Photo Diode Array ?¬ëŸ¬ë¦?ë¶„ì„ê¸?, 2),
        create_callout("Photo Diode Arrayë¥??¬ìš©??ê´‘í•™???¬ëŸ¬ë¦?ë¶„ì„ ?¥ë¹„", "?“ˆ"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: PDA ê¸°ë°˜ ê´‘í•™???¬ëŸ¬ë¦?ë¶„ì„ê¸?),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32G0 ?œë¦¬ì¦?),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[SSC ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("Photo Diode Array: ?¤ì±„???™ì‹œ ê´‘í•™ ì¸¡ì •"),
        create_bulleted_list("UART ?¸í„°?½íŠ¸ + ?? ?¼ì„œ ?°ì´???¤ì‹œê°??˜ì§‘"),
        create_bulleted_list("ë¶„ê´‘ ë¶„ì„: ?Œì¥ë³??¡ê´‘??ì¸¡ì •"),
        create_bulleted_list("RS485 ?µì‹ : ?¸ë? ?¥ë¹„ ?°ë™"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - UART ?¸í„°?½íŠ¸ ë°??°ì´???˜ì§‘", 3),
        create_code_block(ssc_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=ssc_blocks)
    print("  ??SSC ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 6. MS - Protocol ?µì‹  =====
    print("\n?“ MS ì¶”ê? ì¤?..")
    ms_code = '''// CS_610F_comm.c - Horiba CS-610F ?„ë¡œ? ì½œ êµ¬í˜„
// DRV Board?ì„œ ?¸ë? ?¥ë¹„?€ ?µì‹  ?„ë¡œ? ì½œ ì²˜ë¦¬

/* ?„ë¡œ? ì½œ ?‘ë‹µ ?ì„± - ì¸¡ì • ?°ì´???„ì†¡ */
// Command : R,DD[CR][LF]
// Response: DD,XXX,X,X,XXX.XX,XXXXXX,...[CR][LF]
void cmd_RDD_svc(void) {
    char resp[70];
    int  cnt = 0;
    int  resp_cnt = bd.cs610f.resp_cnt;
    
    if (++bd.cs610f.resp_cnt >= 1000)
        bd.cs610f.resp_cnt = 0;
    
    // ?‘ë‹µ ?„ë ˆ??êµ¬ì„±
    resp[cnt++] = 'D';
    resp[cnt++] = 'D';
    resp[cnt++] = ',';
    
    cnt += sprintf(&resp[cnt], "%03d", resp_cnt);           // No
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%01d", bd.cs610f.meas_stat); // Status
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%01d", bd.cs610f.chem_type); // Chemical type
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%06.2f", bd.concent_str.temp_1);    // Temperature
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%06.2f", bd.concent_str.concent1);  // Concentration 1
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%06.2f", bd.concent_str.concent2);  // Concentration 2
    // ... ?˜ë¨¸ì§€ ?„ë“œ
    
    // ?‘ë‹µ ?„ì†¡
    term_printf(TPID_RS232, "%s\\r\\n", resp);
}'''

    ms_blocks = [
        create_heading("?§ª MS (Aston) - ì§ˆëŸ‰ë¶„ì„ê¸?DRV Board", 2),
        create_callout("ì§ˆëŸ‰ë¶„ì„ê¸?Drive Board - ?Œí”„, ?¼ì„œ ???¥ë¹„ ?œì–´", "?™ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?˜ì§ˆë¶„ì„??ì§ˆëŸ‰ë¶„ì„ê¸?(Mass Spectrometer)"),
        create_bulleted_list("?´ë‹¹: DRV Board ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[MS DRV Board ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("DRV Board ?µì‹¬ ê¸°ëŠ¥", 3),
        create_bulleted_list("?¸ë? ?„ë¡œ? ì½œ ?µì‹ : Horiba CS-610F ?„ë¡œ? ì½œ êµ¬í˜„"),
        create_bulleted_list("?Œí”„ ?œì–´: ?œë£Œ/?œì•½ ì£¼ì… ?Œí”„ êµ¬ë™"),
        create_bulleted_list("?¼ì„œ ?¸í„°?˜ì´?? ?¤ì±„??ADC ?¼ì„œ ? í˜¸ ì¸¡ì •"),
        create_bulleted_list("ë¦´ë ˆ???œì–´: ?¸ë? ?¥ë¹„ On/Off ?œì–´"),
        create_bulleted_list("SDì¹´ë“œ ë¡œê¹…: FATFS ?°ì´???€??),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ?¸ë? ?¥ë¹„ ?„ë¡œ? ì½œ ?µì‹ ", 3),
        create_code_block(ms_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=ms_blocks)
    print("  ??MS ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 7. JIG Board - ?¸í„°?½íŠ¸/??=====
    print("\n?“ JIG Board ì¶”ê? ì¤?..")
    jig_code = '''// ATIK_JIG_main.c - UART ?¸í„°?½íŠ¸ ë°???ê¸°ë°˜ ?µì‹ 
// ?¤ì¤‘ UART ?¬íŠ¸ ?™ì‹œ ?¸ë“¤ë§?

/* ì§€???œí’ˆ ëª©ë¡ */
const char hw_model_str[16][30] = {
    "1.L-Titrator", "2.Nu-2000", "3.Sigma-3000/4000",
    "4.Psi-3000", "5.Psi-1000", "15.ATIK_JIG"
};

/* UART ?˜ì‹  ?¸í„°?½íŠ¸ ?¸ë“¤??- ??ê¸°ë°˜ ?°ì´??ê´€ë¦?*/
void my_uart_rx_irq_handler(UART_HandleTypeDef *huart) {
    TERM_PORT   tpid;
    UART_STR*   pUART;
    uint32_t    isrflags = READ_REG(huart->Instance->SR);
    
    // 1. UART ?¬íŠ¸ ?ë³„ (ë©€???¬íŠ¸ ì§€??
    if (huart->Instance == USART1)      { tpid = TPID_DEBUG; pUART = &uart[TPID_DEBUG]; }
    else if (huart->Instance == USART2) { tpid = TPID_PANEL; pUART = &uart[TPID_PANEL]; }
    else if (huart->Instance == USART3) { tpid = TPID_RS232; pUART = &uart[TPID_RS232]; }
    else if (huart->Instance == USART6) { tpid = TPID_RS485; pUART = &uart[TPID_RS485]; }
    else return;
    
    // 2. ?°ì´??ë°??ëŸ¬ ?½ê¸° (SR ??DR ?œì„œë¡??ëŸ¬ ?Œë˜ê·??´ë¦¬??
    pUART->rx_err_flag = (isrflags & (USART_SR_PE | USART_SR_FE | USART_SR_ORE | USART_SR_NE));
    pUART->rx_data = (uint8_t)(huart->Instance->DR & 0x00FF);
    
    // 3. ?˜ì‹  ?ì— ?€??(ë§?ë²„í¼)
    rxQ_write(tpid, pUART->rx_data);
    
    // 4. ?¤ë²„?????ëŸ¬ ë¡œê¹…
    if (pUART->rx_err_flag)
        term_printf(TPID_DEBUG, "UART[%d] Err: 0x%08x\\n", tpid, pUART->rx_err_flag);
}'''

    jig_blocks = [
        create_heading("?”§ ATIK JIG Board - ë³´ë“œ ?…ê³  ?ŒìŠ¤???¥ë¹„", 2),
        create_callout("ë³´ë“œ ?…ê³  ?ŒìŠ¤?¸ë? ?„í•œ ?ë™???ŒìŠ¤???¥ë¹„", "?› ï¸?),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ì‚°?¼ì¸ ë³´ë“œ ?…ê³  ?ŒìŠ¤?¸ìš© JIG"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[JIG Board ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?¤ì¤‘ UART ?¸í„°?½íŠ¸: 4ê°??¬íŠ¸ ?™ì‹œ ?¸ë“¤ë§?),
        create_bulleted_list("??ê¸°ë°˜ ?°ì´??ê´€ë¦? ë§?ë²„í¼ë¡??°ì´???ì‹¤ ë°©ì?"),
        create_bulleted_list("?¤ì¤‘ ?œí’ˆ ì§€?? L-Titrator, Nu-2000, Sigma, Psi ??),
        create_bulleted_list("?ë™???ŒìŠ¤?? ADC/DAC, ?µì‹ , GPIO ?ë™ ê²€ì¦?),
        create_bulleted_list("ë¦´ë ˆ??MUX ?œì–´: ?ŒìŠ¤???¬ì¸???ë™ ? íƒ"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ë©€??UART ?¸í„°?½íŠ¸ ë°???, 3),
        create_code_block(jig_code, "c"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=jig_blocks)
    print("  ??JIG Board ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== 8. Sigma - ?Œë¡œ?¤ê³„ë§?=====
    print("\n?“ Sigma ì¶”ê? ì¤?..")
    sigma_blocks = [
        create_heading("?—ï¸ Sigma-1000/3000/4000 - COD ë¶„ì„ ?œìŠ¤??, 2),
        create_callout("COD/BOD/TOC ?˜ì§ˆë¶„ì„ ?„ìš© ?¥ë¹„", "?“Š"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: COD(?”í•™???°ì†Œ?”êµ¬?? ?ë™ë¶„ì„ ?œìŠ¤??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ (Hardware Only)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Sigma ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?Œë¡œ ?¤ê³„ ì£¼ìš” ?¬í•­", 3),
        create_bulleted_list("UV-IR LED ?œë¼?´ë²„ ?Œë¡œ: ?•ì „ë¥?êµ¬ë™"),
        create_bulleted_list("Photo Diode ? í˜¸ ì¦í­: ?€?¡ìŒ ?¸ëœ?¤ì„?¼ë˜???°í”„"),
        create_bulleted_list("?¤ì±„??ADC ?¸í„°?˜ì´?? AD7682 16ë¹„íŠ¸ ADC"),
        create_bulleted_list("?œë¦°ì§€ ?Œí”„ ?œë¼?´ë²„: ?¤í…Œ??ëª¨í„° êµ¬ë™ ?Œë¡œ"),
        create_bulleted_list("?¨ë„ ?œì–´ ?Œë¡œ: ?ˆí„°/ì¿¨ëŸ¬ PWM ?œë¼?´ë²„"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=sigma_blocks)
    print("  ??Sigma ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== ê¸°ìˆ  ?¤íƒ + ?°ë½ì²?=====
    print("\n?“ ê¸°ìˆ  ?¤íƒ/?°ë½ì²?ì¶”ê? ì¤?..")
    final_blocks = [
        create_heading("?› ï¸?ê¸°ìˆ  ?¤íƒ", 2),
        create_heading("?„ë¡œê·¸ë˜ë°??¸ì–´", 3),
        create_bulleted_list("C/C++ (Embedded), Python, Verilog"),
        create_heading("MCU/?„ë¡œ?¸ì„œ", 3),
        create_bulleted_list("STM32 ?œë¦¬ì¦?(F1, F4, G0, H7), Renesas RA6M4"),
        create_bulleted_list("Xilinx Zynq (FPGA + ARM)"),
        create_heading("ê°œë°œ ?„êµ¬", 3),
        create_bulleted_list("IDE: STM32CubeIDE, VS Code, Keil"),
        create_bulleted_list("EDA: Altium Designer (?Œë¡œ?¤ê³„/PCB)"),
        create_bulleted_list("ë²„ì „ê´€ë¦? Git, GitHub"),
        create_divider(),
        create_heading("?“ ?°ë½ì²?, 2),
        create_bulleted_list("?“§ Email: gari210@naver.com"),
        create_bulleted_list("?™ GitHub: github.com/gari210404"),
        create_paragraph(""),
        create_callout("Thank you for reviewing my portfolio!", "?™"),
    ]
    notion.blocks.children.append(block_id=page_id, children=final_blocks)
    print("  ???„ë£Œ")

    print(f"\n{'='*50}")
    print("?‰ ?¸ì…˜ ?¬íŠ¸?´ë¦¬??v5 ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print(f"{'='*50}")
    return page_id

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Notion Portfolio v6 - ? ê?(?œë) ?•ì‹?¼ë¡œ ?„ë¡œ?íŠ¸ ?‘ê¸°/?¼ì¹˜ê¸?ê°€??
"""

from notion_client import Client
import time

NOTION_TOKEN = "YOUR_NOTION_TOKEN_HERE"
PARENT_PAGE_ID = "2e8f0746-9821-809c-9331-e34ae2d5e03e"
notion = Client(auth=NOTION_TOKEN)

def create_heading(text, level=2):
    heading_type = f"heading_{level}"
    return {"object": "block", "type": heading_type, heading_type: {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_paragraph(text, bold=False):
    return {"object": "block", "type": "paragraph", "paragraph": {"rich_text": [{"type": "text", "text": {"content": text}, "annotations": {"bold": bold}}] if text else []}}

def create_code_block(code, language="c"):
    return {"object": "block", "type": "code", "code": {"rich_text": [{"type": "text", "text": {"content": code}}], "language": language}}

def create_bulleted_list(text):
    return {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {"rich_text": [{"type": "text", "text": {"content": text}}]}}

def create_divider():
    return {"object": "block", "type": "divider", "divider": {}}

def create_callout(text, emoji="?’¡"):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": text}}], "icon": {"type": "emoji", "emoji": emoji}}}

def create_image_placeholder(caption):
    return {"object": "block", "type": "callout", "callout": {"rich_text": [{"type": "text", "text": {"content": caption}}], "icon": {"type": "emoji", "emoji": "?“Š"}, "color": "gray_background"}}

def create_toggle(title, emoji="?“"):
    """? ê?(?œë) ë¸”ë¡ ?ì„± - ?ì‹ ë¸”ë¡?€ ?˜ì¤‘??ì¶”ê?"""
    return {
        "object": "block",
        "type": "toggle",
        "toggle": {
            "rich_text": [
                {"type": "text", "text": {"content": f"{emoji} {title}"}, "annotations": {"bold": True}}
            ],
            "children": []  # ?ì‹ ë¸”ë¡?€ appendë¡?ì¶”ê?
        }
    }

def add_children_to_toggle(page_id, toggle_block_id, children):
    """? ê? ë¸”ë¡???ì‹ ë¸”ë¡ ì¶”ê?"""
    notion.blocks.children.append(block_id=toggle_block_id, children=children)

def main():
    print("=== Notion Portfolio v6 (? ê? ?•ì‹) ?…ë°?´íŠ¸ ?œì‘ ===\n")
    
    new_page = notion.pages.create(
        parent={"page_id": PARENT_PAGE_ID},
        properties={"title": {"title": [{"type": "text", "text": {"content": "ì¡°ë½???¬íŠ¸?´ë¦¬??v6 - Toggle"}}]}},
        icon={"type": "emoji", "emoji": "??"},
        cover={"type": "external", "external": {"url": "https://images.unsplash.com/photo-1518770660439-4636190af475?w=1200"}}
    )
    page_id = new_page["id"]
    print(f"?????˜ì´ì§€ ?ì„±: {page_id}")
    
    # ===== ?„ë¡œ??=====
    blocks = [
        create_heading("?‘¤ ì¡°ë½??(Rakhyun Cho)", 1),
        create_callout("Senior Embedded Engineer | 9?„ì°¨ | HW/FW Full-Stack ê°œë°œ??, "??),
        create_paragraph(""),
        create_heading("?“‹ ?µì‹¬ ??Ÿ‰", 2),
        create_bulleted_list("?„ë² ?”ë“œ ?œìŠ¤?? STM32 (F1, F4, H7), ARM Cortex-M ê¸°ë°˜ ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("?˜ë“œ?¨ì–´ ?¤ê³„: Altium Designer ê¸°ë°˜ ?Œë¡œ ?¤ê³„ ë°?PCB Layout"),
        create_bulleted_list("?¤ì‹œê°??œì–´: FreeRTOS, PID ?œì–´, ëª¨í„° ?œì–´ (Stepper, BLDC)"),
        create_bulleted_list("ê´‘í•™ ?œìŠ¤?? UV/IR LED, Photo Diode Array, ?ˆì´?€ ê´‘í•™ê³?),
        create_bulleted_list("?µì‹  ?„ë¡œ? ì½œ: UART, SPI, I2C, Ethernet (LwIP), TCP/UDP"),
        create_divider(),
    ]
    notion.blocks.children.append(block_id=page_id, children=blocks)
    print("???„ë¡œ???¹ì…˜ ì¶”ê?")
    time.sleep(0.5)

    # ===== 1. Nu-2000 - ? ê?ë¡?ë³€ê²?=====
    print("\n?“ Nu-2000 ? ê? ì¶”ê? ì¤?..")
    nu2000_code = '''// ADS1115.c - I2C ADC ?¼ì„œ ?°ì´???˜ì§‘ ë°??´ë™?‰ê·  ê³„ì‚°
// UV/IR Photo Diode ? í˜¸ë¥?16ë¹„íŠ¸ ADCë¡?ì¸¡ì •

/* I2Cë¥??µí•œ ADC ?°ì´???½ê¸° */
int ads1115_read(uint16_t addr, uint8_t *pdata) {
    uint16_t i2c_addr = (addr | 0x01);  // Read operation
    int result = HAL_I2C_Master_Receive(&hi2c3, i2c_addr, pdata, 2, 10);
    ads1115_i2c_err_check(addr, result);
    return result;
}

/* ?´ë™?‰ê·  ê³„ì‚° - ?¸ì´ì¦??œê±° ë°??ˆì •??ì¸¡ì • */
void adc_pd_svc_beta(uint8_t ch, int sig_dark, uint16_t val) {
    uint32_t sum;
    int i, cnt;
    
    // ë²„í¼ ?¬ì¸??ì¦ê? (?œí™˜ ë²„í¼)
    cnt = pPD_STR->ma_cnt;
    if (++cnt >= ((bd.env.working[FILTER] & 0x00FFFFFF) * 5))
        cnt = 0;
    pPD_STR->ma_cnt = cnt;
    
    // ADC ê°??€??(?‘ìˆ˜ë§??ˆìš©)
    pPD_STR->raw[cnt] = val;
    if (val < 0x8000) {
        pPD_STR->raw_pos[cnt] = val;
        pPD_STR->vtg_raw = (float)val * 0.000125;  // 125uV/step
    } else {
        pPD_STR->raw_pos[cnt] = 0;
    }
    
    // ?´ë™?‰ê·  ê³„ì‚° (Filter * 5???˜í”Œ)
    sum = 0;
    for (i = 0; i < ((bd.env.working[FILTER] & 0x00FFFFFF) * 5); i++) 
        sum += pPD_STR->raw_pos[i];
    
    pPD_STR->raw_mavg = (uint16_t)((float)sum / (float)((bd.env.working[FILTER] & 0x00FFFFFF) * 5));
    pPD_STR->vtg_mavg = (float)pPD_STR->raw_mavg * 0.000125;  // ?„ì•• ë³€??
}'''

    # ? ê? ë¸”ë¡ ?ì„±
    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("Nu-2000 (Lux) - UV+IR ?©ì•¡ ?ë„ ë¶„ì„ê¸?, "?’§")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    # ? ê? ?´ë? ì»¨í…ì¸?ì¶”ê?
    nu2000_content = [
        create_callout("UV+IR ê´‘í•™???´ìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„", "?”¬"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: UV+IR(ê´‘í•™)ë¥??´ìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4, 168MHz)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Nu-2000 ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("UV/IR LED ê´‘í•™ ì¸¡ì •: ?¡ê´‘??ê¸°ë°˜ ?ë„ ë¶„ì„"),
        create_bulleted_list("ADS1115 16ë¹„íŠ¸ ADC: I2C ?¸í„°?˜ì´?? Delta-Sigma ë°©ì‹"),
        create_bulleted_list("?´ë™?‰ê·  ?„í„°: ?¸ì´ì¦??œê±° ë°?ì¸¡ì • ?ˆì •??),
        create_bulleted_list("?œë¦°ì§€ ?Œí”„: A3977 ?¤í…Œ??ëª¨í„° ë§ˆì´?¬ë¡œ?¤í… ?œì–´"),
        create_bulleted_list("Nextion HMI LCD: ?°ì¹˜?¤í¬ë¦?UI"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - I2C ?¼ì„œ ?°ì´???´ë™?‰ê· ", 3),
        create_code_block(nu2000_code, "c"),
    ]
    add_children_to_toggle(page_id, toggle_id, nu2000_content)
    print("  ??Nu-2000 ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    # ===== êµ¬ë¶„??=====
    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== 2. L-Titrator - ? ê? =====
    print("\n?“ L-Titrator ? ê? ì¶”ê? ì¤?..")
    titrator_code = '''// syringe_pump_ctrl.c - Hamilton ?œë¦°ì§€ ?Œí”„ ?íƒœë¨¸ì‹  ?œì–´
// RS485 ?µì‹ ?¼ë¡œ ?œë¦°ì§€ ?Œí”„ ëª…ë ¹/?‘ë‹µ ì²˜ë¦¬

/* ?œë¦°ì§€ ?Œí”„ ?íƒœ ë¨¸ì‹  */
void syr_pump_st10_svc(void) {
    int data_pos;
    
    switch (syr_pump[syr_pump_ch].st) {
        case 0x10:  // ?œë¦°ì§€ ?íƒœ ì¡°íšŒ
            syr_pump_query_cmd(QUERY_SRG_STATUS);
            syr_pump[syr_pump_ch].st++;
            syr_pump[syr_pump_ch].resp_cnt = 0;
            break;
            
        case 0x11:  // ?‘ë‹µ ?€ê¸?ë°?ì²˜ë¦¬
            if (++syr_pump[syr_pump_ch].resp_cnt >= SRG_PUMP_RESP_TIMEOUT) {
                // ?€?„ì•„??- ì´ˆê¸° ?íƒœë¡?ë³µê?
                syr_pump[syr_pump_ch].st = SRG_PUMP_ST00_DETECT;
                trace_printf(TPID_DEBUG, "syr_pump[%d] : no response\\n", syr_pump_ch);
            } else {
                if (syr_pump_rx_flag == 1) {
                    syr_pump_rx_flag = 0;
                    data_pos = is_Q_ack_packet(uart[TPID_RS485].command, uart[TPID_RS485].cmd_index);
                    
                    if (data_pos > 0) {
                        syr_pump[syr_pump_ch].syr_status = syr_pump[syr_pump_ch].temp;
                        if (syr_pump[syr_pump_ch].syr_status == 0) {
                            trace_printf(TPID_DEBUG, "syr_pump[%d] : initialized\\n", syr_pump_ch);
                            syr_pump[syr_pump_ch].st++;
                        }
                    }
                }
            }
            break;
    }
}'''

    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("L-Titrator - ?ì •???©ì•¡ ?ë„ ë¶„ì„ê¸?, "?§ª")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    titrator_content = [
        create_callout("?ì •(Titration) ë°©ì‹???œìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„", "?—ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ì • ë°©ì‹???œìš©???©ì•¡ ?ë„ ë¶„ì„ ?¥ë¹„"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[L-Titrator ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("Hamilton ?œë¦°ì§€ ?Œí”„: RS485 ?„ë¡œ? ì½œ ?œì–´"),
        create_bulleted_list("?íƒœë¨¸ì‹  ê¸°ë°˜ ?œì–´: ?ˆì •?ì¸ ?Œí”„ ?™ì‘ ê´€ë¦?),
        create_bulleted_list("pH/?„ë„???¼ì„œ: ?ì • ì¢…ë§???ë™ ê²€ì¶?),
        create_bulleted_list("ë°°í„°ë¦?ê´€ë¦? ?´ë???ê¸°ê¸° ?€?„ë ¥ ?¤ê³„"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ?œë¦°ì§€ ?Œí”„ ?íƒœë¨¸ì‹ ", 3),
        create_code_block(titrator_code, "c"),
    ]
    add_children_to_toggle(page_id, toggle_id, titrator_content)
    print("  ??L-Titrator ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== 3. Psi - ? ê? =====
    print("\n?“ Psi ? ê? ì¶”ê? ì¤?..")
    psi_code = '''// PrsCtrl.c - ?•ë? ê°€???•ë ¥ PID ?œì–´
// ì§„ê³µ ê²Œì´ì§€ ëª¨ë‹ˆ?°ë§ + 250ms ì£¼ê¸° ê³ ì† ?¼ë“œë°?

/* ?•ë ¥ ?œì–´ ë³€??*/
INTERNAL float32_t sglPrsCtrlFilteredPressureP1;   // ?…ë ¥ ?•ë ¥ (torr)
INTERNAL float32_t sglPrsCtrlControlPeriod = 0.25f; // 250ms ?œì–´ ì£¼ê¸°

/* PID ?•ë ¥ ?œì–´ ?Œê³ ë¦¬ì¦˜ */
void pid_pressure_compute(PID_STR *pPID) {
    float error = pPID->sp - pPID->pv;  // ëª©í‘œ?•ë ¥ - ?„ì¬?•ë ¥
    
    // P (ë¹„ë??? - ì¦‰ê°?ì¸ ?¤ì°¨ ë°˜ì‘
    pPID->p_term = pPID->kp * error;
    
    // I (?ë¶„?? - ?„ì  ?¤ì°¨ ë³´ì •
    pPID->i_term += pPID->ki * error;
    if (pPID->i_term > pPID->i_max) pPID->i_term = pPID->i_max;
    if (pPID->i_term < pPID->i_min) pPID->i_term = pPID->i_min;
    
    // D (ë¯¸ë¶„?? - ê¸‰ê²©??ë³€???µì œ
    pPID->d_term = pPID->kd * (error - pPID->prev_error);
    pPID->prev_error = error;
    
    // ìµœì¢… ì¶œë ¥ = ë°¸ë¸Œ PWM ?€??
    pPID->co = pPID->p_term + pPID->i_term + pPID->d_term;
}'''

    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("Psi-1000/3000 - ?•ë? ê°€???œì–´ ?œìŠ¤??, "?”¬")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    psi_content = [
        create_callout("ì§„ê³µ ê²Œì´ì§€ ëª¨ë‹ˆ?°ë§ ê¸°ë°˜ ?•ë? ê°€???œì–´ ?œìŠ¤??, "?¯"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?•ë? ê°€???œì–´ ?œìŠ¤??(ë°˜ë„ì²??”ìŠ¤?Œë ˆ??ê³µì •??"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4, 168MHz)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Psi ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("ì§„ê³µ ê²Œì´ì§€ ?¸í„°?˜ì´?? Pirani/Capacitance Gauge ?°ë™"),
        create_bulleted_list("ê³ ì† PID ?¼ë“œë°??œì–´: 250ms ì£¼ê¸° ?•ë ¥ ?œì–´"),
        create_bulleted_list("FreeRTOS ë©€?°íƒœ?¤í‚¹: ?•ë ¥?œì–´, ?¼ì„œ, ?µì‹  ë³‘ë ¬ ì²˜ë¦¬"),
        create_bulleted_list("Ethernet (LwIP): TCP/UDP PC ?œì–´ ?¸í„°?˜ì´??),
        create_bulleted_list("PID Auto-Tuning: ìµœì  ?Œë¼ë¯¸í„° ?ë™ ?ìƒ‰"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ê°€???•ë ¥ PID ?œì–´", 3),
        create_code_block(psi_code, "c"),
    ]
    add_children_to_toggle(page_id, toggle_id, psi_content)
    print("  ??Psi ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== 4. LPC - ? ê? =====
    print("\n?“ LPC ? ê? ì¶”ê? ì¤?..")
    lpc_code = '''// ethernet.c - TCP ?œë²„ ?µì‹  (LwIP netconn API)
// FreeRTOS ?œìŠ¤?¬ë¡œ TCP ?´ë¼?´ì–¸???°ê²° ê´€ë¦?

void TcpTask(void *arg) {
    struct netconn *conn, *newconn;
    err_t err, recv_err;
    
    // 1. TCP ?°ê²° ?ì„± ë°?ë°”ì¸??
    conn = netconn_new(NETCONN_TCP);
    err = netconn_bind(conn, NULL, TELNET_TCP_PORT);  // Port 23
    netconn_listen(conn);
    
    while (1) {
        task_cnt_tcp++;
        osDelay(1);
        
        // 2. ?´ë¼?´ì–¸???°ê²° ?˜ë½ (1ms ?€?„ì•„??
        conn->recv_timeout = 1;
        err = netconn_accept(conn, &newconn);
        
        if (err == ERR_OK) {
            bd.eth.tcp.connected = 1;
            netconn_getaddr(newconn, &bd.eth.tcp.remote_ip, &bd.eth.tcp.remote_port, 0);
            
            for (;;) {
                // 3. ?°ì´???˜ì‹  ì²˜ë¦¬
                newconn->recv_timeout = 1;
                recv_err = netconn_recv(newconn, &buf);
                
                if (recv_err == ERR_OK) {
                    netbuf_data(buf, &data, &len);
                    for (i = 0; i < len; i++)
                        debug_command_process(TPID_TCP, ((char *)data)[i]);
                    netbuf_delete(buf);
                } else if (recv_err == ERR_CLSD) {
                    netconn_close(newconn);
                    break;  // ?°ê²° ì¢…ë£Œ
                }
            }
        }
    }
}'''

    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("LPC - ?ˆì´?€ ?Œí‹°??ë¶„ì„ê¸?, "?”´")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    lpc_content = [
        create_callout("?ˆì´?€(ê´‘í•™)ë¥??´ìš©???¬ëŸ¬ë¦??Œí‹°??ë¶„ì„ ?¥ë¹„", "?’¡"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ˆì´?€ ê´‘í•™ ê¸°ë°˜ ?¬ëŸ¬ë¦??Œí‹°??ì¹´ìš´??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[LPC ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?ˆì´?€ ê´‘í•™ê³? ?°ë?ê´?ê²€ì¶?ê¸°ë°˜ ?Œí‹°??ê³„ìˆ˜"),
        create_bulleted_list("TCP/IP ?µì‹  (LwIP): PC ?ê²© ?œì–´ ë°??°ì´???„ì†¡"),
        create_bulleted_list("FreeRTOS: ë©€?°íƒœ?¤í‚¹ ê¸°ë°˜ ?¤ì‹œê°?ì²˜ë¦¬"),
        create_bulleted_list("ê³ ì† ADC ?˜í”Œë§? ?¤ì‹œê°??Œí‹°??? í˜¸ ê²€ì¶?),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - TCP ?œë²„ ?µì‹ ", 3),
        create_code_block(lpc_code, "c"),
    ]
    add_children_to_toggle(page_id, toggle_id, lpc_content)
    print("  ??LPC ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== 5. SSC - ? ê? =====
    print("\n?“ SSC ? ê? ì¶”ê? ì¤?..")
    ssc_code = '''// SSC_main.c - UART ?˜ì‹  ?¸í„°?½íŠ¸ ë°???ì²˜ë¦¬
// Photo Diode Array ?¼ì„œ ?°ì´???˜ì§‘

/* UART ?˜ì‹  ?¸í„°?½íŠ¸ ?¸ë“¤??*/
void my_uart_rx_irq_handler(UART_HandleTypeDef *huart) {
    TERM_PORT   tpid;
    UART_STR*   pUART;
    uint32_t    isrflags = READ_REG(huart->Instance->ISR);
    
    // 1. UART ?ŒìŠ¤ ?ë³„
    if (huart->Instance == USART1) {
        tpid  = TPID_DEBUG;
        pUART = &uart[TPID_DEBUG];
    } else if (huart->Instance == USART2) {
        tpid  = TPID_RS485;
        pUART = &uart[TPID_RS485];
    } else {
        return;
    }
    
    // 2. ?˜ì‹  ?°ì´??ë°??ëŸ¬ ?Œë˜ê·??½ê¸°
    pUART->rx_err_flag = (isrflags & (USART_ISR_PE | USART_ISR_FE | USART_ISR_ORE | USART_ISR_NE));
    pUART->rx_data = (uint8_t)(huart->Instance->RDR & 0x00FF);
    
    // 3. ?˜ì‹  ?ì— ?€??
    rxQ_write(tpid, pUART->rx_data);
    
    // 4. ?ëŸ¬ ë°œìƒ ??ì¶œë ¥
    if (pUART->rx_err_flag)
        term_printf(TPID_DEBUG, "UART[%d] Rx Err: 0x%08x\\n", tpid, pUART->rx_err_flag);
}

/* ?€?´ë¨¸ ?¸í„°?½íŠ¸ (1ms/10ms ???ì„±) */
void HAL_IncTick(void) {
    uwTick += uwTickFreq;
    tick_1ms++;
    if ((tick_1ms % 10) == 0) tick_10ms++;
}'''

    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("SSC - Photo Diode Array ?¬ëŸ¬ë¦?ë¶„ì„ê¸?, "?“Š")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    ssc_content = [
        create_callout("Photo Diode Arrayë¥??¬ìš©??ê´‘í•™???¬ëŸ¬ë¦?ë¶„ì„ ?¥ë¹„", "?“ˆ"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: PDA ê¸°ë°˜ ê´‘í•™???¬ëŸ¬ë¦?ë¶„ì„ê¸?),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32G0 ?œë¦¬ì¦?),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[SSC ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("Photo Diode Array: ?¤ì±„???™ì‹œ ê´‘í•™ ì¸¡ì •"),
        create_bulleted_list("UART ?¸í„°?½íŠ¸ + ?? ?¼ì„œ ?°ì´???¤ì‹œê°??˜ì§‘"),
        create_bulleted_list("ë¶„ê´‘ ë¶„ì„: ?Œì¥ë³??¡ê´‘??ì¸¡ì •"),
        create_bulleted_list("RS485 ?µì‹ : ?¸ë? ?¥ë¹„ ?°ë™"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - UART ?¸í„°?½íŠ¸ ë°??°ì´???˜ì§‘", 3),
        create_code_block(ssc_code, "c"),
    ]
    add_children_to_toggle(page_id, toggle_id, ssc_content)
    print("  ??SSC ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== 6. MS - ? ê? =====
    print("\n?“ MS ? ê? ì¶”ê? ì¤?..")
    ms_code = '''// CS_610F_comm.c - Horiba CS-610F ?„ë¡œ? ì½œ êµ¬í˜„
// DRV Board?ì„œ ?¸ë? ?¥ë¹„?€ ?µì‹  ?„ë¡œ? ì½œ ì²˜ë¦¬

/* ?„ë¡œ? ì½œ ?‘ë‹µ ?ì„± - ì¸¡ì • ?°ì´???„ì†¡ */
// Command : R,DD[CR][LF]
// Response: DD,XXX,X,X,XXX.XX,XXXXXX,...[CR][LF]
void cmd_RDD_svc(void) {
    char resp[70];
    int  cnt = 0;
    int  resp_cnt = bd.cs610f.resp_cnt;
    
    if (++bd.cs610f.resp_cnt >= 1000)
        bd.cs610f.resp_cnt = 0;
    
    // ?‘ë‹µ ?„ë ˆ??êµ¬ì„±
    resp[cnt++] = 'D';
    resp[cnt++] = 'D';
    resp[cnt++] = ',';
    
    cnt += sprintf(&resp[cnt], "%03d", resp_cnt);           // No
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%01d", bd.cs610f.meas_stat); // Status
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%01d", bd.cs610f.chem_type); // Chemical type
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%06.2f", bd.concent_str.temp_1);    // Temperature
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%06.2f", bd.concent_str.concent1);  // Concentration 1
    resp[cnt++] = ',';
    cnt += sprintf(&resp[cnt], "%06.2f", bd.concent_str.concent2);  // Concentration 2
    // ... ?˜ë¨¸ì§€ ?„ë“œ
    
    // ?‘ë‹µ ?„ì†¡
    term_printf(TPID_RS232, "%s\\r\\n", resp);
}'''

    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("MS (Aston) - ì§ˆëŸ‰ë¶„ì„ê¸?DRV Board", "?§ª")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    ms_content = [
        create_callout("ì§ˆëŸ‰ë¶„ì„ê¸?Drive Board - ?Œí”„, ?¼ì„œ ???¥ë¹„ ?œì–´", "?™ï¸"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?˜ì§ˆë¶„ì„??ì§ˆëŸ‰ë¶„ì„ê¸?(Mass Spectrometer)"),
        create_bulleted_list("?´ë‹¹: DRV Board ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[MS DRV Board ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("DRV Board ?µì‹¬ ê¸°ëŠ¥", 3),
        create_bulleted_list("?¸ë? ?„ë¡œ? ì½œ ?µì‹ : Horiba CS-610F ?„ë¡œ? ì½œ êµ¬í˜„"),
        create_bulleted_list("?Œí”„ ?œì–´: ?œë£Œ/?œì•½ ì£¼ì… ?Œí”„ êµ¬ë™"),
        create_bulleted_list("?¼ì„œ ?¸í„°?˜ì´?? ?¤ì±„??ADC ?¼ì„œ ? í˜¸ ì¸¡ì •"),
        create_bulleted_list("ë¦´ë ˆ???œì–´: ?¸ë? ?¥ë¹„ On/Off ?œì–´"),
        create_bulleted_list("SDì¹´ë“œ ë¡œê¹…: FATFS ?°ì´???€??),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ?¸ë? ?¥ë¹„ ?„ë¡œ? ì½œ ?µì‹ ", 3),
        create_code_block(ms_code, "c"),
    ]
    add_children_to_toggle(page_id, toggle_id, ms_content)
    print("  ??MS ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== 7. JIG Board - ? ê? =====
    print("\n?“ JIG Board ? ê? ì¶”ê? ì¤?..")
    jig_code = '''// ATIK_JIG_main.c - UART ?¸í„°?½íŠ¸ ë°???ê¸°ë°˜ ?µì‹ 
// ?¤ì¤‘ UART ?¬íŠ¸ ?™ì‹œ ?¸ë“¤ë§?

/* ì§€???œí’ˆ ëª©ë¡ */
const char hw_model_str[16][30] = {
    "1.L-Titrator", "2.Nu-2000", "3.Sigma-3000/4000",
    "4.Psi-3000", "5.Psi-1000", "15.ATIK_JIG"
};

/* UART ?˜ì‹  ?¸í„°?½íŠ¸ ?¸ë“¤??- ??ê¸°ë°˜ ?°ì´??ê´€ë¦?*/
void my_uart_rx_irq_handler(UART_HandleTypeDef *huart) {
    TERM_PORT   tpid;
    UART_STR*   pUART;
    uint32_t    isrflags = READ_REG(huart->Instance->SR);
    
    // 1. UART ?¬íŠ¸ ?ë³„ (ë©€???¬íŠ¸ ì§€??
    if (huart->Instance == USART1)      { tpid = TPID_DEBUG; pUART = &uart[TPID_DEBUG]; }
    else if (huart->Instance == USART2) { tpid = TPID_PANEL; pUART = &uart[TPID_PANEL]; }
    else if (huart->Instance == USART3) { tpid = TPID_RS232; pUART = &uart[TPID_RS232]; }
    else if (huart->Instance == USART6) { tpid = TPID_RS485; pUART = &uart[TPID_RS485]; }
    else return;
    
    // 2. ?°ì´??ë°??ëŸ¬ ?½ê¸° (SR ??DR ?œì„œë¡??ëŸ¬ ?Œë˜ê·??´ë¦¬??
    pUART->rx_err_flag = (isrflags & (USART_SR_PE | USART_SR_FE | USART_SR_ORE | USART_SR_NE));
    pUART->rx_data = (uint8_t)(huart->Instance->DR & 0x00FF);
    
    // 3. ?˜ì‹  ?ì— ?€??(ë§?ë²„í¼)
    rxQ_write(tpid, pUART->rx_data);
    
    // 4. ?¤ë²„?????ëŸ¬ ë¡œê¹…
    if (pUART->rx_err_flag)
        term_printf(TPID_DEBUG, "UART[%d] Err: 0x%08x\\n", tpid, pUART->rx_err_flag);
}'''

    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("ATIK JIG Board - ë³´ë“œ ?…ê³  ?ŒìŠ¤???¥ë¹„", "?”§")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    jig_content = [
        create_callout("ë³´ë“œ ?…ê³  ?ŒìŠ¤?¸ë? ?„í•œ ?ë™???ŒìŠ¤???¥ë¹„", "?› ï¸?),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: ?ì‚°?¼ì¸ ë³´ë“œ ?…ê³  ?ŒìŠ¤?¸ìš© JIG"),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ + ?Œì›¨??ê°œë°œ (100%)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[JIG Board ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?µì‹¬ ê¸°ìˆ ", 3),
        create_bulleted_list("?¤ì¤‘ UART ?¸í„°?½íŠ¸: 4ê°??¬íŠ¸ ?™ì‹œ ?¸ë“¤ë§?),
        create_bulleted_list("??ê¸°ë°˜ ?°ì´??ê´€ë¦? ë§?ë²„í¼ë¡??°ì´???ì‹¤ ë°©ì?"),
        create_bulleted_list("?¤ì¤‘ ?œí’ˆ ì§€?? L-Titrator, Nu-2000, Sigma, Psi ??),
        create_bulleted_list("?ë™???ŒìŠ¤?? ADC/DAC, ?µì‹ , GPIO ?ë™ ê²€ì¦?),
        create_bulleted_list("ë¦´ë ˆ??MUX ?œì–´: ?ŒìŠ¤???¬ì¸???ë™ ? íƒ"),
        create_paragraph(""),
        create_heading("ì½”ë“œ ?˜í”Œ - ë©€??UART ?¸í„°?½íŠ¸ ë°???, 3),
        create_code_block(jig_code, "c"),
    ]
    add_children_to_toggle(page_id, toggle_id, jig_content)
    print("  ??JIG Board ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== 8. Sigma - ? ê? =====
    print("\n?“ Sigma ? ê? ì¶”ê? ì¤?..")
    toggle_response = notion.blocks.children.append(
        block_id=page_id,
        children=[create_toggle("Sigma-1000/3000/4000 - COD ë¶„ì„ ?œìŠ¤??, "?—ï¸")]
    )
    toggle_id = toggle_response["results"][0]["id"]
    
    sigma_content = [
        create_callout("COD/BOD/TOC ?˜ì§ˆë¶„ì„ ?„ìš© ?¥ë¹„", "?“Š"),
        create_paragraph(""),
        create_heading("?„ë¡œ?íŠ¸ ê°œìš”", 3),
        create_bulleted_list("?œí’ˆ: COD(?”í•™???°ì†Œ?”êµ¬?? ?ë™ë¶„ì„ ?œìŠ¤??),
        create_bulleted_list("??• : ?Œë¡œ ?¤ê³„ (Hardware Only)"),
        create_bulleted_list("MCU: STM32F407 (ARM Cortex-M4)"),
        create_paragraph(""),
        create_heading("?œìŠ¤??ë¸”ë¡ ?¤ì´?´ê·¸??, 3),
        create_image_placeholder("[Sigma ë¸”ë¡ ?¤ì´?´ê·¸???½ì…]"),
        create_paragraph(""),
        create_heading("?Œë¡œ ?¤ê³„ ì£¼ìš” ?¬í•­", 3),
        create_bulleted_list("UV-IR LED ?œë¼?´ë²„ ?Œë¡œ: ?•ì „ë¥?êµ¬ë™"),
        create_bulleted_list("Photo Diode ? í˜¸ ì¦í­: ?€?¡ìŒ ?¸ëœ?¤ì„?¼ë˜???°í”„"),
        create_bulleted_list("?¤ì±„??ADC ?¸í„°?˜ì´?? AD7682 16ë¹„íŠ¸ ADC"),
        create_bulleted_list("?œë¦°ì§€ ?Œí”„ ?œë¼?´ë²„: ?¤í…Œ??ëª¨í„° êµ¬ë™ ?Œë¡œ"),
        create_bulleted_list("?¨ë„ ?œì–´ ?Œë¡œ: ?ˆí„°/ì¿¨ëŸ¬ PWM ?œë¼?´ë²„"),
    ]
    add_children_to_toggle(page_id, toggle_id, sigma_content)
    print("  ??Sigma ? ê? ì¶”ê? ?„ë£Œ")
    time.sleep(0.5)

    notion.blocks.children.append(block_id=page_id, children=[create_divider()])

    # ===== ê¸°ìˆ  ?¤íƒ + ?°ë½ì²?=====
    print("\n?“ ê¸°ìˆ  ?¤íƒ/?°ë½ì²?ì¶”ê? ì¤?..")
    final_blocks = [
        create_heading("?› ï¸?ê¸°ìˆ  ?¤íƒ", 2),
        create_heading("?„ë¡œê·¸ë˜ë°??¸ì–´", 3),
        create_bulleted_list("C/C++ (Embedded), Python, Verilog"),
        create_heading("MCU/?„ë¡œ?¸ì„œ", 3),
        create_bulleted_list("STM32 ?œë¦¬ì¦?(F1, F4, G0, H7), Renesas RA6M4"),
        create_bulleted_list("Xilinx Zynq (FPGA + ARM)"),
        create_heading("ê°œë°œ ?„êµ¬", 3),
        create_bulleted_list("IDE: STM32CubeIDE, VS Code, Keil"),
        create_bulleted_list("EDA: Altium Designer (?Œë¡œ?¤ê³„/PCB)"),
        create_bulleted_list("ë²„ì „ê´€ë¦? Git, GitHub"),
        create_divider(),
        create_heading("?“ ?°ë½ì²?, 2),
        create_bulleted_list("?“§ Email: gari210@naver.com"),
        create_bulleted_list("?™ GitHub: github.com/gari210404"),
        create_paragraph(""),
        create_callout("Thank you for reviewing my portfolio!", "?™"),
    ]
    notion.blocks.children.append(block_id=page_id, children=final_blocks)
    print("  ???„ë£Œ")

    print(f"\n{'='*60}")
    print("?‰ ?¸ì…˜ ?¬íŠ¸?´ë¦¬??v6 (? ê? ?•ì‹) ?ì„± ?„ë£Œ!")
    print(f"?“ ?˜ì´ì§€ ID: {page_id}")
    print(f"?”— URL: https://notion.so/{page_id.replace('-', '')}")
    print(f"{'='*60}")
    return page_id

if __name__ == "__main__":
    main()


